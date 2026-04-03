# 교대근무간호사 근무표 자동 생성기
# 근무표 생성: OR-Tools CP-SAT 전용 (`schedule_cpsat.py` — cp_model.CpModel + model.Add).

from flask import Flask, render_template, request, send_file, redirect, url_for
from datetime import date, timedelta
import calendar as _calendar
import json
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

app = Flask(__name__)

# ── 기본 상수 ──────────────────────────────────────────────────────────────────
# 인원 수(num_nurses 등)는 항상 수간호사를 포함한 총원이다.
# 예: 간호사 11명 = 수간호사 1 + 일반간호사 10 → num_nurses == 11 (인덱스 0..10).
YEAR, MONTH, NUM_DAYS = 2026, 5, 31


def set_period(year: int, month: int):
    """연도·월을 변경할 때 전역 상수를 갱신합니다."""
    global YEAR, MONTH, NUM_DAYS
    YEAR = year
    MONTH = month
    NUM_DAYS = _calendar.monthrange(year, month)[1]
    # 말일 단독 N(예: 3-3-1·2-3-1의 끝 1): 허용 일자는 항상 NUM_DAYS(당월 마지막 날).
    # 31일로 끝나는 달은 NUM_DAYS==31이므로 31일에 단독 N 1개 적용이 가능하다.

# NO: 야간(N) 누적 20회마다 생기는 휴무(OF 성격). 발생일은 사람마다 다름(대략 3개월에 1회 수준).
#     자동 근무 생성 시 배정하지 않음 — 신청/근무표에서 수기 입력.
SHIFT_NAMES = ['A1', 'D', 'E', 'N', 'OF', 'EDU', '연', '공', '병', '경', 'OH', 'NO']
A1_S, D_S, E_S, N_S, OF_S, EDU_S, YUN_S, GONG_S, BYUNG_S, GYUNG_S, OH_S, NO_S = range(12)

# 연속근무 최대 5일 제한에 포함되는 시프트 (연차·병가·경조 등 휴가는 제외 — D/E/N/공/EDU만)
STREAK_WORK_SHIFTS = frozenset({'D', 'E', 'N', 'EDU', '공'})

# 화면 색상 (연한 파스텔 배경 — 눈 피로 완화, 글자는 진한 톤으로 대비 유지)
SHIFT_COLORS = {
    'A1': '#E3EEF9', 'D': '#FFF9C4', 'E': '#FFE8E0', 'N': '#E8EAF6',
    'OF': '#F5F5F5', 'EDU': '#E8F5E9', '연': '#FCE4EC', '공': '#F3E5F5',
    '병': '#FFEBEE', '경': '#E0F2F1', 'OH': '#FFF3E0',
    'NO': '#ECEFF1',   # N 20회 등 수기 휴무 (OF와 구분)
}
SHIFT_TEXT_COLORS = {
    'A1': '#1565C0', 'D': '#F57F17', 'E': '#E65100', 'N': '#283593',
    'OF': '#78909C', 'EDU': '#2E7D32', '연': '#AD1457', '공': '#6A1B9A',
    '병': '#C62828', '경': '#00695C', 'OH': '#E65100', 'NO': '#37474F',
}

# 마지막 생성 결과 임시 저장 (단일 사용자 로컬 용도)
_last_result = {}


# ── 유틸리티 함수 ──────────────────────────────────────────────────────────────
def get_nurse_names(num_nurses):
    return ['수간호사'] + [f'간호사{i}' for i in range(1, num_nurses)]


def get_april_days(holidays=()):
    holiday_set = set(holidays)
    weekday_names = ['월', '화', '수', '목', '금', '토', '일']
    days = []
    for d in range(1, NUM_DAYS + 1):
        dt = date(YEAR, MONTH, d)
        days.append({
            'day': d,
            'date': dt,
            'weekday': dt.weekday(),
            'weekday_name': weekday_names[dt.weekday()],
            'is_weekend': dt.weekday() >= 5,
            'is_holiday': d in holiday_set,
        })
    return days


WEEKLY_REST_SHIFTS = frozenset({'OF', 'OH', 'NO', '연', '공', '병', '경'})

# CP·검증 공통: 엔진이 임의 배정하면 안 되는 시프트(반드시 신청 칸에만)
REQUEST_ONLY_SHIFTS = frozenset({'경', '공', 'EDU', '연', '병', 'NO'})


def _week_sunday(dt: date) -> date:
    """해당 날짜가 속한 주의 일요일(일~토 주간)."""
    return dt - timedelta(days=(dt.weekday() + 1) % 7)


def _carry_week_prev_month_off_counts(
    carry: dict, n: int, week_start: date, month_first: date,
) -> tuple[int, int, int, int]:
    """
    일~토 한 주의 시작일 week_start(일요일), month_first(당월 1일) **이전**인 날만 carry에서 읽어
    OF/OH/NO 개수와 그 일수를 반환.
    """
    pre_of = pre_oh = pre_no = 0
    n_prev = 0
    for i in range(7):
        d = week_start + timedelta(days=i)
        if d >= month_first:
            break
        n_prev += 1
        k = (month_first - d).days
        c = carry.get(n) or ()
        if not (1 <= k <= len(c)):
            continue
        s = c[-k]
        if s == 'OF':
            pre_of += 1
        elif s == 'OH':
            pre_oh += 1
        elif s == 'NO':
            pre_no += 1
    return pre_of, pre_oh, pre_no, n_prev


def _carry_week_prev_rest_total(
    carry: dict, n: int, week_start: date, month_first: date,
) -> tuple[int, int]:
    """전월 쪽 동일 주: 휴무(OF/OH/NO/연/공/병/경)일 수·해당 일수."""
    cnt = 0
    n_prev = 0
    for i in range(7):
        d = week_start + timedelta(days=i)
        if d >= month_first:
            break
        n_prev += 1
        k = (month_first - d).days
        c = carry.get(n) or ()
        if 1 <= k <= len(c) and c[-k] in WEEKLY_REST_SHIFTS:
            cnt += 1
    return cnt, n_prev


def _carry_week_next_month_off_counts(
    carry_next: dict, n: int, week_start: date, month_last: date,
) -> tuple[int, int, int, int]:
    """
    같은 일~토 주 중 month_last(당월 말일) 이후(차월) OF/OH/NO·일수 n_next.
    """
    post_of = post_oh = post_no = 0
    n_next = 0
    next_first = month_last + timedelta(days=1)
    for i in range(7):
        d = week_start + timedelta(days=i)
        if d <= month_last:
            continue
        n_next += 1
        if not carry_next:
            continue
        seq = carry_next.get(n)
        day_next = (d - next_first).days + 1
        s = None
        if isinstance(seq, dict):
            s = seq.get(day_next) or seq.get(str(day_next))
        elif isinstance(seq, (list, tuple)) and 1 <= day_next <= len(seq):
            s = seq[day_next - 1]
        if s == 'OF':
            post_of += 1
        elif s == 'OH':
            post_oh += 1
        elif s == 'NO':
            post_no += 1
    return post_of, post_oh, post_no, n_next


def _carry_prev_week_tail_complete(carry: dict, n: int, n_prev: int) -> bool:
    """일~토 주에서 당월 이전에 포함된 일수 n_prev만큼 carry 꼬리가 있는지."""
    if n_prev <= 0:
        return True
    c = carry.get(n) or ()
    return len(c) >= n_prev


def _carry_week_next_rest_total(
    carry_next: dict, n: int, week_start: date, month_last: date,
) -> tuple[int, int]:
    post_cnt = 0
    n_next = 0
    next_first = month_last + timedelta(days=1)
    for i in range(7):
        d = week_start + timedelta(days=i)
        if d <= month_last:
            continue
        n_next += 1
        if not carry_next:
            continue
        seq = carry_next.get(n)
        day_next = (d - next_first).days + 1
        s = None
        if isinstance(seq, dict):
            s = seq.get(day_next) or seq.get(str(day_next))
        elif isinstance(seq, (list, tuple)) and 1 <= day_next <= len(seq):
            s = seq[day_next - 1]
        if s is not None and s in WEEKLY_REST_SHIFTS:
            post_cnt += 1
    return post_cnt, n_next


def _weekly_off_rule_met(
    of_vis: int,
    oh_vis: int,
    no_vis: int,
    n_prev: int,
    len_wdays: int,
    post_of: int,
    post_oh: int,
    post_no: int,
    n_next: int,
    carry_next_provided: bool,
) -> bool:
    """
    주(월~일) 2 OF 규칙 충족 여부.
    - 마지막주가 당월 안에서 토·일까지 끝나지 않으면(n_next>0) 차월 초와 한 주로 합산.
    - carry_next_month 가 없으면: 당월·전월 구간만으로 주 2 OF를 강제하지 않음(차월에서 맞출 수 있음).
    - 있으면: 전월+당월+차월 합산으로 weekly_of_equiv_satisfied.
    """
    of_t = of_vis + post_of
    oh_t = oh_vis + post_oh
    no_t = no_vis + post_no
    m = n_prev + len_wdays + n_next
    if m <= 0:
        return True
    if n_next > 0 and not carry_next_provided:
        return True
    return weekly_of_equiv_satisfied(of_t, oh_t, no_t, m)


def weekly_of_equiv_satisfied(of_c: int, oh_c: int, no_c: int, m: int) -> bool:
    """
    월~일 한 주 중 평가에 포함된 일수 m일에 대한 '주 2 OF' 충족 여부.
    (당월 일자 + 전월 동주 carry 일자를 합친 m일에 대해 동일 규칙 적용 가능)
    - of_c: OF 개수만 (연차는 주 2 휴무 인정에 포함되지 않음 — OF+연으로 대체 불가).
    - OF가 2개 이상이면 충족.
    - OH만 2개 이상이어도 충족.
    - 또는 OF+OH, OF+NO, OH+NO (각 1개 이상). NO는 주당 최대 1개(별도 검증).
    - m==1: OF/OH/NO 중 하나라도 있으면 충족.
    """
    if m <= 0:
        return True
    if m == 1:
        return of_c + oh_c + no_c >= 1
    return (
        of_c >= 2
        or oh_c >= 2
        or (of_c >= 1 and oh_c >= 1)
        or (of_c >= 1 and no_c >= 1)
        or (oh_c >= 1 and no_c >= 1)
    )


def _weekly_off_ok_after_of_to_yun(
    of_vis: int,
    oh_vis: int,
    no_vis: int,
    n_prev: int,
    wdays: list,
    post_of: int,
    post_oh: int,
    post_no: int,
    n_next: int,
    carry_next_provided: bool,
) -> bool:
    """
    OF 한 칸을 연으로 바꾼 뒤(OF만 -1, 연은 주 2휴무 인정에 불포함)에도
    weekly_of_equiv 가 성립하는지. 말주가 차월로 이어지고 carry_next 가 없을 때는
    당월에 보이는 일수만으로 엄격히 판정(기존 _weekly_off_rule_met 의 무조건 True 회피).
    """
    new_of = max(0, of_vis - 1)
    len_w = len(wdays)
    if n_next > 0 and not carry_next_provided:
        m = n_prev + len_w
        if m <= 0:
            return True
        return weekly_of_equiv_satisfied(new_of, oh_vis, no_vis, m)
    of_t = new_of + post_of
    oh_t = oh_vis + post_oh
    no_t = no_vis + post_no
    m = n_prev + len_w + n_next
    if m <= 0:
        return True
    return weekly_of_equiv_satisfied(of_t, oh_t, no_t, m)


def _weekly_off_ok_after_oh_to_yun(
    of_vis: int,
    oh_vis: int,
    no_vis: int,
    n_prev: int,
    wdays: list,
    post_of: int,
    post_oh: int,
    post_no: int,
    n_next: int,
    carry_next_provided: bool,
) -> bool:
    """
    OH 한 칸을 연(또는 D)으로 바꾼 뒤(OH만 -1)에도 weekly_of_equiv 가 성립하는지.
    """
    new_oh = max(0, oh_vis - 1)
    len_w = len(wdays)
    if n_next > 0 and not carry_next_provided:
        m = n_prev + len_w
        if m <= 0:
            return True
        return weekly_of_equiv_satisfied(of_vis, new_oh, no_vis, m)
    of_t = of_vis + post_of
    oh_t = new_oh + post_oh
    no_t = no_vis + post_no
    m = n_prev + len_w + n_next
    if m <= 0:
        return True
    return weekly_of_equiv_satisfied(of_t, oh_t, no_t, m)


def _weekly_off_strict_satisfied_for_week(
    sched, n: int, wdays: list, carry, carry_next, week_start: date,
    month_first: date, month_last: date, carry_next_provided: bool,
) -> bool:
    """연차 제외 OF/OH/NO만으로 주 2휴무 충족 여부(차월 미입력 시 당월 구간만 엄격 판정)."""
    pre_of, pre_oh, pre_no, n_prev = _carry_week_prev_month_off_counts(
        carry, n, week_start, month_first,
    )
    post_of, post_oh, post_no, n_next = _carry_week_next_month_off_counts(
        carry_next, n, week_start, month_last,
    )
    of_vis = pre_of + sum(1 for d2 in wdays if sched[n].get(d2) == 'OF')
    oh_vis = pre_oh + sum(1 for d2 in wdays if sched[n].get(d2) == 'OH')
    no_vis = pre_no + sum(1 for d2 in wdays if sched[n].get(d2) == 'NO')
    len_w = len(wdays)
    if n_next > 0 and not carry_next_provided:
        m = n_prev + len_w
        if m <= 0:
            return True
        return weekly_of_equiv_satisfied(of_vis, oh_vis, no_vis, m)
    of_t = of_vis + post_of
    oh_t = oh_vis + post_oh
    no_t = no_vis + post_no
    m = n_prev + len_w + n_next
    if m <= 0:
        return True
    return weekly_of_equiv_satisfied(of_t, oh_t, no_t, m)


def infer_unit_profile(dept_name=None) -> str:
    """
    부서명 → 인원 규칙 프로파일.
    'icu' | 'er' | 'ward' — 알 수 없으면 일반 병동(ward).
    """
    raw = (dept_name or '').strip()
    u = raw.upper()
    if 'ICU' in u or '중환자' in raw:
        return 'icu'
    if 'ER' in u or '응급' in raw:
        return 'er'
    return 'ward'


def daily_regular_staff_targets(
    num_nurses: int, day: dict, head_shift: str, unit_profile: str = 'ward',
) -> tuple[int, int, tuple[int, int]]:
    """
    일반 간호사(수간 제외) 일별 목표: (E명, N명, (D하한, D상한)).
    부서별 분기는 if / elif / else 만 사용(ICU·ER·일반병동 로직 분리).
    """
    up = (unit_profile or 'ward').strip().lower()
    if up not in ('icu', 'er', 'ward'):
        up = 'ward'
    if up == 'icu':
        return (4, 3, (4, 4))
    elif up == 'er':
        d_lo, d_hi = d_regular_d_bounds(num_nurses, day, head_shift, 'er')
        return (2, 2, (d_lo, d_hi))
    else:
        d_lo, d_hi = d_regular_d_bounds(num_nurses, day, head_shift, 'ward')
        return (2, 2, (d_lo, d_hi))


def d_regular_d_bounds(
    num_nurses: int,
    day: dict,
    head_shift: str,
    unit_profile: str = 'ward',
) -> tuple[int, int]:
    """
    일반 간호사(수간 제외) 일별 D 인원 하한·상한. CP-SAT·검증 공통.

    - ICU: D=4 고정(타겟은 daily_regular_staff_targets 참고).
    - ER: 평일이고 수간 A1이면 일반간 D=1(정확). 주말/공휴이거나 수간 비A1이면 일반간 D=2(정확).
    - ward: 총 10명 — 주말/공휴 D=2; 평일 수간 A1이면 D=1, 비A1이면 D=2.
            총 11명 — 수간 A1일 때 D 1~2, 비A1(휴가 등)일 때 D=2만(최소 2 하드).
            총 12명 이상 — D 2~3(수간과 무관 최소 2).
    """
    h_is_a1 = head_shift == 'A1'
    is_we = day['is_weekend'] or day['is_holiday']
    up = (unit_profile or 'ward').strip().lower()
    if up not in ('icu', 'er', 'ward'):
        up = 'ward'

    if up == 'icu':
        return (4, 4)
    elif up == 'er':
        if h_is_a1 and not is_we:
            return (1, 1)
        return (2, 2)
    else:
        # ── 일반 병동(ward): 인원·수간 A1 여부 — if / elif 명시 ───────────────
        if num_nurses == 11:
            if h_is_a1:
                return (1, 2)
            else:
                return (2, 2)
        elif num_nurses >= 12:
            return (2, 3)
        elif num_nurses == 10:
            if is_we:
                return (2, 2)
            elif h_is_a1:
                return (1, 1)
            else:
                return (2, 2)
        else:
            if is_we:
                return (2, 2)
            elif not h_is_a1:
                return (2, 2)
            else:
                return (2, 2)


def d_slots_per_day(num_nurses: int, day: dict, head_is_a1: bool, unit_profile: str = 'ward') -> int:
    """호환용. 가능한 경우 상한 쪽. 신규 코드는 d_regular_d_bounds 사용."""
    lo, hi = d_regular_d_bounds(num_nurses, day, 'A1' if head_is_a1 else 'OF', unit_profile)
    return hi


def d_assignment_target(num_nurses: int, day: dict, head_is_a1: bool, unit_profile: str = 'ward') -> int:
    """호환용 하한. 신규 코드는 d_regular_d_bounds 사용."""
    lo, hi = d_regular_d_bounds(num_nurses, day, 'A1' if head_is_a1 else 'OF', unit_profile)
    return lo


def _monthly_head_nurse_of_count(sched: dict, days: list) -> int:
    """수간호사(0)의 월간 'OF' 칸 개수. 일반 간호사 OF 상한·쿼터를 이 값에 맞춘다."""
    return sum(1 for day in days if sched.get(0, {}).get(day['day']) == 'OF')


def _monthly_head_nurse_oh_count(sched: dict, days: list) -> int:
    """수간호사(0)의 월간 'OH' 칸 개수(공휴 휴무)."""
    return sum(1 for day in days if sched.get(0, {}).get(day['day']) == 'OH')

# ── 스케줄 생성 (OR-Tools CP-SAT) ───────────────────────────────────────────
def _normalize_pregnant_nurses(raw, num_nurses, nurse_names=None):
    """
    임산부(야간 불가) — 간호사 인덱스 frozenset. 수간(0) 제외. 이름 또는 정수 리스트.
    """
    names = nurse_names if nurse_names is not None else get_nurse_names(num_nurses)
    name_to_i = {str(nm).strip(): i for i, nm in enumerate(names)}
    out: set[int] = set()
    if not raw:
        return frozenset()
    seq = raw if isinstance(raw, (list, tuple, set)) else []
    for item in seq:
        try:
            ni = int(item)
        except (TypeError, ValueError):
            sk = str(item).strip()
            ni = name_to_i.get(sk)
            if ni is None:
                continue
        if ni == 0 or not (1 <= ni < num_nurses):
            continue
        out.add(ni)
    return frozenset(out)


def solve_schedule(num_nurses, requests, holidays=(), forbidden_pairs=None, carry_in=None,
                   regenerate=False, rng_seed=None, nurse_names=None, carry_next_month=None,
                   shift_bans=None, not_available=None, pregnant_nurses=None,
                   unit_profile: str = 'ward'):
    """
    CP-SAT: 신청 근무 하드 고정; 함께 근무 불가는 총원 12명 이상이면 같은 날·같은 D/E/N 하드,
    미만이면 5M급 벌점. 일일 E/N/D·기타 패턴은 벌점 완화. 임산부·OH 등 사전 검증 유지.
    반환 4번째: 가변 규칙 위반 등 `validate_schedule` issue 목록.
    """
    try:
        from schedule_cpsat import solve_schedule_cpsat
        return solve_schedule_cpsat(
            num_nurses,
            requests,
            holidays,
            forbidden_pairs=forbidden_pairs,
            carry_in=carry_in,
            carry_next_month=carry_next_month,
            shift_bans=shift_bans,
            not_available=not_available,
            pregnant_nurses=pregnant_nurses,
            nurse_names=nurse_names,
            regenerate=regenerate,
            rng_seed=rng_seed,
            unit_profile=unit_profile,
        )
    except Exception as e:
        print(f'[오류] {e}')
        return None, False, str(e), []


def _normalize_forbidden_pairs(forbidden_pairs, num_nurses):
    """
    forbidden_pairs: [(i,j), ...] 또는 [(i,j, ['D','E']), ...] 등
    반환: dict (min,max) -> frozenset({'D','E','N'}의 부분집합) — 수간호사(0) 포함.
    """
    out: dict[tuple[int, int], frozenset] = {}
    if not forbidden_pairs:
        return out
    for pair in forbidden_pairs:
        if not pair:
            continue
        shifts = frozenset({'D', 'E', 'N'})
        if len(pair) == 2:
            try:
                a, b = int(pair[0]), int(pair[1])
            except (TypeError, ValueError):
                continue
        elif len(pair) >= 3:
            try:
                a, b = int(pair[0]), int(pair[1])
            except (TypeError, ValueError):
                continue
            raw = pair[2]
            if isinstance(raw, (list, tuple, set, frozenset)):
                shifts = frozenset(x for x in raw if x in ('D', 'E', 'N'))
            elif isinstance(raw, str):
                shifts = frozenset(
                    x.strip() for x in raw.replace(',', ' ').split() if x.strip() in ('D', 'E', 'N')
                )
            if not shifts:
                shifts = frozenset({'D', 'E', 'N'})
        else:
            continue
        if not (0 <= a < num_nurses and 0 <= b < num_nurses and a != b):
            continue
        key = (min(a, b), max(a, b))
        out[key] = out.get(key, frozenset()) | shifts
    return out


def _normalize_shift_bans(shift_bans, num_nurses):
    """
    shift_bans: dict[int, str] — 간호사 인덱스 → 'd_only' | 'no_d' | 'no_e' | 'no_n'
    d_only: D/E/N 중 E·N 자동 배제(데이만 가능).
    반환: dict[int, frozenset] — 간호사별 금지 D/E/N (수간호사 0은 무시).
    """
    mode_to_banned = {
        'd_only': frozenset({'E', 'N'}),
        'no_d': frozenset({'D'}),
        'no_e': frozenset({'E'}),
        'no_n': frozenset({'N'}),
    }
    out: dict[int, frozenset] = {}
    if not shift_bans:
        return out
    for k, mode in shift_bans.items():
        try:
            ni = int(k)
        except (TypeError, ValueError):
            continue
        if not (0 <= ni < num_nurses):
            continue
        m = str(mode).strip() if mode is not None else ''
        if m not in mode_to_banned:
            continue
        out[ni] = mode_to_banned[m]
    return out


# CP-SAT·검증: 날짜·근무 지정 불가(소프트 벌점). 신청 불일치(1천만)보다 낮게 두어 신청 우선.
UNAVAILABLE_PENALTY_WEIGHT = 100_000

# 엔진이 날짜별로 막을 수 있는 시프트(OF/OH 포함)
UNAVAILABLE_ENTRY_SHIFTS = frozenset({'D', 'E', 'N', 'OF', 'OH'})


def unavailable_violation_warn_message(nurse_name: str, day_num: int, shift: str) -> str:
    return (
        f'{nurse_name}님, 불가 신청한 {YEAR}년 {MONTH}월 {day_num}일/{shift} 근무가 '
        '인원 부족으로 배정되었습니다. 수기 수정 바랍니다.'
    )


def _request_shift_for_cell(requests, n: int, dn: int):
    """requests dict에서 (n, dn) 신청 시프트 또는 None."""
    if not requests or not isinstance(requests, dict):
        return None
    ds = requests.get(n)
    if not isinstance(ds, dict):
        ds = requests.get(str(n))
    if not isinstance(ds, dict):
        return None
    v = ds.get(dn)
    if v is None:
        v = ds.get(str(dn))
    if v is None:
        return None
    return str(v).strip()


def _normalize_not_available(raw, num_nurses, nurse_names=None):
    """
    프론트·API용 not_available 리스트 → (간호사 인덱스, 일, 시프트) 집합.
    항목 예: {"nurse":"간호사3","day":5,"shift":"N"}, {"n":1,"d":5,"s":"N"}, [1,5,"N"]
    수간(0) 행은 무시. 시프트는 D/E/N/OF/OH만.
    """
    names = nurse_names if nurse_names is not None else get_nurse_names(num_nurses)
    name_to_i = {str(nm).strip(): i for i, nm in enumerate(names)}
    out: set[tuple[int, int, str]] = set()
    if not raw:
        return frozenset()
    items = raw if isinstance(raw, (list, tuple)) else []
    for item in items:
        n_idx = dni = None
        sh = None
        if isinstance(item, (list, tuple)) and len(item) >= 3:
            try:
                n_idx = int(item[0])
            except (TypeError, ValueError):
                n_idx = name_to_i.get(str(item[0]).strip())
            try:
                dni = int(item[1])
            except (TypeError, ValueError):
                continue
            sh = item[2]
        elif isinstance(item, dict):
            nk = None
            for key in ('nurse', 'nurse_name', 'name', 'nurse_idx'):
                if key in item and item[key] is not None:
                    nk = item[key]
                    break
            if nk is None and 'n' in item:
                nk = item['n']
            if nk is None:
                continue
            try:
                n_idx = int(nk)
            except (TypeError, ValueError):
                sk = str(nk).strip()
                if sk not in name_to_i:
                    continue
                n_idx = name_to_i[sk]
            dk = None
            for key in ('day', 'd', 'dn', 'date'):
                if key in item and item[key] is not None:
                    dk = item[key]
                    break
            if dk is None:
                continue
            try:
                dni = int(dk)
            except (TypeError, ValueError):
                continue
            skk = None
            for key in ('shift', 's', 'sh'):
                if key in item and item[key] is not None:
                    skk = item[key]
                    break
            sh = skk
        else:
            continue
        if n_idx is None or dni is None or sh is None:
            continue
        if not (0 <= n_idx < num_nurses) or n_idx == 0:
            continue
        if not (1 <= dni <= NUM_DAYS):
            continue
        sh = str(sh).strip()
        if sh not in UNAVAILABLE_ENTRY_SHIFTS:
            continue
        out.add((n_idx, dni, sh))
    return frozenset(out)


# 전월 말 근무 꼬리 (월 경계 규칙용) — 최대 14일, 오래된 날 → 최근 날 순
CARRY_MAX_DAYS = 14

# 간호사당 월 N(야간) 상한 — 수간 포함 총원이 11명 이상이어도 동일(7개까지).
# 일일 N 2명·목표 합=2×말일·공평 분배 — CP-SAT·validate_schedule 공통.
N_ABS_MAX = 7


def _compute_n_targets_fair(num_reg: int, total_slots: int, n_max: int = N_ABS_MAX) -> list:
    """
    일반간호사 num_reg명에게 total_slots개의 N 배정 목표(합=total_slots, 각 n_max 이하).
    먼저 균등 분배 후 min(·, n_max) 적용, 부족분은 가장 적게 받은 사람부터 +1(상한 내).
    인원이 늘어도(예: 수간 포함 11명 이상) N 상한·공평 목표 분배 방식은 동일하며, 변하는 것은 num_reg·합계 슬롯뿐이다.
    (예: 8명·60슬롯은 7×8=56으로 상한만으로는 60을 못 채움 — 물리적 한계)
    """
    if num_reg <= 0:
        return []
    if total_slots <= 0:
        return [0] * num_reg
    base = total_slots // num_reg
    rem = total_slots % num_reg
    targets = [min(base + (1 if i < rem else 0), n_max) for i in range(num_reg)]
    s = sum(targets)
    if s < total_slots:
        deficit = total_slots - s
        while deficit > 0:
            best_i = -1
            best_val = None
            for i in range(num_reg):
                if targets[i] >= n_max:
                    continue
                if best_i < 0 or targets[i] < best_val:
                    best_i, best_val = i, targets[i]
            if best_i < 0:
                break
            targets[best_i] += 1
            deficit -= 1
    elif s > total_slots:
        surplus = s - total_slots
        while surplus > 0:
            worst_i = -1
            worst_val = None
            for i in range(num_reg):
                if targets[i] <= 0:
                    continue
                if worst_i < 0 or targets[i] > worst_val:
                    worst_i, worst_val = i, targets[i]
            if worst_i < 0:
                break
            targets[worst_i] -= 1
            surplus -= 1
    return targets
def _normalize_carry_in(carry_in, num_nurses):
    """
    carry_in: { 간호사인덱스: [시프트, ...], ... } 또는 None
    값은 전월 말에서 이어지는 날들(오래된 것부터).
    """
    if not carry_in:
        return {}
    out = {}
    if not isinstance(carry_in, dict):
        return {}
    for k, v in carry_in.items():
        try:
            ni = int(k)
        except (TypeError, ValueError):
            continue
        if not (0 <= ni < num_nurses):
            continue
        if isinstance(v, (list, tuple)):
            seq = list(v)
        elif isinstance(v, str):
            seq = [x.strip() for x in v.replace(',', ' ').split() if x.strip()]
        else:
            continue
        if not seq:
            continue
        out[ni] = tuple(seq[-CARRY_MAX_DAYS:])
    return out


def _shift_k_days_before(sched_map, carry, n, dn, k):
    """
    이번 달 dn일의 k일 전 근무.
    carry[n] = 전월 말쪽 날들(오래된→최근), 마지막 원소가 전월 마지막 날(= 이번 달 1일 직전).
    """
    if k <= 0:
        return None
    td = dn - k
    if td >= 1:
        return sched_map.get(n, {}).get(td)
    # dn일보다 앞선 날짜가 당월에 없으면 전월 carry에서 이어짐.
    # k일 전 = (dn-1)일은 당월, 나머지 need일은 carry 끝에서 need칸 (need = k - dn + 1).
    # 예: dn=1,k=1 → need=1 → c[-1] (전월 말). 예전 idx=-td는 td=0일 때 0이 되어 N-D 검사가 누락됨.
    c = carry.get(n) or ()
    need = k - dn + 1
    if need >= 1 and need <= len(c):
        return c[-need]
    return None


def _req_shift_get(requests: dict | None, n: int, dn: int) -> str | None:
    """requests에서 n번 간호사·dn일 신청 시프트(없으면 None)."""
    if not requests:
        return None
    ds = requests.get(n)
    if not isinstance(ds, dict):
        ds = requests.get(str(n))
    if not isinstance(ds, dict):
        return None
    v = ds.get(dn)
    if v is None:
        v = ds.get(str(dn))
    if v is None:
        return None
    return str(v).strip()


def collect_request_advice_warnings(
    schedule: dict,
    num_nurses: int,
    holidays: tuple | list,
    nurse_names: list | tuple | None,
    carry_in: dict | None,
    carry_next_month: dict | None,
    requests: dict | None,
) -> list[dict]:
    """
    신청 근무(연·OF 등)에 대한 지능형 권고만 반환한다.
    근무표 배정이나 신청 값은 변경하지 않으며, 경고만 추가한다.
    """
    if not requests:
        return []
    days = get_april_days(holidays)
    names = nurse_names if nurse_names is not None else get_nurse_names(num_nurses)
    carry = _normalize_carry_in(carry_in, num_nurses)
    carry_next_provided = carry_next_month is not None
    carry_next = _normalize_carry_in(carry_next_month, num_nurses) if carry_next_provided else {}
    month_first = date(YEAR, MONTH, 1)
    month_last = date(YEAR, MONTH, NUM_DAYS)
    out: list[dict] = []

    wk_map: dict[int, list] = {}
    for day in days:
        sun = _week_sunday(day['date'])
        wk_map.setdefault(sun.toordinal(), []).append(day['day'])

    def sh(ni: int, dni: int) -> str:
        return schedule.get(ni, {}).get(dni, '')

    for ni, ds in list(requests.items()):
        try:
            ni = int(ni)
        except (TypeError, ValueError):
            continue
        if not (0 <= ni < num_nurses):
            continue
        nm = names[ni]
        for dn, req_raw in (ds or {}).items():
            try:
                dni = int(dn)
            except (TypeError, ValueError):
                continue
            if not (1 <= dni <= NUM_DAYS):
                continue
            rs = str(req_raw).strip()
            if sh(ni, dni) != rs:
                continue

            if rs == '연':
                sun = _week_sunday(date(YEAR, MONTH, dni))
                wdays = wk_map.get(sun.toordinal(), [])
                if not wdays:
                    continue
                pre_rest, n_prev = _carry_week_prev_rest_total(carry, ni, sun, month_first)
                post_rest, _nn = _carry_week_next_rest_total(carry_next, ni, sun, month_last)
                post_part = post_rest if carry_next_provided else 0
                carry_prev_ok = _carry_prev_week_tail_complete(carry, ni, n_prev)
                if n_prev > 0 and not carry_prev_ok:
                    continue
                rest_others = sum(
                    1 for d2 in wdays
                    if d2 != dni and sh(ni, d2) in WEEKLY_REST_SHIFTS
                )
                if pre_rest + rest_others + post_part >= 2:
                    out.append({
                        'level': 'warn',
                        'msg': (
                            f"{nm}님, 해당 날짜는 주간 오프(OF)로 대체 가능한 자리입니다. "
                            f"연차를 아끼려면 OF로 변경을 고려해보세요."
                        ),
                    })

            if rs == 'OF':
                sun = _week_sunday(date(YEAR, MONTH, dni))
                wdays = wk_map.get(sun.toordinal(), [])
                if not wdays:
                    continue
                pre_of, pre_oh, _pre_no, _np = _carry_week_prev_month_off_counts(
                    carry, ni, sun, month_first,
                )
                prior_carry_of_oh = pre_of + pre_oh
                req_other = 0
                for d2 in wdays:
                    if d2 == dni:
                        continue
                    q = _req_shift_get(requests, ni, d2)
                    if q in ('OF', 'OH'):
                        req_other += 1
                if prior_carry_of_oh + req_other >= 2:
                    out.append({
                        'level': 'warn',
                        'msg': (
                            f"{nm}님, 이번 주는 이미 2회 오프를 사용하셨습니다. "
                            f"추가 휴무가 필요하시면 '연차(연)'로 변경을 고려해보세요."
                        ),
                    })

    return out


def validate_schedule(schedule, num_nurses, holidays=(), forbidden_pairs=None,
                      nurse_names=None, carry_in=None, requests=None, carry_next_month=None,
                      shift_bans=None, not_available=None, engine_soft_report: bool = False,
                      unit_profile: str = 'ward', cell_highlights_out: list | None = None):
    """
    생성된 스케줄을 규칙에 따라 검증하고 위반 사항 목록을 반환한다.
    num_nurses: 수간호사 포함 총원(예: 11 = 수간 1 + 일반 10).
    forbidden_pairs: [(i,j), ...] 또는 [(i,j,['D','E']), ...] — 같은 날 동시 배치 금지(수간 0 포함)
    nurse_names: 표시용 이름 (없으면 기본 수간호사/간호사1…)
    carry_in: (선택) 전월 말 근무 꼬리 — 월 경계 규칙 검증용
    carry_next_month: (선택) 차월 초 근무 — 마지막주(당월 말~차월 일요일) 주 2 OF 합산 검증용
    requests: (선택) 생성 시 사용한 신청 — 있으면 스케줄 셀과 반드시 일치해야 함
    shift_bans: (선택) dict[int,str] — 간호사 인덱스별 근무불가(d_only|no_d|no_e|no_n) — 엔진은 소프트, 검증은 경고
    not_available: (선택) 날짜·근무 지정 불가 리스트 — 엔진 소프트, 위반 시 동일 경고 문구
    engine_soft_report: True면 기존 error 급도 warn으로 올리고 수기검토용 접두를 붙임(CP-SAT 소프트 엔진 결과 표시용).
    cell_highlights_out: 있으면 {'level','msg','cells': frozenset[(간호사인덱스, 일자)]} 를 누적(시각화용).
    Returns: list of dict  { 'level': 'error'|'warn', 'msg': str }
    신청이 있으면 맨 앞에 `collect_request_advice_warnings` 지능형 권고가 붙을 수 있다.
    """
    issues = []
    cell_records = cell_highlights_out

    def _freeze_cells(cells):
        if not cells:
            return frozenset()
        return frozenset((int(a), int(b)) for a, b in cells)
    days = get_april_days(holidays)
    _dn_holiday = {d['day']: bool(d['is_holiday']) for d in days}
    nurses = list(range(1, num_nurses))
    names = nurse_names if nurse_names is not None else get_nurse_names(num_nurses)
    fp_map = _normalize_forbidden_pairs(forbidden_pairs, num_nurses)
    den_bans = _normalize_shift_bans(shift_bans, num_nurses)
    na_set = _normalize_not_available(not_available, num_nurses, nurse_names=names)
    carry = _normalize_carry_in(carry_in, num_nurses)
    carry_next = _normalize_carry_in(carry_next_month, num_nurses) if carry_next_month else {}
    carry_next_provided = carry_next_month is not None
    month_last = date(YEAR, MONTH, NUM_DAYS)
    OFF_SET = {'OF', 'OH', 'NO'}
    REST_GAP = frozenset(OFF_SET | {'연'})
    GAP_WORK = frozenset({'D', 'E', 'N', 'EDU', '공', '병', '경', 'A1'})

    def sh(n, dn):
        return schedule.get(n, {}).get(dn, '')

    def vk(n, dn, k):
        return _shift_k_days_before(schedule, carry, n, dn, k)

    def is_off(s):
        return s in OFF_SET or s in ('', None)

    def err(msg, cells=()):
        if engine_soft_report:
            m = f'【자동배정·규칙위반 수기검토】{msg}'
            issues.append({
                'level': 'warn',
                'msg': m,
            })
            if cell_records is not None:
                cell_records.append({
                    'level': 'warn',
                    'msg': m,
                    'cells': _freeze_cells(cells),
                })
        else:
            issues.append({'level': 'error', 'msg': msg})
            if cell_records is not None:
                cell_records.append({
                    'level': 'error',
                    'msg': msg,
                    'cells': _freeze_cells(cells),
                })

    def warn(msg, cells=()):
        issues.append({'level': 'warn',  'msg': msg})
        if cell_records is not None:
            cell_records.append({
                'level': 'warn',
                'msg': msg,
                'cells': _freeze_cells(cells),
            })

    if requests:
        for ni, ds in requests.items():
            try:
                ni = int(ni)
            except (TypeError, ValueError):
                continue
            if not (0 <= ni < num_nurses):
                continue
            nm = names[ni]
            for dn, req_shift in (ds or {}).items():
                try:
                    dni = int(dn)
                except (TypeError, ValueError):
                    continue
                if not (1 <= dni <= NUM_DAYS):
                    continue
                cur = schedule.get(ni, {}).get(dni, '')
                if cur != req_shift:
                    err(
                        f"{nm} 신청 불일치(절대): {dni}일 신청={req_shift!r} "
                        f"근무표={cur!r}",
                        [(ni, dni)],
                    )

    if requests:
        for n in nurses:
            nm = names[n]
            ds = requests.get(n)
            if not isinstance(ds, dict):
                ds = requests.get(str(n))
            if not isinstance(ds, dict):
                ds = {}
            for dn in range(1, NUM_DAYS + 1):
                s = sh(n, dn)
                if s not in REQUEST_ONLY_SHIFTS:
                    continue
                rs = ds.get(dn)
                if rs is None:
                    rs = ds.get(str(dn))
                if rs != s:
                    err(
                        f"{nm} {s}(신청 전용)은 해당 일에 신청이 있을 때만 허용: {dn}일 "
                        f"(자동 배정 불가)",
                        [(n, dn)],
                    )

    if den_bans:
        for n in range(num_nurses):
            b = den_bans.get(n)
            if not b:
                continue
            nm = names[n]
            for dn in range(1, NUM_DAYS + 1):
                rs = _request_shift_for_cell(requests, n, dn) if requests else None
                s = sh(n, dn)
                if s not in b:
                    continue
                if rs == s:
                    continue
                warn(
                    unavailable_violation_warn_message(nm, dn, s),
                    [(n, dn)],
                )

    if na_set:
        for (n, dn, s) in sorted(na_set):
            if not (0 <= n < num_nurses):
                continue
            nm = names[n]
            rs = _request_shift_for_cell(requests, n, dn) if requests else None
            if sh(n, dn) != s:
                continue
            if rs == s:
                continue
            warn(
                unavailable_violation_warn_message(nm, dn, s),
                [(n, dn)],
            )

    _holiday_day_nums = {d['day'] for d in days if d['is_holiday']}
    for n in range(num_nurses):
        nm = names[n]
        for dn in range(1, NUM_DAYS + 1):
            if sh(n, dn) == 'OH' and dn not in _holiday_day_nums:
                err(
                    f"{nm} OH는 화면·폼에 입력한 공휴일 목록에 포함된 날에만 가능합니다: {dn}일",
                    [(n, dn)],
                )

    # ── ① 일일 인력 요구 (일반 간호사만 집계; 수간 A1은 제외) ─────────────────
    _uprof = (unit_profile or 'ward').strip().lower()
    if _uprof not in ('icu', 'er', 'ward'):
        _uprof = 'ward'
    for day in days:
        dn = day['day']
        label = f"{dn}일({day['weekday_name']})"
        head = sh(0, dn)
        is_we = day['is_weekend'] or day['is_holiday']
        d_cnt = sum(1 for n in nurses if sh(n, dn) == 'D')
        e_cnt = sum(1 for n in nurses if sh(n, dn) == 'E')
        n_cnt = sum(1 for n in nurses if sh(n, dn) == 'N')

        need_e, need_n, (d_lo, d_hi) = daily_regular_staff_targets(
            num_nurses, day, head, _uprof,
        )
        if _uprof == 'icu':
            unit_lbl = '(중환자실·일반간호사)'
        elif _uprof == 'er':
            unit_lbl = '(응급실·일반간호사)'
        else:
            unit_lbl = '(일반병동·일반간호사)'

        _day_col = [(nn, dn) for nn in range(num_nurses)]
        if e_cnt != need_e:
            err(
                f"{label} {unit_lbl} E 인원: {e_cnt}명 "
                f"(필요 정확히 {need_e}명, 평일·주말·공휴 동일)",
                _day_col,
            )
        if n_cnt != need_n:
            err(
                f"{label} {unit_lbl} N 인원: {n_cnt}명 "
                f"(필요 정확히 {need_n}명, 평일·주말·공휴 동일)",
                _day_col,
            )

        tag = '[주말/공휴일]' if is_we else '[평일]'
        if d_cnt < d_lo:
            err(
                f"{label} {tag} {unit_lbl} D 인원 부족: {d_cnt}명 "
                f"(허용 {d_lo}~{d_hi}명, 수간 {head or '—'})",
                _day_col,
            )
        if d_cnt > d_hi:
            err(
                f"{label} {tag} {unit_lbl} D 인원 초과: {d_cnt}명 "
                f"(허용 {d_lo}~{d_hi}명, 수간 {head or '—'})",
                _day_col,
            )

    # ── ①b 함께 근무 불가 (선택한 시프트에 한해 같은 날 동시 배치 금지, 수간 포함) ─
    if fp_map:
        all_idx = list(range(num_nurses))
        for day in days:
            dn = day['day']
            label = f"{dn}일({day['weekday_name']})"
            for shift in ('D', 'E', 'N'):
                team = [n for n in all_idx if sh(n, dn) == shift]
                for i in range(len(team)):
                    for j in range(i + 1, len(team)):
                        a, b = team[i], team[j]
                        key = (min(a, b), max(a, b))
                        if key in fp_map and shift in fp_map[key]:
                            err(
                                f"{label} 함께 근무 불가: {names[a]} · {names[b]} "
                                f"동시 {shift}",
                                [(a, dn), (b, dn)],
                            )

    # ── ② 개인별 규칙 ───────────────────────────────────────────────────────
    for n in nurses:
        ns   = schedule.get(n, {})
        nm   = names[n]

        # N 총 개수
        n_total = sum(1 for v in ns.values() if v == 'N')
        if n_total > N_ABS_MAX:
            err(
                f"{nm} N 초과: {n_total}개 (최대 {N_ABS_MAX}개)",
                [(n, d) for d, v in ns.items() if v == 'N'],
            )

        # N 블록 분석
        n_days = sorted(d for d, s in ns.items() if s == 'N')
        blocks = []
        if n_days:
            blk = [n_days[0]]
            for d in n_days[1:]:
                if d == blk[-1] + 1:
                    blk.append(d)
                else:
                    blocks.append(blk); blk = [d]
            blocks.append(blk)

        for blk in blocks:
            if len(blk) < 2:
                # 말일 단독 N: 3-3-1 / 2-3-1 / 3-2-1 등 — 당월 마지막 날(NUM_DAYS)만 (31일 말달은 31일)
                if blk[0] != NUM_DAYS:
                    err(
                        f"{nm} N 블록 단독: {blk[0]}일 "
                        f"(1개, 당월 말일({NUM_DAYS}일)만 단독 허용 — 3-3-1·2-3-1·3-2-1)",
                        [(n, blk[0])],
                    )
            elif len(blk) > 3:
                err(
                    f"{nm} N 블록 초과: {blk[0]}~{blk[-1]}일 ({len(blk)}개, 최대 3개)",
                    [(n, d) for d in blk],
                )

        n_gap_min = 5
        for i in range(len(blocks) - 1):
            gap = blocks[i+1][0] - blocks[i][-1] - 1
            if gap < n_gap_min:
                warn(
                    f"{nm} N 블록 간격 부족: {blocks[i][-1]}일→{blocks[i+1][0]}일 "
                    f"({gap}일, 최소 {n_gap_min}일)",
                    [(n, blocks[i][-1]), (n, blocks[i + 1][0])],
                )

        # 전월 말 N → 당월 1일(연속 N 아님): 공휴 OH / 평일 OF · 1일 직접 공가 불가
        cseq = list(carry.get(n, ()))
        if cseq and cseq[-1] == 'N':
            s_first = sh(n, 1)
            if s_first == '공':
                err(
                    f"{nm} 전월 말 N 직후 당월 1일 공가 금지: 야간 후 공가는 OF/OH 휴가 "
                    f"이틀 이상 연속된 뒤에만 배치 가능합니다.",
                    [(n, 1)],
                )
            if s_first != 'N':
                need0 = 'OH' if days[0]['is_holiday'] else 'OF'
                if s_first != need0:
                    err(
                        f"{nm} N블록 직후 휴무 위반: 전월 말 N 이후 1일 "
                        f"({s_first}, 필요 {need0})",
                        [(n, 1)],
                    )

        # N블록 말 직후(당월): 공휴이면 OH, 아니면 OF
        for blk in blocks:
            end = blk[-1]
            if end >= NUM_DAYS:
                continue
            s1 = sh(n, end + 1)
            need = 'OH' if _dn_holiday.get(end + 1) else 'OF'
            if s1 != need:
                err(
                    f"{nm} N블록 직후 휴무 위반: {end}일 N 다음 {end+1}일 "
                    f"({s1 or '빈칸'}, 필요 {need})",
                    [(n, end), (n, end + 1)],
                )

        for blk in blocks:
            end = blk[-1]
            if end >= NUM_DAYS - 1:
                continue
            s1 = sh(n, end + 1)
            s2 = sh(n, end + 2)
            if s1 in ('OF', 'OH') and s2 == 'D':
                err(
                    f"{nm} N-휴무-D 금지: {end}일N→{end+1}일{s1}→{end+2}일D",
                    [(n, end), (n, end + 1), (n, end + 2)],
                )
            if s1 in ('OF', 'OH') and s2 == 'EDU':
                err(
                    f"{nm} N-휴무-교육 금지: {end}일N→{end+1}일{s1}→{end+2}일EDU",
                    [(n, end), (n, end + 1), (n, end + 2)],
                )

        # E 직후 D·EDU 직접 금지 (전월 말 E → 당월 1일 포함). E→OF/OH 후 공(E-OF-공)은 허용.
        for dn in range(1, NUM_DAYS + 1):
            if vk(n, dn, 1) == 'E' and sh(n, dn) in ('D', 'EDU'):
                _cells_ed = [(n, dn)]
                if dn > 1:
                    _cells_ed.append((n, dn - 1))
                err(
                    f"{nm} E 직후 {sh(n, dn)} 금지: 전일E→{dn}일{sh(n, dn)}",
                    _cells_ed,
                )

        # N-D 금지 (전날 야간 직후 데이 — 절대 불가, 전월 말 N 포함)
        for dn in range(1, NUM_DAYS + 1):
            if vk(n, dn, 1) == 'N' and sh(n, dn) == 'D':
                _cells_nd = [(n, dn)]
                if dn > 1:
                    _cells_nd.append((n, dn - 1))
                err(f"{nm} N-D 금지: 전일N→{dn}일D", _cells_nd)

        # 연속 근무 최대 5일 (전월 꼬리 + 당월) — D/E/N/공/EDU만 합산(연차 등은 끊김)
        seq = list(carry.get(n, ())) + [sh(n, d) for d in range(1, NUM_DAYS + 1)]
        streak_max = 5
        streak = 0
        for s in seq:
            if s in STREAK_WORK_SHIFTS:
                streak += 1
                if streak > streak_max:
                    err(
                        f"{nm} 연속근무 초과: 전월이월·당월 합산 {streak}일 "
                        f"(최대 {streak_max}일 연속)",
                        [(n, d) for d in range(1, NUM_DAYS + 1) if sh(n, d) in STREAK_WORK_SHIFTS],
                    )
            else:
                streak = 0

        # 쉬는 날(OF/OH/NO/연) 사이 근무: 0일(붙은 휴무) OK, 1일은 섬 경고,
        # 연속근무 5일 하드와 맞추어 사이 근무 2~5일만 허용(5일 초과 시 오류).
        gap_anchors = sorted(d for d, s in ns.items() if s in REST_GAP)
        prev_a = None
        for od in gap_anchors:
            if prev_a is not None:
                work_btw = sum(
                    1 for d in range(prev_a + 1, od)
                    if sh(n, d) in STREAK_WORK_SHIFTS
                )
                la = sh(n, prev_a) or "?"
                ra = sh(n, od) or "?"
                _btw_cells = [(n, d) for d in range(prev_a + 1, od) if sh(n, d) in STREAK_WORK_SHIFTS]
                if work_btw == 1:
                    warn(
                        f"{nm} 쉬는 날 사이 근무 1일(섬): {prev_a}일{la}~{od}일{ra} "
                        f"— 0일 또는 2~5일이어야 함",
                        _btw_cells,
                    )
                elif work_btw > 5:
                    err(
                        f"{nm} 쉬는 날 사이 근무 과다(절대): {prev_a}일{la}~{od}일{ra} "
                        f"사이 근무 {work_btw}일 — 최대 5일, 공가·교육 포함",
                        _btw_cells,
                    )
            prev_a = od

        # OF/OH 수간 대비(경고만 — 생성·화면 표시는 막지 않음)
        head_of_q = _monthly_head_nurse_of_count(schedule, days)
        head_oh_q = _monthly_head_nurse_oh_count(schedule, days)
        of_only = sum(1 for v in ns.values() if v == 'OF')
        oh_only = sum(1 for v in ns.values() if v == 'OH')
        if of_only > head_of_q:
            warn(
                f"{nm} OF 초과: {of_only}개 (수간호사 OF {head_of_q}개까지)",
                [(n, d) for d, v in ns.items() if v == 'OF'],
            )
        if oh_only > head_oh_q:
            warn(
                f"{nm} OH 초과: {oh_only}개 (수간호사 OH {head_oh_q}개까지)",
                [(n, d) for d, v in ns.items() if v == 'OH'],
            )

        # 주(일~토): 휴무(OF/OH/NO/연/공/병/경) 합산 최소 2일 하드, 기타 OF한도·NO 등은 보조
        if days:
            wk_map: dict[int, list] = {}
            for day in days:
                dt  = day['date']
                sun = _week_sunday(dt)
                wk_map.setdefault(sun.toordinal(), []).append(day['day'])
            month_first = date(YEAR, MONTH, 1)
            for _wk, wdays in wk_map.items():
                if not wdays:
                    continue
                sun_date = date.fromordinal(_wk)
                pre_of, pre_oh, pre_no, n_prev = _carry_week_prev_month_off_counts(
                    carry, n, sun_date, month_first,
                )
                post_of, post_oh, post_no, n_next = _carry_week_next_month_off_counts(
                    carry_next, n, sun_date, month_last,
                )
                pre_rest, _np = _carry_week_prev_rest_total(
                    carry, n, sun_date, month_first,
                )
                post_rest, _nn = _carry_week_next_rest_total(
                    carry_next, n, sun_date, month_last,
                )
                of_vis = pre_of + sum(1 for d2 in wdays if sh(n, d2) == 'OF')
                tot_of_wk = of_vis + (post_of if carry_next_provided else 0)
                _wk_cells_base = [(n, d2) for d2 in wdays]
                if tot_of_wk > 3:
                    warn(
                        f"【주간 휴무·OF한도】{nm} 같은 주(일~토)에 OF가 {tot_of_wk}일 — "
                        f"주당 OF는 최대 3일(수간 OF 맞춤 시). 수동으로 조정해 주세요.",
                        _wk_cells_base,
                    )
                elif tot_of_wk == 3:
                    warn(
                        f"{nm} 같은 주(일~토)에 OF 3일 — 월간 수간 OF 개수 맞춤을 위한 완화 구간입니다",
                        _wk_cells_base,
                    )
                oh_vis = pre_oh + sum(1 for d2 in wdays if sh(n, d2) == 'OH')
                no_vis = pre_no + sum(1 for d2 in wdays if sh(n, d2) == 'NO')
                no_week_total = no_vis + post_no
                m = n_prev + len(wdays) + n_next
                d_range = f"{min(wdays)}~{max(wdays)}일"
                if n_prev:
                    d_range = (
                        f"{d_range} (일~토 주에 전월 {n_prev}일 + 당월 {len(wdays)}일, carry 합산)"
                    )
                if n_next > 0:
                    d_range = f"{d_range} · 말주~차월 동일 주(차월 {n_next}일)"
                vis_rest = sum(1 for d2 in wdays if sh(n, d2) in WEEKLY_REST_SHIFTS)
                tot_rest = pre_rest + vis_rest + (post_rest if carry_next_provided else 0)
                carry_prev_ok = _carry_prev_week_tail_complete(carry, n, n_prev)
                if n_prev > 0 and not carry_prev_ok:
                    warn(
                        f"【주간 휴무·carry】{nm} {d_range} — 전월 동일 주({n_prev}일) "
                        f"합산에 필요한 carry_in 꼬리(len≥{n_prev})가 없어 주 2휴무 하드를 "
                        f"검사하지 않습니다. 전월 말 근무를 넘기면 완전 검증됩니다.",
                        _wk_cells_base,
                    )
                elif (
                    m > 0
                    and (n_next == 0 or carry_next_provided)
                    and carry_prev_ok
                    and tot_rest < 2
                ):
                    err(
                        f"【주간 휴무 2일 미달·하드】{nm} {d_range} — "
                        f"휴무 합 {tot_rest}일(필요 ≥2, OF/OH/NO/연/공/병/경)",
                        _wk_cells_base,
                    )
                if no_week_total > 1:
                    warn(
                        f"【주간 휴무·NO】{nm} 주간 NO 초과: {d_range} — "
                        f"같은 주에 NO 최대 1개 권장 (현재 {no_week_total}개). 수동 검토 바랍니다.",
                        [(n, d2) for d2 in wdays if sh(n, d2) == 'NO'],
                    )
                if not _weekly_off_rule_met(
                    of_vis, oh_vis, no_vis, n_prev, len(wdays),
                    post_of, post_oh, post_no, n_next, carry_next_provided,
                ):
                    warn(
                        f"【주간 OF/OH 조합 권고】{nm} {d_range} — "
                        f"OF{of_vis + post_of} OH{oh_vis + post_oh} NO{no_vis + post_no} "
                        f"(평가 {m}일, OF≥2·OH≥2·OF+OH·OF+NO·OH+NO 패턴 권장)",
                        _wk_cells_base,
                    )

    # ── ②b 법적 휴식·공가: N 후 공가는 N-OF-OF-공(OF/OH/NO 2일+), E-공 금지·E-OF-공 허용 — 전원·carry
    _GONG_AFTER_N_OFF = frozenset({'OF', 'OH', 'NO'})
    for n in range(num_nurses):
        nm = names[n]
        cseq = list(carry.get(n, ()))
        full = cseq + [sh(n, d) for d in range(1, NUM_DAYS + 1)]
        for dn in range(1, NUM_DAYS + 1):
            if sh(n, dn) != '공':
                continue
            idx = len(cseq) + dn - 1
            i = idx - 1
            rest_cnt = 0
            while i >= 0 and full[i] in _GONG_AFTER_N_OFF:
                rest_cnt += 1
                i -= 1
            if i >= 0 and full[i] == 'N' and rest_cnt < 2:
                err(
                    f"{nm} 야간(N) 후 공가: OF/OH/NO 휴가 최소 2일 연속 필요(N-OF-OF-공). "
                    f"{dn}일 공가",
                    [(n, dn)],
                )
            if idx > 0 and full[idx - 1] == 'E':
                err(
                    f"{nm} 이브닝(E) 직후 공가 금지(E-공). E-OF-공 등 휴가 1일 후만 허용. {dn}일 공가",
                    [(n, dn)],
                )

    advice = collect_request_advice_warnings(
        schedule,
        num_nurses,
        holidays,
        nurse_names,
        carry_in,
        carry_next_month,
        requests,
    )
    return advice + issues


VIOLATION_CELL_BG_ERROR = '#87CEFA'  # LightSkyBlue — error 급 위반 셀
VIOLATION_CELL_BG_WARN = '#98FB98'  # PaleGreen — 경고 급 위반 셀
VIOLATION_CELL_FG = '#000000'


def merge_validation_cell_highlights(records: list | None) -> dict[tuple, list[dict]]:
    """validate_schedule(..., cell_highlights_out=…) 결과를 (간호사idx, 일) → 메시지 목록으로 병합."""
    if not records:
        return {}
    merged: dict[tuple, list[dict]] = {}
    for rec in records:
        cells = rec.get('cells') or frozenset()
        for c in cells:
            merged.setdefault(c, []).append({
                'level': rec.get('level', 'warn'),
                'msg': rec.get('msg', ''),
            })
    return merged


def cell_highlight_background_for_messages(msgs: list[dict]) -> str:
    """해당 셀에 error가 하나라도 있으면 error 색, 아니면 warn 색."""
    if any(m.get('level') == 'error' for m in msgs):
        return VIOLATION_CELL_BG_ERROR
    return VIOLATION_CELL_BG_WARN


def build_stats(schedule, num_nurses):
    """간호사별 시프트 통계 및 날짜별 인원 통계 계산"""
    nurse_stats = {}
    for n in range(num_nurses):
        shifts = schedule.get(n, {})
        nurse_stats[n] = {
            'N':  sum(1 for v in shifts.values() if v == 'N'),
            'D':  sum(1 for v in shifts.values() if v == 'D'),
            'E':  sum(1 for v in shifts.values() if v == 'E'),
            'OF': sum(1 for v in shifts.values() if v in ('OF', 'OH')),
            'NO': sum(1 for v in shifts.values() if v == 'NO'),
            'A1': sum(1 for v in shifts.values() if v == 'A1'),
        }

    day_stats = {}
    for d in range(1, NUM_DAYS + 1):
        day_stats[d] = {
            'D': sum(1 for n in range(num_nurses) if schedule.get(n, {}).get(d) == 'D'),
            'E': sum(1 for n in range(num_nurses) if schedule.get(n, {}).get(d) == 'E'),
            'N': sum(1 for n in range(num_nurses) if schedule.get(n, {}).get(d) == 'N'),
        }
    return nurse_stats, day_stats


# ── Flask 라우트 ───────────────────────────────────────────────────────────────
@app.route('/', methods=['GET'])
def index():
    days = get_april_days()
    return render_template(
        'index.html',
        days=days,
        shift_names=SHIFT_NAMES,
        shift_colors=SHIFT_COLORS,
        shift_text_colors=SHIFT_TEXT_COLORS,
        num_nurses=10,
        schedule=None,
        nurse_names=get_nurse_names(10),
        holidays=[],
        engine_issues=(),
        period_year=YEAR,
        period_month=MONTH,
        num_days=NUM_DAYS,
    )


@app.route('/generate', methods=['POST'])
def generate():
    global _last_result

    num_nurses = int(request.form.get('num_nurses', 10))
    num_nurses = max(9, min(25, num_nurses))

    # 공휴일 파싱
    holidays_str = request.form.get('holidays', '').strip()
    holidays = []
    if holidays_str:
        for h in holidays_str.split(','):
            try:
                hday = int(h.strip())
                if 1 <= hday <= NUM_DAYS:
                    holidays.append(hday)
            except ValueError:
                pass

    # 개인 신청 근무 파싱 (숨겨진 입력 필드: req_{n_idx}_{day_num})
    requests = {}
    for key, val in request.form.items():
        if key.startswith('req_') and val.strip() in SHIFT_NAMES:
            parts = key.split('_')
            if len(parts) == 3:
                try:
                    n_idx = int(parts[1])
                    day_num = int(parts[2])
                    if 0 <= n_idx < num_nurses and 1 <= day_num <= NUM_DAYS:
                        requests.setdefault(n_idx, {})[day_num] = val.strip()
                except ValueError:
                    pass

    not_available = None
    _na_raw = request.form.get('not_available', '').strip()
    if _na_raw:
        try:
            _na_p = json.loads(_na_raw)
            if isinstance(_na_p, list):
                not_available = _na_p
        except json.JSONDecodeError:
            not_available = None

    pregnant_nurses = None
    _pg_raw = request.form.get('pregnant_nurses', '').strip()
    if _pg_raw:
        try:
            _pg_p = json.loads(_pg_raw)
            if isinstance(_pg_p, list):
                pregnant_nurses = _pg_p
        except json.JSONDecodeError:
            pregnant_nurses = None

    try:
        _sol = solve_schedule(
            num_nurses,
            requests,
            holidays,
            not_available=not_available,
            pregnant_nurses=pregnant_nurses,
        )
        schedule = _sol[0]
        success = _sol[1]
        status_str = _sol[2]
        engine_issues = _sol[3] if len(_sol) > 3 else []
    except Exception as e:
        schedule, success, status_str, engine_issues = None, False, f'예외 발생: {e}', []

    days = get_april_days(holidays)
    nurse_names = get_nurse_names(num_nurses)

    if success:
        nurse_stats, day_stats = build_stats(schedule, num_nurses)
        _last_result = {
            'schedule': schedule,
            'num_nurses': num_nurses,
            'holidays': holidays,
            'nurse_names': nurse_names,
            'engine_issues': engine_issues,
        }
        return render_template(
            'index.html',
            days=days,
            shift_names=SHIFT_NAMES,
            shift_colors=SHIFT_COLORS,
            shift_text_colors=SHIFT_TEXT_COLORS,
            num_nurses=num_nurses,
            schedule=schedule,
            nurse_names=nurse_names,
            nurse_stats=nurse_stats,
            day_stats=day_stats,
            status=status_str,
            holidays=holidays,
            error=None,
            engine_issues=engine_issues,
            period_year=YEAR,
            period_month=MONTH,
            num_days=NUM_DAYS,
        )
    else:
        return render_template(
            'index.html',
            days=days,
            shift_names=SHIFT_NAMES,
            shift_colors=SHIFT_COLORS,
            shift_text_colors=SHIFT_TEXT_COLORS,
            num_nurses=num_nurses,
            schedule=None,
            nurse_names=nurse_names,
            holidays=holidays,
            error=(
                f'해결책을 찾지 못했습니다 ({status_str}). '
                '개인 신청 근무 내용을 줄이거나, 간호사 수를 조정 후 다시 시도해주세요.'
            ),
            engine_issues=(),
            period_year=YEAR,
            period_month=MONTH,
            num_days=NUM_DAYS,
        )


@app.route('/download')
def download():
    global _last_result
    if not _last_result:
        return redirect(url_for('index'))

    schedule = _last_result['schedule']
    num_nurses = _last_result['num_nurses']
    holidays = _last_result['holidays']
    nurse_names = _last_result['nurse_names']
    days = get_april_days(holidays)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{YEAR}년 {MONTH}월 근무표'

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def _xrgb(h: str) -> str:
        return h.replace('#', '').upper()

    EXCEL_BG = {
        sk: (_xrgb(SHIFT_COLORS[sk]), _xrgb(SHIFT_TEXT_COLORS[sk]))
        for sk in SHIFT_COLORS
    }

    N_COL = NUM_DAYS + 2
    OF_COL = NUM_DAYS + 3
    D_COL = NUM_DAYS + 4

    # ─ 타이틀 행
    ws.merge_cells(f'A1:{get_column_letter(D_COL)}1')
    tc = ws['A1']
    tc.value = f'교대근무간호사 {YEAR}년 {MONTH}월 근무표'
    tc.font = Font(bold=True, size=14, color=_xrgb(SHIFT_TEXT_COLORS['N']))
    tc.fill = PatternFill(
        start_color=_xrgb(SHIFT_COLORS['N']), end_color=_xrgb(SHIFT_COLORS['N']), fill_type='solid',
    )
    tc.alignment = center
    ws.row_dimensions[1].height = 28

    # ─ 헤더 행 (날짜)
    ws.cell(row=2, column=1, value='간호사')
    hdr = ws.cell(row=2, column=1)
    hdr.fill = PatternFill(
        start_color=_xrgb(SHIFT_COLORS['OF']), end_color=_xrgb(SHIFT_COLORS['OF']), fill_type='solid',
    )
    hdr.font = Font(bold=True, color=_xrgb(SHIFT_TEXT_COLORS['NO']), size=10)
    hdr.alignment = center
    hdr.border = thin
    ws.column_dimensions['A'].width = 10

    for d, day in enumerate(days):
        col = d + 2
        cell = ws.cell(row=2, column=col, value=f"{day['day']}\n{day['weekday_name']}")
        cell.alignment = center
        cell.border = thin
        if day['is_holiday']:
            bg, tfg = 'FFEBEE', 'C62828'
        elif day['is_weekend']:
            bg, tfg = 'E3F2FD', '1565C0'
        else:
            bg, tfg = 'F5F5F5', '455A64'
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
        cell.font = Font(bold=True, color=tfg, size=9)
        ws.column_dimensions[get_column_letter(col)].width = 4.5

    for col, label, sk in [(N_COL, 'N\n합계', 'N'), (OF_COL, 'OF\n합계', 'OF'), (D_COL, 'D\n합계', 'D')]:
        c = ws.cell(row=2, column=col, value=label)
        c.alignment = center
        c.border = thin
        c.fill = PatternFill(start_color=EXCEL_BG[sk][0], end_color=EXCEL_BG[sk][0], fill_type='solid')
        c.font = Font(bold=True, color=EXCEL_BG[sk][1], size=9)
        ws.column_dimensions[get_column_letter(col)].width = 5.5

    ws.row_dimensions[2].height = 28

    # ─ 간호사 행
    for n_idx, nurse_name in enumerate(nurse_names):
        row = n_idx + 3
        nc = ws.cell(row=row, column=1, value=nurse_name)
        nc.fill = PatternFill(
            start_color=_xrgb(SHIFT_COLORS['OF']), end_color=_xrgb(SHIFT_COLORS['OF']), fill_type='solid',
        )
        nc.font = Font(bold=True, color=_xrgb(SHIFT_TEXT_COLORS['NO']), size=9)
        nc.alignment = center
        nc.border = thin
        ws.row_dimensions[row].height = 18

        nurse_sched = schedule.get(n_idx, {})
        n_cnt = d_cnt = of_cnt = 0

        for d, day in enumerate(days):
            shift = nurse_sched.get(d + 1, '')
            col = d + 2
            cell = ws.cell(row=row, column=col, value=shift)
            cell.alignment = center
            cell.border = thin
            if shift in EXCEL_BG:
                bg, fg = EXCEL_BG[shift]
                cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
                cell.font = Font(color=fg, size=9, bold=True)
            if shift == 'N':
                n_cnt += 1
            if shift == 'D':
                d_cnt += 1
            if shift in ('OF', 'OH', 'NO'):
                of_cnt += 1

        for col, val, sk in [(N_COL, n_cnt, 'N'), (OF_COL, of_cnt, 'OF'), (D_COL, d_cnt, 'D')]:
            bg, fg = EXCEL_BG[sk]
            c = ws.cell(row=row, column=col, value=val)
            c.alignment = center
            c.border = thin
            c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
            c.font = Font(color=fg, bold=True, size=10)

    # ─ 일별 인원 합계 행
    summary_start = len(nurse_names) + 3
    for idx, (label, shift_key) in enumerate([
        ('D 인원', 'D'),
        ('E 인원', 'E'),
        ('N 인원', 'N'),
    ]):
        row = summary_start + idx
        lb, lf = EXCEL_BG[shift_key]
        lc = ws.cell(row=row, column=1, value=label)
        lc.fill = PatternFill(start_color=lb, end_color=lb, fill_type='solid')
        lc.font = Font(bold=True, color=lf, size=9)
        lc.alignment = center
        lc.border = thin
        ws.row_dimensions[row].height = 16

        for d in range(NUM_DAYS):
            day_num = d + 1
            cnt = sum(1 for n in range(num_nurses) if schedule.get(n, {}).get(day_num) == shift_key)
            cell = ws.cell(row=row, column=d + 2, value=cnt)
            cell.alignment = center
            cell.border = thin
            cell.fill = PatternFill(start_color=lb, end_color=lb, fill_type='solid')
            cell.font = Font(bold=True, color=lf, size=9)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'{YEAR}년_{MONTH}월_근무표.xlsx',
    )


if __name__ == '__main__':
    print('=' * 55)
    print(f'  교대근무간호사 {YEAR}년 {MONTH}월 근무표 생성기 시작!')
    print('  브라우저에서 http://127.0.0.1:5000 을 열어주세요')
    print('=' * 55)
    # threaded=True: 계산 중에도 서버가 다른 요청에 응답할 수 있도록 함
    app.run(debug=False, port=5000, threaded=True)