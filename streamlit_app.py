"""
교대근무간호사 근무표 생성기 – Streamlit UI v2
- 명단 행 수(총원)는 수간호사를 포함한다. 예: 간호사 11명 = 수간 1 + 일반 10.
- 연도·월 선택 가능
- 부서(Department) CRUD
- 간호사(Staff) CRUD: 추가 / 이름 수정 / 삭제
- 부서별 신청 근무 입력 달력 (data_editor)
- 함께 근무 불가(2~4명, 선택 D/E/N에 한해 같은 날 같은 근무 동시 배치 금지)
- 임산부 간호사 N 배정 제외(법적)
- 부서별 근무표 생성 + 컬러 테이블 + 엑셀 다운로드
- st.session_state 영속 저장
- 전월 말 근무 이월(JSON) — 월 경계 N-D·연속근무 등
"""

import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import io
import json
import re
import calendar as _calendar
import os
from pathlib import Path

import app as _app                          # 전역 상수(YEAR/MONTH/NUM_DAYS) 동적 갱신
from app import (
    solve_schedule,
    get_april_days,
    validate_schedule,
    SHIFT_NAMES,
    SHIFT_COLORS,
    SHIFT_TEXT_COLORS,
    error_cells_from_validation_issues,
)
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from streamlit_local_storage import LocalStorage as _LocalStorageCls
except ImportError:  # optional; falls back to schedule_requests.json
    _LocalStorageCls = None

# 신청 근무 st.data_editor 전용 드롭다운(생성 근무표의 SHIFT_NAMES와 별개)
REQUEST_SHIFT_OPTIONS: list[str] = ["", "D", "E", "N", "OF", "OH", "NO", "공", "A1", "EDU", "연"]

# ════════════════════════════════════════════════════════════════════════════════
#  페이지 설정
# ════════════════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="교대근무간호사 근무표 생성",
    page_icon="🏥",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ════════════════════════════════════════════════════════════════════════════════
#  전역 CSS
# ════════════════════════════════════════════════════════════════════════════════
st.markdown("""
<style>
/* 브라우저가 폼 컨트롤을 다크모드로 그리지 않게 */
.stApp, section[data-testid="stSidebar"], section[data-testid="stSidebar"] input {
    color-scheme: light !important;
}

.stApp { background-color: #F0F2F6; }

/* ── Duty Solution 최상단 브랜드 헤더 (클로버 로고) ── */
.ds-brand-header {
  background: #FFFFFF !important;
  border-bottom: 1px solid #ddd !important;
  box-sizing: border-box !important;
  min-height: 60px !important;
  height: 60px !important;
  display: flex !important;
  align-items: center !important;
  padding: 10px 20px !important;
  margin: -0.12rem -0.2rem 0.45rem -0.2rem !important;
  width: calc(100% + 0.4rem) !important;
  max-width: none !important;
  position: relative !important;
  z-index: 20 !important;
  font-family: "Pretendard Variable", Pretendard, Inter, -apple-system, BlinkMacSystemFont,
    "Segoe UI", Roboto, "Helvetica Neue", Arial, "Noto Sans KR", sans-serif !important;
}
.ds-brand-header__inner {
  display: flex !important;
  flex-direction: row !important;
  align-items: center !important;
  gap: 10px !important;
}
.ds-brand-header__icon {
  flex-shrink: 0 !important;
  display: flex !important;
  align-items: center !important;
  justify-content: center !important;
  line-height: 0 !important;
}
.ds-brand-header__icon svg {
  display: block !important;
}
.ds-brand-header__title {
  margin: 0 !important;
  padding: 0 !important;
  font-size: 1.2rem !important;
  font-weight: 700 !important;
  letter-spacing: -0.02em !important;
  color: #2D2E83 !important;
  line-height: 1.2 !important;
  -webkit-font-smoothing: antialiased !important;
}

/* 빈 사이드바 숨김 + 메인 가로 전체 사용 (접기 버튼·열 제거) */
section[data-testid="stSidebar"] {
    display: none !important;
    width: 0 !important;
    min-width: 0 !important;
}
[data-testid="stSidebarCollapsedControl"],
[data-testid="collapsedControl"] {
    display: none !important;
}
section[data-testid="stMain"] {
    width: 100% !important;
    max-width: 100% !important;
    margin-left: 0 !important;
}
[data-testid="stAppViewContainer"] {
    padding-left: 0 !important;
    padding-right: 0 !important;
}
header[data-testid="stHeader"] {
    padding-left: 0.35rem !important;
    padding-right: 0.35rem !important;
    padding-top: 0.25rem !important;
    padding-bottom: 0.25rem !important;
}

/* 부서 암호 — 상단 로그인 막대 */
section[data-testid="stMain"] [data-testid="stTextInput"] input[placeholder="부서 암호"] {
    min-height: 34px !important;
    max-width: 100% !important;
    padding: 0.28rem 0.45rem !important;
    font-size: 0.8rem !important;
    line-height: 1.3 !important;
    box-sizing: border-box !important;
    background-color: #ECEFF1 !important;
    border: 1px solid #90A4AE !important;
    border-radius: 5px !important;
}
section[data-testid="stMain"] [data-testid="stHorizontalBlock"]:has(input[placeholder="부서 암호"]) {
    gap: 0.2rem !important;
    row-gap: 0.2rem !important;
}
section[data-testid="stMain"] [data-testid="stHorizontalBlock"]:has(input[placeholder="부서 암호"]) button {
    min-height: 34px !important;
    height: auto !important;
    padding: 0.25rem 0.65rem !important;
    font-size: 0.78rem !important;
    font-weight: 700 !important;
}

/* 메인 영역 — 상하좌우 여백 최소화 */
section[data-testid="stMain"] .block-container {
    max-width: 100% !important;
    padding: 0.06rem 0.2rem 0.2rem 0.2rem !important;
}
section[data-testid="stMain"] [data-testid="stVerticalBlock"] {
    gap: 0.12rem !important;
    row-gap: 0.12rem !important;
}
/* 테두리 패널(로그인·설정): 연한 배경·얇은 여백 */
section[data-testid="stMain"] [data-testid="stVerticalBlockBorderWrapper"] {
    background: #FAFAFA !important;
    border: 1px solid #BDBDBD !important;
    border-radius: 8px !important;
    padding: 0.28rem 0.2rem 0.32rem 0.28rem !important;
    margin-bottom: 0.08rem !important;
}
section[data-testid="stMain"] [data-testid="stVerticalBlockBorderWrapper"] [data-testid="stElementContainer"] {
    margin-bottom: 0 !important;
    margin-top: 0 !important;
}
section[data-testid="stMain"] hr,
section[data-testid="stMain"] [data-testid="stHorizontalRule"] {
    margin: 0.12rem 0 !important;
}
/* st.data_editor / 표 — 전폭·가운데 정렬(모바일에서 편집 입력이 왼쪽으로 쏠리는 현상 완화) */
div[data-testid="stDataFrame"],
div[data-testid="stDataEditor"] {
    width: 100% !important;
    max-width: 100% !important;
    margin: 0 auto !important;
    box-sizing: border-box !important;
}
div[data-testid="stDataEditor"] > div {
    width: 100% !important;
    max-width: 100% !important;
    margin-left: auto !important;
    margin-right: auto !important;
    box-sizing: border-box !important;
}

/* 메인 select — 모바일·좁은 화면 대비 min-height·화살표 우측 여백·16px·세로 가운데 */
section[data-testid="stMain"] [data-testid="stSelectbox"] [data-baseweb="select"] {
    width: 100% !important;
    max-width: 100% !important;
    min-width: 0 !important;
}
section[data-testid="stMain"] [data-testid="stSelectbox"] [data-baseweb="select"] > div {
    background-color: #ffffff !important;
    border: none !important;
    border-radius: 8px !important;
    box-shadow: none !important;
    color: #000000 !important;
    display: flex !important;
    flex-direction: row !important;
    align-items: center !important;
    min-height: 52px !important;
    height: auto !important;
    max-height: none !important;
    padding: 0.65rem 3rem 0.65rem 0.85rem !important;
    box-sizing: border-box !important;
    overflow: visible !important;
    width: 100% !important;
    max-width: 100% !important;
    min-width: 0 !important;
}
section[data-testid="stMain"] [data-testid="stSelectbox"] [data-baseweb="select"] > div > div {
    overflow: visible !important;
    min-height: 0 !important;
    min-width: 0 !important;
    flex: 1 1 auto !important;
    align-self: center !important;
}
section[data-testid="stMain"] [data-testid="stSelectbox"] [data-baseweb="select"] [role="combobox"] {
    color: #0d1117 !important;
    -webkit-text-fill-color: #0d1117 !important;
    display: flex !important;
    flex: 1 1 auto !important;
    align-items: center !important;
    min-height: 44px !important;
    max-height: none !important;
    height: auto !important;
    line-height: 1.5 !important;
    padding: 0 2px 0 0 !important;
    margin: 0 !important;
    overflow: visible !important;
    min-width: 0 !important;
}
section[data-testid="stMain"] [data-testid="stSelectbox"] [data-baseweb="select"] [role="combobox"] p,
section[data-testid="stMain"] [data-testid="stSelectbox"] [data-baseweb="select"] [role="combobox"] span,
section[data-testid="stMain"] [data-testid="stSelectbox"] [data-baseweb="select"] [role="combobox"] div {
    color: #0d1117 !important;
    -webkit-text-fill-color: #0d1117 !important;
    font-weight: 600 !important;
    line-height: 1.5 !important;
    font-size: 16px !important;
    overflow: visible !important;
    margin: 0 !important;
    padding: 0 !important;
    max-height: none !important;
}
section[data-testid="stMain"] [data-testid="stSelectbox"] [data-baseweb="select"] [role="combobox"] p {
    white-space: nowrap !important;
    text-overflow: clip !important;
}
@media (max-width: 768px) {
    section[data-testid="stMain"] [data-testid="stSelectbox"] [data-baseweb="select"] [role="combobox"] p {
        white-space: normal !important;
        word-break: keep-all !important;
        overflow: visible !important;
        text-overflow: unset !important;
    }
}
section[data-testid="stMain"] [data-baseweb="select"] [role="combobox"] p,
section[data-testid="stMain"] [data-baseweb="select"] [role="combobox"] span,
section[data-testid="stMain"] [data-testid="stSelectbox"] [data-baseweb="select"] p {
    color: #000000 !important;
    -webkit-text-fill-color: #000000 !important;
    font-weight: 600 !important;
}
section[data-testid="stMain"] [data-testid="stCheckbox"] label span,
section[data-testid="stMain"] [data-testid="stCheckbox"] label p {
    color: #111111 !important;
}
section[data-testid="stMain"] [data-testid="stExpander"] summary,
section[data-testid="stMain"] [data-testid="stExpander"] summary span {
    color: #212121 !important;
}

/* 상단 설정 패널(부서·연월) — 최대 압축 */
section[data-testid="stMain"] [data-testid="stExpander"] details > summary {
    font-size: 10px !important;
    font-weight: 600 !important;
    padding: 0.02rem 0.22rem !important;
    min-height: 1.35rem !important;
    list-style: none;
}
section[data-testid="stMain"] [data-testid="stExpander"] [data-testid="stVerticalBlock"] {
    gap: 0.35rem !important;
}
section[data-testid="stMain"] [data-testid="stHorizontalBlock"] {
    align-items: center !important;
    gap: 0.06rem !important;
    row-gap: 0.06rem !important;
}
section[data-testid="stMain"] [data-testid="stHorizontalBlock"] > div [data-testid="stSelectbox"] [data-baseweb="select"] > div {
    display: flex !important;
    align-items: center !important;
    min-height: 52px !important;
    height: auto !important;
    max-height: none !important;
    padding: 0.65rem 3rem 0.65rem 0.85rem !important;
    box-sizing: border-box !important;
    overflow: visible !important;
    min-width: 0 !important;
    border: none !important;
    box-shadow: none !important;
}
section[data-testid="stMain"] [data-testid="stHorizontalBlock"] > div div.stButton > button {
    min-height: 50px !important;
    font-size: 11px !important;
    padding: 6px 12px !important;
}
/* 메인 전역 text_input — 높이·패딩 통일 (가로 줄 포함) */
section[data-testid="stMain"] [data-testid="stTextInput"] input:not([type="hidden"]) {
    min-height: 50px !important;
    padding-top: 0.55rem !important;
    padding-bottom: 0.55rem !important;
    padding-left: 0.65rem !important;
    padding-right: 0.65rem !important;
    line-height: 1.45 !important;
    font-size: 0.9rem !important;
    box-sizing: border-box !important;
}

/* 사이드바 — Streamlit CSS 변수(다크 텍스트 색이 입력에 전달되도록) */
section[data-testid="stSidebar"] {
    --text-color: #262730 !important;
    --stTextColor: #262730 !important;
    --widget-text-color: #000000 !important;
}

/* 사이드바 — 흰색 배경 + 선명한 검정 계열 글자 */
section[data-testid="stSidebar"] > div:first-child,
section[data-testid="stSidebar"] [data-testid="stSidebarContent"],
section[data-testid="stSidebar"] [data-testid="stSidebarNavLink"] {
    background: #ffffff !important;
    border-right: 1px solid #e0e0e0;
}
section[data-testid="stSidebar"] {
    color: #212121 !important;
    background-color: #ffffff !important;
}
section[data-testid="stSidebar"] .stMarkdown,
section[data-testid="stSidebar"] .stMarkdown p,
section[data-testid="stSidebar"] .stMarkdown li,
section[data-testid="stSidebar"] .stMarkdown h1,
section[data-testid="stSidebar"] .stMarkdown h2,
section[data-testid="stSidebar"] .stMarkdown h3,
section[data-testid="stSidebar"] .stMarkdown h4 {
    color: #212121 !important;
}
section[data-testid="stSidebar"] hr { border-color: #e0e0e0 !important; margin:0.6rem 0; }

/* 연도·월 2열 — 한 겹 겹친 느낌 제거(그림자·반투명·블러 없음) */
section[data-testid="stSidebar"] [data-testid="stVerticalBlock"],
section[data-testid="stSidebar"] [data-testid="stHorizontalBlock"] {
    opacity: 1 !important;
    filter: none !important;
    box-shadow: none !important;
}
section[data-testid="stSidebar"] [data-testid="stVerticalBlock"] > [data-testid="stElementContainer"],
section[data-testid="stSidebar"] [data-testid="stHorizontalBlock"] > [data-testid="stElementContainer"] {
    opacity: 1 !important;
}
section[data-testid="stSidebar"] [data-testid="column"] {
    background: transparent !important;
    box-shadow: none !important;
}
section[data-testid="stSidebar"] [data-testid="stElementContainer"] {
    box-shadow: none !important;
}

section[data-testid="stSidebar"] [data-testid="stSelectbox"] label p,
section[data-testid="stSidebar"] [data-testid="stSelectbox"] label span,
section[data-testid="stSidebar"] .stTextInput label p,
section[data-testid="stSidebar"] .stTextInput label span {
    color: #000000 !important;
    font-weight: 700 !important;
    font-size: 14px !important;
    opacity: 1 !important;
    -webkit-font-smoothing: antialiased;
}

section[data-testid="stSidebar"] [data-testid="stTextInput"] input,
section[data-testid="stSidebar"] [data-testid="stTextInput"] textarea,
section[data-testid="stSidebar"] [data-testid="stExpander"] [data-testid="stTextInput"] input,
section[data-testid="stSidebar"] [data-testid="stExpander"] [data-testid="stTextInput"] textarea {
    background: #fafafa !important;
    border: 1px solid #bdbdbd !important;
    color: #0d1117 !important;
    -webkit-text-fill-color: #0d1117 !important;
    caret-color: #0d1117 !important;
    opacity: 1 !important;
    border-radius: 8px;
    min-height: 50px !important;
    padding-top: 0.55rem !important;
    padding-bottom: 0.55rem !important;
    line-height: 1.45 !important;
    font-size: 0.9rem !important;
    box-sizing: border-box !important;
}
section[data-testid="stSidebar"] [data-testid="stTextInput"] input::placeholder,
section[data-testid="stSidebar"] [data-testid="stTextInput"] textarea::placeholder,
section[data-testid="stSidebar"] [data-testid="stExpander"] [data-testid="stTextInput"] input::placeholder {
    color: #616161 !important;
    -webkit-text-fill-color: #616161 !important;
    opacity: 1 !important;
}

/* 모든 사이드바 텍스트 입력 — Streamlit은 .stTextInput + data-baseweb="input" 조합 사용 */
section[data-testid="stSidebar"] .stTextInput input,
section[data-testid="stSidebar"] .stTextInput textarea,
section[data-testid="stSidebar"] [data-baseweb="input"] input,
section[data-testid="stSidebar"] [data-baseweb="textarea"] textarea {
    color: #000000 !important;
    -webkit-text-fill-color: #000000 !important;
    caret-color: #000000 !important;
    background-color: #ffffff !important;
    opacity: 1 !important;
}
section[data-testid="stSidebar"] .stTextInput input:-webkit-autofill,
section[data-testid="stSidebar"] [data-baseweb="input"] input:-webkit-autofill {
    -webkit-text-fill-color: #000000 !important;
    caret-color: #000000 !important;
    box-shadow: 0 0 0px 1000px #ffffff inset !important;
}

/* Expander 안 입력(새 부서 추가 등) — 추가 보강 */
section[data-testid="stSidebar"] [data-testid="stExpander"] input:not([role="combobox"]),
section[data-testid="stSidebar"] [data-testid="stExpanderDetails"] input:not([role="combobox"]),
section[data-testid="stSidebar"] [data-testid="stExpander"] textarea,
section[data-testid="stSidebar"] [data-testid="stExpanderDetails"] textarea {
    color: #000000 !important;
    -webkit-text-fill-color: #000000 !important;
    caret-color: #000000 !important;
    background-color: #ffffff !important;
    opacity: 1 !important;
}

/* 사이드바 내 모든 텍스트형 input (타입 미지정 포함) — 콤보박스 제외 */
section[data-testid="stSidebar"] input[type="text"],
section[data-testid="stSidebar"] input[type="search"],
section[data-testid="stSidebar"] input:not([type]):not([role="combobox"]) {
    color: #000000 !important;
    -webkit-text-fill-color: #000000 !important;
    caret-color: #000000 !important;
    background-color: #ffffff !important;
}

/* 앱 전역: text_input 위젯은 항상 검정 글자 (사이드바 DOM 분리 대비) */
div[data-testid="stTextInput"] input,
.stTextInput input {
    color: #000000 !important;
    -webkit-text-fill-color: #000000 !important;
    caret-color: #000000 !important;
}

/* select — 순배경(회색끔 제거) + 순검정 글자, 덧씌운 느낌 나는 그림자 제거 */
section[data-testid="stSidebar"] [data-testid="stSelectbox"] [data-baseweb="select"] > div {
    background-color: #ffffff !important;
    background-image: none !important;
    border: none !important;
    border-radius: 8px !important;
    box-shadow: none !important;
    opacity: 1 !important;
    color: #111111 !important;
    display: flex !important;
    align-items: center !important;
    min-height: 52px !important;
    height: auto !important;
    max-height: none !important;
    padding: 0.65rem 3rem 0.65rem 0.85rem !important;
    box-sizing: border-box !important;
    overflow: visible !important;
}
section[data-testid="stSidebar"] [data-baseweb="select"] > div {
    background-color: #ffffff !important;
    background-image: none !important;
    border: none !important;
    border-radius: 8px !important;
    box-shadow: none !important;
    opacity: 1 !important;
    color: #111111 !important;
    display: flex !important;
    align-items: center !important;
    min-height: 52px !important;
    height: auto !important;
    max-height: none !important;
    padding: 0.65rem 3rem 0.65rem 0.85rem !important;
    box-sizing: border-box !important;
    overflow: visible !important;
}
section[data-testid="stSidebar"] [data-baseweb="select"] [role="combobox"] {
    color: #0d1117 !important;
    font-weight: 600 !important;
    font-size: 16px !important;
    line-height: 1.5 !important;
    -webkit-font-smoothing: antialiased;
    display: flex !important;
    align-items: center !important;
    flex: 1 1 auto !important;
    min-height: 44px !important;
    min-width: 0 !important;
    overflow: visible !important;
    padding: 0 2px 0 0 !important;
}
section[data-testid="stSidebar"] [data-baseweb="select"] > div > div {
    color: #000000 !important;
}
section[data-testid="stSidebar"] [data-baseweb="select"] [role="combobox"] span,
section[data-testid="stSidebar"] [data-baseweb="select"] [role="combobox"] div,
section[data-testid="stSidebar"] [data-testid="stSelectbox"] [data-baseweb="select"] p {
    color: #000000 !important;
    -webkit-text-fill-color: #000000 !important;
}
/* 사이드바 셀렉트: 좁은 열에서 '간…' 말줄임 방지 — 전체 이름 표시 */
section[data-testid="stSidebar"] [data-testid="stSelectbox"] [data-baseweb="select"] > div {
    min-width: 0 !important;
    width: 100% !important;
    max-width: 100% !important;
    display: flex !important;
    align-items: center !important;
    min-height: 52px !important;
    height: auto !important;
    max-height: none !important;
    padding: 0.65rem 3rem 0.65rem 0.85rem !important;
    box-sizing: border-box !important;
    overflow: visible !important;
}
section[data-testid="stSidebar"] [data-testid="stSelectbox"] [data-baseweb="select"] [role="combobox"] p {
    white-space: normal !important;
    overflow: visible !important;
    text-overflow: clip !important;
}
/* 드롭다운 목록 항목 — multiselect/select 공통 가독성 */
div[data-baseweb="popover"] li[role="option"] {
    white-space: normal !important;
    overflow: visible !important;
    text-overflow: clip !important;
    font-size: 15px !important;
    line-height: 1.4 !important;
    min-height: 40px !important;
    padding: 8px 14px !important;
    color: #111111 !important;
    -webkit-text-fill-color: #111111 !important;
}

section[data-testid="stSidebar"] [data-testid="stExpander"] summary,
section[data-testid="stSidebar"] [data-testid="stExpander"] summary span,
section[data-testid="stSidebar"] [data-testid="stExpander"] summary p {
    color: #212121 !important;
}
section[data-testid="stSidebar"] .stCaption { color: #616161 !important; }

/* 사이드바 체크박스 라벨 — 테마에 밝은 글자가 묻는 경우 방지 */
section[data-testid="stSidebar"] [data-testid="stCheckbox"] label span,
section[data-testid="stSidebar"] [data-testid="stCheckbox"] label p {
    color: #111111 !important;
}

/* 펼친 드롭다운 목록은 밝은 배경 + 검정 글자 (포털로 body에 그려질 수 있음) */
div[data-baseweb="popover"] ul[role="listbox"] li,
div[data-baseweb="popover"] li[role="option"] {
    color: #111111 !important;
    -webkit-text-fill-color: #111111 !important;
    background-color: #ffffff !important;
}
div[data-baseweb="popover"] ul[role="listbox"] {
    background-color: #ffffff !important;
}
/* 플레이스홀더 — selectbox 전용(multiselect는 라벨만 사용) */
section[data-testid="stMain"] [data-testid="stSelectbox"] [data-baseweb="select"] input::placeholder {
    color: #212121 !important;
    opacity: 1 !important;
    font-size: 16px !important;
    font-weight: 500 !important;
    -webkit-text-fill-color: #212121 !important;
}
/* multiselect(Choose options) 패널 전체 흰 배경 — 메인·사이드바 공통 */
div[data-baseweb="popover"] {
    background-color: #ffffff !important;
    box-shadow: 0 4px 16px rgba(0,0,0,0.12) !important;
}
div[data-baseweb="popover"] [data-baseweb="menu"],
div[data-baseweb="popover"] ul {
    background-color: #ffffff !important;
}
/* st.selectbox·st.multiselect — 래퍼 흰 배경·회색 테두리·포커스 인디고 (헤더 stHeader 미적용) */
section[data-testid="stMain"] [data-testid="stSelectbox"] [data-baseweb="select"],
section[data-testid="stMain"] [data-testid="stMultiSelect"] [data-baseweb="select"],
section[data-testid="stSidebar"] [data-testid="stSelectbox"] [data-baseweb="select"],
section[data-testid="stSidebar"] [data-testid="stMultiSelect"] [data-baseweb="select"] {
    background-color: #ffffff !important;
    border: 1px solid #d1d5db !important;
    border-radius: 8px !important;
    box-sizing: border-box !important;
}
section[data-testid="stMain"] [data-testid="stSelectbox"] [data-baseweb="select"]:focus-within,
section[data-testid="stMain"] [data-testid="stMultiSelect"] [data-baseweb="select"]:focus-within,
section[data-testid="stSidebar"] [data-testid="stSelectbox"] [data-baseweb="select"]:focus-within,
section[data-testid="stSidebar"] [data-testid="stMultiSelect"] [data-baseweb="select"]:focus-within {
    border-color: #4f46e5 !important;
    box-shadow: 0 0 0 1px #4f46e5 !important;
}
section[data-testid="stMain"] [data-testid="stMultiSelect"] [data-baseweb="select"] > div,
section[data-testid="stSidebar"] [data-testid="stMultiSelect"] [data-baseweb="select"] > div {
    border: none !important;
    box-shadow: none !important;
}
/* 익스팬더 안 multiselect 겹침 방지 */
section[data-testid="stMain"] [data-testid="stExpanderDetails"] {
    overflow: visible !important;
}
.fp-multiselect-anchor {
    height: 10px;
    min-height: 10px;
    display: block;
}
/* 메인 multiselect — 위젯 간격·라벨 강조(플레이스홀더 미사용·상자는 기본 스타일) */
section[data-testid="stMain"] [data-testid="stMultiSelect"] {
    margin-top: 6px !important;
    margin-bottom: 12px !important;
    position: relative;
    z-index: 1;
    width: 100% !important;
    max-width: 100% !important;
    min-width: 0 !important;
}
section[data-testid="stMain"] [data-testid="stMultiSelect"] label p,
section[data-testid="stMain"] [data-testid="stMultiSelect"] label span,
section[data-testid="stMain"] [data-testid="stMultiSelect"] [data-testid="stWidgetLabel"] p,
section[data-testid="stMain"] [data-testid="stMultiSelect"] [data-testid="stWidgetLabel"] span {
    font-size: 1.05rem !important;
    font-weight: 700 !important;
    color: #0d1117 !important;
    -webkit-text-fill-color: #0d1117 !important;
}
@media (max-width: 768px) {
    section[data-testid="stMain"] [data-testid="stMultiSelect"] label p,
    section[data-testid="stMain"] [data-testid="stMultiSelect"] label span,
    section[data-testid="stMain"] [data-testid="stMultiSelect"] [data-testid="stWidgetLabel"] p,
    section[data-testid="stMain"] [data-testid="stMultiSelect"] [data-testid="stWidgetLabel"] span {
        font-size: 1.12rem !important;
    }
}
/* multiselect 선택 태그(칩) — 연회색·검정 글자 */
section[data-testid="stMain"] [data-testid="stMultiSelect"] [data-baseweb="tag"] {
    font-size: 14px !important;
    line-height: 1.4 !important;
    padding: 4px 6px 4px 8px !important;
    margin: 2px 4px 2px 0 !important;
    border-radius: 6px !important;
    max-width: calc(100% - 6px) !important;
    box-sizing: border-box !important;
    background-color: #f0f2f6 !important;
    background-image: none !important;
    border: 1px solid #e0e4eb !important;
    color: #0d1117 !important;
    -webkit-text-fill-color: #0d1117 !important;
    box-shadow: none !important;
    align-items: center !important;
    display: inline-flex !important;
    flex-wrap: nowrap !important;
    gap: 4px !important;
    min-height: 0 !important;
}
section[data-testid="stMain"] [data-testid="stMultiSelect"] [data-baseweb="tag"] span {
    font-size: 14px !important;
    line-height: 1.4 !important;
    color: #0d1117 !important;
    -webkit-text-fill-color: #0d1117 !important;
    font-weight: 600 !important;
    white-space: nowrap !important;
    overflow: visible !important;
    text-overflow: clip !important;
}
section[data-testid="stMain"] [data-testid="stMultiSelect"] [data-baseweb="tag"] svg,
section[data-testid="stMain"] [data-testid="stMultiSelect"] [data-baseweb="tag"] svg path {
    fill: #424242 !important;
    color: #424242 !important;
}
section[data-testid="stMain"] [data-testid="stMultiSelect"] [data-baseweb="tag"] button,
section[data-testid="stMain"] [data-testid="stMultiSelect"] [data-baseweb="tag"] [role="button"] {
    color: #424242 !important;
    background: transparent !important;
    opacity: 1 !important;
    padding: 2px !important;
    min-width: 22px !important;
    min-height: 22px !important;
    flex-shrink: 0 !important;
}

/* 사이드바 selectbox 보강 */
section[data-testid="stSidebar"] [data-testid="stSelectbox"] [data-baseweb="select"] [role="combobox"] svg {
    fill: #111111 !important;
}

/* 버튼 */
div.stButton > button[kind="primary"] {
    background:#2E7D32; border:none; border-radius:10px;
    font-size:15px; font-weight:700; height:52px;
}
div.stButton > button[kind="primary"]:hover { background:#1B5E20; }

/* 작은 보조 버튼 */
div.stButton > button[kind="secondary"] {
    border-radius:6px; font-size:12px; padding:2px 8px;
}

/* 섹션 카드 */
.card {
    background:white; border-radius:12px; padding:16px 20px;
    box-shadow:0 2px 10px rgba(0,0,0,0.07); margin-bottom:14px;
}
.card-title { font-size:17px; font-weight:800; color:#1A237E; margin-bottom:6px; }
.card-sub   { font-size:12px; color:#546E7A; margin-top:2px; }

/* 배지 */
.dept-badge {
    display:inline-block; background:#E8EAF6; color:#1A237E;
    border-radius:20px; padding:3px 12px; font-size:12px; font-weight:600; margin-right:4px;
}

/* 신청·생성 근무 data_editor — 이름 열 + 1~말일 한 화면 (가로 균등·날짜 헤더 세로) */
section[data-testid="stMain"] div[data-testid="stDataFrame"] table {
    table-layout: fixed !important;
    width: 100% !important;
    max-width: 100% !important;
}
section[data-testid="stMain"] div[data-testid="stDataFrame"] thead th:not(:first-child) {
    writing-mode: vertical-rl !important;
    text-orientation: mixed !important;
    transform: rotate(180deg);
    font-size: 6px !important;
    font-weight: 700 !important;
    padding: 0 1px !important;
    height: 2.65em !important;
    line-height: 1 !important;
    vertical-align: middle !important;
}
section[data-testid="stMain"] div[data-testid="stDataFrame"] thead th:first-child {
    width: 13% !important;
    min-width: 7.25rem !important;
    max-width: none !important;
    font-size: 8px !important;
    padding: 1px 4px !important;
    line-height: 1.15 !important;
    vertical-align: bottom !important;
    overflow: visible !important;
    white-space: normal !important;
    word-break: keep-all !important;
}
section[data-testid="stMain"] div[data-testid="stDataFrame"] td,
section[data-testid="stMain"] div[data-testid="stDataFrame"] tbody th {
    font-size: 7px !important;
    padding: 0 !important;
    line-height: 1.05 !important;
}
section[data-testid="stMain"] div[data-testid="stDataFrame"] td:first-child,
section[data-testid="stMain"] div[data-testid="stDataFrame"] tbody th:first-child {
    width: 13% !important;
    min-width: 7.25rem !important;
    max-width: none !important;
    font-size: 8px !important;
    padding: 1px 4px !important;
    white-space: normal !important;
    word-break: keep-all !important;
    overflow: visible !important;
    overflow-x: visible !important;
    overflow-y: visible !important;
    text-overflow: unset !important;
}
/* 이름 셀 내부 래퍼 — 가로·세로 스크롤바 없이 전체 표시 */
section[data-testid="stMain"] div[data-testid="stDataFrame"] td:first-child > div,
section[data-testid="stMain"] div[data-testid="stDataFrame"] tr td:first-child [data-testid="cell"] {
    overflow: visible !important;
    overflow-x: visible !important;
    overflow-y: visible !important;
    max-height: none !important;
    white-space: normal !important;
    word-break: keep-all !important;
}
section[data-testid="stMain"] div[data-testid="stDataFrame"] [data-baseweb="select"] > div {
    font-size: 7px !important;
    min-height: 15px !important;
    padding: 0 1px !important;
}
section[data-testid="stMain"] div[data-testid="stDataFrame"] [data-baseweb="select"] [role="combobox"],
section[data-testid="stMain"] div[data-testid="stDataFrame"] [data-baseweb="select"] p {
    font-size: 7px !important;
    line-height: 1.05 !important;
}
section[data-testid="stMain"] div[data-testid="stDataFrame"] [data-baseweb="popover"] li,
section[data-testid="stMain"] div[data-testid="stDataFrame"] ul[role="listbox"] li {
    font-size: 8px !important;
    min-height: 18px !important;
    padding: 1px 5px !important;
}
/* 내부 가로 스크롤 최소화(균등 분배 우선) */
section[data-testid="stMain"] [data-testid="stDataEditor"] > div [data-testid="stHorizontalBlock"],
section[data-testid="stMain"] [data-testid="stDataFrame"] {
    max-width: 100% !important;
}

/* 신청 근무 확정 박스 — 본문 글자 검정 (테마 간섭 방지) */
.req-save-panel, .req-save-panel h4, .req-save-panel p,
.req-save-status { color: #111111 !important; }

/* 생성된 근무표(HTML 미리보기) — 가로 스크롤(말일·합계 열까지) */
section[data-testid="stMain"] .duty-generated-schedule-wrap {
    overflow-x: scroll !important;
    overflow-y: hidden;
    width: 100% !important;
    max-width: 100% !important;
    min-width: 0 !important;
    box-sizing: border-box !important;
    -webkit-overflow-scrolling: touch;
    scrollbar-gutter: stable;
}
section[data-testid="stMain"] .duty-generated-schedule-wrap table {
    width: max-content !important;
    min-width: unset !important;
    max-width: none !important;
    table-layout: auto !important;
}
section[data-testid="stMain"] .duty-generated-schedule-wrap th:first-child,
section[data-testid="stMain"] .duty-generated-schedule-wrap td:first-child {
    min-width: 80px !important;
    padding: 5px 8px !important;
    font-size: 11px !important;
    box-sizing: border-box !important;
}
section[data-testid="stMain"] [data-testid="stMarkdownContainer"]:has(.duty-generated-schedule-wrap),
section[data-testid="stMain"] [data-testid="stElementContainer"]:has(.duty-generated-schedule-wrap) {
    overflow-x: visible !important;
    max-width: 100% !important;
}
/* 생성 근무표 편집 data_editor — 바로 다음 블록에 가로 스크롤 허용 */
section[data-testid="stMain"] [data-testid="stElementContainer"]:has(.duty-schedule-editor-hscroll)
    + [data-testid="stElementContainer"] [data-testid="stDataFrame"] {
    overflow-x: auto !important;
    max-width: 100% !important;
}
section[data-testid="stMain"] [data-testid="stElementContainer"]:has(.duty-schedule-editor-hscroll)
    + [data-testid="stElementContainer"] [data-testid="stDataFrame"] > div {
    overflow-x: auto !important;
    max-width: 100% !important;
}
</style>
""", unsafe_allow_html=True)

# ════════════════════════════════════════════════════════════════════════════════
#  Session State 초기화
# ════════════════════════════════════════════════════════════════════════════════
def _default_nurses(n: int = 9) -> list[str]:
    return ["수간호사"] + [f"간호사{i}" for i in range(1, n + 1)]


# 부서 상수: 설정 누락·로드 실패 시에도 세션·셀렉트에 반드시 노출
_ER_DEPT_NAME = "응급실"


def _er_department_hospital_row() -> dict:
    """`hospital_config.json`의 응급실 블록: 총 10명(수간+9), 평일 E2/N2·수간 A1 시 일반간 D=1(절대), 그 외 평일 D2 등(app `er`)."""
    return {
        "nurses": _default_nurses(9),
        "general_code": "er1004",
        "admin_code": "er777",
        "unit_profile": "er",
        "rule_note": "E2/N2, 평일 수간 A1 시 일반간 D=1(절대)·그 외 D2 등 — unit_profile er, 총원 10",
    }


def _ordered_dept_keys(depts: dict) -> list[str]:
    """응급실을 목록 맨 앞에 두어 selectbox에서 최우선 노출."""
    keys = list(depts.keys())
    if _ER_DEPT_NAME in keys:
        return [_ER_DEPT_NAME] + [k for k in keys if k != _ER_DEPT_NAME]
    return keys


def _primary_dept_key(depts: dict) -> str:
    od = _ordered_dept_keys(depts)
    return od[0] if od else _ER_DEPT_NAME


def _bundle_ensure_emergency_room(bundle: dict | None) -> None:
    """로드된 bundle에 응급실·er 메타가 없거나 잘못된 경우 보강."""
    if bundle is None:
        return
    dep = bundle.get("departments")
    if not isinstance(dep, dict):
        return
    meta = bundle.setdefault("dept_meta", {})
    if not isinstance(meta, dict):
        bundle["dept_meta"] = {}
        meta = bundle["dept_meta"]
    row = _er_department_hospital_row()
    if _ER_DEPT_NAME not in dep or not isinstance(dep.get(_ER_DEPT_NAME), list) or len(dep[_ER_DEPT_NAME]) < 2:
        dep[_ER_DEPT_NAME] = list(row["nurses"])
        meta[_ER_DEPT_NAME] = _default_dept_meta(
            "er", row["general_code"], row["admin_code"], row["rule_note"]
        )
        return
    em = meta.get(_ER_DEPT_NAME)
    if not isinstance(em, dict):
        meta[_ER_DEPT_NAME] = _default_dept_meta(
            "er", row["general_code"], row["admin_code"], row["rule_note"]
        )
        return
    up = str(em.get("unit_profile") or "").strip().lower()
    if up != "er":
        fix = dict(em)
        fix["unit_profile"] = "er"
        fix.setdefault("general_code", row["general_code"])
        fix.setdefault("admin_code", row["admin_code"])
        fix.setdefault("rule_note", row["rule_note"])
        meta[_ER_DEPT_NAME] = fix


def _repair_hospital_config_file_emergency_dept() -> None:
    """JSON에 응급실 키가 없을 때만 기본 블록을 끼워 넣어 저장(기존 부서·코드는 건드리지 않음)."""
    if not _HOSPITAL_CONFIG_PATH.is_file():
        return
    try:
        with open(_HOSPITAL_CONFIG_PATH, encoding="utf-8") as f:
            data = json.load(f)
    except (OSError, json.JSONDecodeError, TypeError, ValueError):
        return
    if not isinstance(data, dict):
        return
    depts = data.get("departments")
    if not isinstance(depts, dict):
        depts = {}
        data["departments"] = depts
    if _ER_DEPT_NAME in depts:
        return
    row = _er_department_hospital_row()
    depts[_ER_DEPT_NAME] = {k: v for k, v in row.items()}
    try:
        _atomic_write_json(_HOSPITAL_CONFIG_PATH, data)
    except OSError:
        pass


def _ensure_emergency_department_session_state() -> None:
    """로드 실패·구 세션 등으로 빠진 응급실을 세션에 복구."""
    st.session_state.setdefault("departments", {})
    depts = st.session_state.departments
    if not isinstance(depts, dict):
        st.session_state.departments = {}
        depts = st.session_state.departments
    meta = st.session_state.setdefault("dept_meta", {})
    row = _er_department_hospital_row()
    if _ER_DEPT_NAME not in depts or not isinstance(depts.get(_ER_DEPT_NAME), list) or len(depts[_ER_DEPT_NAME]) < 2:
        depts[_ER_DEPT_NAME] = list(row["nurses"])
    else:
        _er_n = _clean_nurse_names_list(depts[_ER_DEPT_NAME])
        if not _er_n:
            _er_n = ["수간호사"]
        if _er_n != depts[_ER_DEPT_NAME]:
            depts[_ER_DEPT_NAME] = _er_n
    if _ER_DEPT_NAME not in meta or not isinstance(meta.get(_ER_DEPT_NAME), dict):
        meta[_ER_DEPT_NAME] = _default_dept_meta(
            "er", row["general_code"], row["admin_code"], row["rule_note"]
        )
    else:
        em = meta[_ER_DEPT_NAME]
        up = str(em.get("unit_profile") or "").strip().lower()
        if up != "er":
            fix = dict(em)
            fix["unit_profile"] = "er"
            fix.setdefault("general_code", row["general_code"])
            fix.setdefault("admin_code", row["admin_code"])
            fix.setdefault("rule_note", row["rule_note"])
            meta[_ER_DEPT_NAME] = fix


# 총원(수간 포함) — hospital_config 기본 시드·플레이스홀더 보강·schedule_cpsat 상수와 동기
DEFAULT_DEPT_TOTAL_HEADCOUNT: dict[str, int] = {
    "응급실": 10,
    "신관 3병동": 12,
    "본관 5병동": 12,
    "본관 6병동": 12,
    "본관 7병동": 12,
    "본관 8병동": 11,
    "중환자실": 22,
}


def _is_blank_nurse_name(x: object) -> bool:
    """명단에서 제거할 셀: None, NaN, 공백, 문자열 'none'/'nan' 등."""
    if x is None:
        return True
    try:
        if pd.api.types.is_scalar(x) and pd.isna(x):
            return True
    except (TypeError, ValueError):
        pass
    s = str(x).strip()
    if not s:
        return True
    low = s.lower()
    if low in ("none", "nan", "<na>", "nat"):
        return True
    return False


def _clean_nurse_names_list(items: list | None) -> list[str]:
    """하위 신청·미리보기·세션에 넘길 '진짜 명단'만 순서 유지로 추출."""
    if not items:
        return []
    out: list[str] = []
    for x in items:
        if _is_blank_nurse_name(x):
            continue
        out.append(str(x).strip())
    return out


def _all_nurse_names_placeholder_like(nurses: list[str]) -> bool:
    """실명이 없고 수간/간호사n 자리만 있으면 True (정원까지 자동 채움 가능)."""
    for i, nm in enumerate(nurses):
        s = (str(nm).strip() if nm is not None else "")
        if i == 0:
            if s and s != "수간호사":
                return False
        else:
            if s and not re.fullmatch(r"간호사\d+", s):
                return False
    return True


def _extend_nurses_to_dept_headcount(dept_name: str, nurses: list[str]) -> list[str]:
    target = DEFAULT_DEPT_TOTAL_HEADCOUNT.get(dept_name)
    if target is None or len(nurses) >= target:
        return list(nurses)
    if not _all_nurse_names_placeholder_like(nurses):
        return list(nurses)
    out = list(nurses)
    if not out:
        return _default_nurses(target - 1)
    idx = len(out)
    while len(out) < target:
        out.append(f"간호사{idx}")
        idx += 1
    return out


def _default_dept_meta(
    unit_profile: str = "ward",
    general_code: str = "",
    admin_code: str = "",
    rule_note: str = "",
) -> dict:
    up = unit_profile if unit_profile in ("icu", "er", "ward") else "ward"
    out = {
        "general_code": str(general_code or "").strip(),
        "admin_code": str(admin_code or "").strip(),
        "unit_profile": up,
    }
    if rule_note:
        out["rule_note"] = str(rule_note).strip()
    return out


def _dept_login_secrets(dm: dict | None) -> frozenset[str]:
    """로그인에 쓸 수 있는 값들 — hospital_config의 여러 필드 중 비어 있지 않은 것 전부(예: general·admin 동시 설정 시 둘 다 허용)."""
    if not isinstance(dm, dict):
        return frozenset()
    out: set[str] = set()
    for _k in ("dept_password", "general_code", "access_code", "admin_code"):
        v = str(dm.get(_k) or "").strip()
        if v:
            out.add(v)
    return frozenset(out)


def _default_hospital_config_payload() -> dict:
    """
    최초 설치용 hospital_config 시드.
    ICU·ER·ward 인원 규칙은 app.unit_profile(cp-sat/검증)과 일치하도록 unit_profile로 박제.
    """
    wards = ("본관 5병동", "본관 6병동", "본관 7병동", "본관 8병동")
    ward_codes_g = ("m51004", "m61004", "m71004", "m81004")
    ward_codes_a = ("m5777", "m6777", "m7777", "m8777")
    departments: dict = {
        "중환자실": {
            "nurses": _default_nurses(21),
            "general_code": "icu1004",
            "admin_code": "icu777",
            "unit_profile": "icu",
            "rule_note": "D4/E4/N3, A1 미차감형 — unit_profile icu, 총원 22",
        },
        _ER_DEPT_NAME: dict(_er_department_hospital_row()),
        "신관 3병동": {
            "nurses": _default_nurses(11),
            "general_code": "n31004",
            "admin_code": "n3777",
            "unit_profile": "ward",
            "rule_note": "D2~3/E2/N2 — ward, 총원 12",
        },
    }
    for wn, cg, ca in zip(wards, ward_codes_g, ward_codes_a):
        n_tail = 10 if wn == "본관 8병동" else 11
        departments[wn] = {
            "nurses": _default_nurses(n_tail),
            "general_code": cg,
            "admin_code": ca,
            "unit_profile": "ward",
            "rule_note": (
                "D2~3/E2/N2, A1 미차감형 — ward, 총원 11"
                if wn == "본관 8병동"
                else "D2~3/E2/N2, A1 미차감형 — ward, 총원 12"
            ),
        }
    return {
        "version": 1,
        "active_dept": "중환자실",
        "departments": departments,
        "forbidden_pairs": {},
        "pregnant_nurses": {},
        "n_max4_nurses": {},
        "dept_holidays": {},
    }


# 부서·간호사·보안·규칙 영속 저장: 우선 hospital_config.json, 없으면 기존 user_departments.json
_HOSPITAL_CONFIG_PATH = Path(__file__).resolve().parent / "hospital_config.json"
_DEPT_SAVE_PATH = Path(__file__).resolve().parent / "user_departments.json"
_SCHEDULE_ARCHIVE_PATH = Path(__file__).resolve().parent / "schedule_month_archive.json"
# 신청 근무: 브라우저 localStorage 키(우선). 서버 JSON은 마이그레이션·오프라인 백업용.
_SCHEDULE_REQUESTS_PATH = Path(__file__).resolve().parent / "schedule_requests.json"
# 이전 기록 불러오기: 키 `{부서}_{연도}_{월}` 스냅샷 (💾 저장 형식과 동일 nurse_names/columns/data)
_SHIFT_REQUESTS_PATH = Path(__file__).resolve().parent / "shift_requests.json"
_LS_COMPONENT_STATE_KEY = "duty_solution_ls_component_v1"
_LS_ARCHIVE_ITEM_KEY = "DutySolution.schedule_requests.v1"
_LS_CARRY_ITEM_KEY_RY = "DutySolution.carry_over_by_session_key.v1"
CARRY_AUTO_DAYS = 7
# False: 「직전 달 마지막 N일 자동」버튼 비표시·비실행 (전월 이월은 JSON 수기만)
CARRY_AUTO_FROM_ARCHIVE_ENABLED = False


def _atomic_write_json(path: Path, payload: dict) -> None:
    tmp = path.with_name(path.name + ".tmp")
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    tmp.replace(path)


def _atomic_write_json_safe(path: Path, payload: dict) -> tuple[bool, str | None]:
    """hospital_config 등 JSON 원자 기록. 실패 시 (False, 메시지) 반환 — 조용히 무시하지 않음."""
    try:
        _atomic_write_json(path, payload)
        return True, None
    except OSError as e:
        return False, f"OSError: {e}"
    except (TypeError, ValueError) as e:
        return False, f"JSON 직렬화 불가: {e}"


def _ensure_hospital_config_file() -> None:
    """hospital_config.json 이 없으면 부서·코드·unit_profile 기본값으로 생성."""
    if _HOSPITAL_CONFIG_PATH.is_file():
        return
    try:
        _atomic_write_json(_HOSPITAL_CONFIG_PATH, _default_hospital_config_payload())
    except OSError:
        pass


_ensure_hospital_config_file()


def _duty_local_storage():
    if _LocalStorageCls is None:
        return None
    return _LocalStorageCls(key=_LS_COMPONENT_STATE_KEY)


def _parse_requests_archive_raw(raw) -> dict:
    if raw is None or raw == "":
        return {}
    if isinstance(raw, dict):
        return raw
    if isinstance(raw, str):
        try:
            data = json.loads(raw)
            return data if isinstance(data, dict) else {}
        except json.JSONDecodeError:
            return {}
    return {}


def _requests_archive_from_local_storage(localS) -> dict:
    raw = localS.getItem(_LS_ARCHIVE_ITEM_KEY)
    arch = _parse_requests_archive_raw(raw)
    if not arch:
        disk = _load_schedule_requests_archive()
        if disk:
            arch = disk
            try:
                ctr = int(st.session_state.get("_ls_write_ctr", 0)) + 1
                st.session_state["_ls_write_ctr"] = ctr
                localS.setItem(_LS_ARCHIVE_ITEM_KEY, json.dumps(disk, ensure_ascii=False), key=f"ls_mig_{ctr}")
            except (TypeError, ValueError):
                pass
    return arch


def _forbidden_pairs_from_disk(raw_fp) -> dict[str, list]:
    fp_out: dict[str, list] = {}
    if not isinstance(raw_fp, dict):
        return fp_out
    for dk, rows in raw_fp.items():
        if not isinstance(rows, list):
            continue
        clean = []
        for row in rows:
            if not isinstance(row, (list, tuple)) or len(row) < 2:
                continue
            if isinstance(row[0], list):
                names = sorted({str(x).strip() for x in row[0] if str(x).strip()})
                sh_src = row[1] if len(row) > 1 else None
            else:
                a, b = str(row[0]).strip(), str(row[1]).strip()
                if not a or not b or a == b:
                    continue
                names = sorted({a, b})
                sh_src = row[2] if len(row) > 2 else None
            if len(names) < 2 or len(names) > 4:
                continue
            if isinstance(sh_src, (list, tuple)):
                sh = [x for x in sh_src if x in ("D", "E", "N")]
                if not sh:
                    sh = ["D", "E", "N"]
            else:
                sh = ["D", "E", "N"]
            sh = sorted(sh, key=lambda x: "DEN".index(x))
            clean.append([names, sh])
        fp_out[str(dk)] = clean
    return fp_out


def _pregnant_nurses_from_disk(raw_pg) -> dict[str, list[str]]:
    pg_out: dict[str, list[str]] = {}
    if not isinstance(raw_pg, dict):
        return pg_out
    for dk, rows in raw_pg.items():
        if not isinstance(rows, list):
            continue
        names = [str(x).strip() for x in rows if str(x).strip()]
        if names:
            pg_out[str(dk)] = names
    return pg_out


def _normalize_departments_blob(raw_dep) -> tuple[dict[str, list[str]], dict[str, dict]] | None:
    if not isinstance(raw_dep, dict) or not raw_dep:
        return None
    flat: dict[str, list[str]] = {}
    meta: dict[str, dict] = {}
    for k, v in raw_dep.items():
        name = str(k).strip()
        if not name:
            continue
        if isinstance(v, list):
            cleaned = _clean_nurse_names_list(v)
            if not cleaned:
                cleaned = ["수간호사"]
            flat[name] = cleaned
            meta[name] = _default_dept_meta()
        elif isinstance(v, dict):
            nurses = v.get("nurses")
            if not isinstance(nurses, list):
                continue
            cleaned = _clean_nurse_names_list(nurses)
            if not cleaned:
                cleaned = ["수간호사"]
            flat[name] = cleaned
            up = str(v.get("unit_profile") or "ward").strip().lower()
            if up not in ("icu", "er", "ward"):
                up = "ward"
            leg = str(v.get("access_code") or "").strip()
            adm = str(v.get("admin_code") or "").strip()
            if not adm and leg:
                adm = leg
            gen = str(v.get("general_code") or "").strip()
            meta[name] = {
                "general_code": gen,
                "admin_code": adm,
                "unit_profile": up,
            }
            rn = v.get("rule_note")
            if rn:
                meta[name]["rule_note"] = str(rn).strip()
    if not flat:
        return None
    return flat, meta


def _build_last_month_by_dept_from_raw(data: dict) -> dict[str, dict]:
    """파일 raw의 부서별·루트(레거시) 이월 근무 메타를 부서명 → blob 으로."""
    out: dict[str, dict] = {}
    raw_dep_full = data.get("departments") or {}
    if isinstance(raw_dep_full, dict):
        for dk, dv in raw_dep_full.items():
            if not isinstance(dv, dict):
                continue
            name = str(dk).strip()
            if not name or "last_month_shifts" not in dv:
                continue
            lm = dv.get("last_month_shifts")
            if not isinstance(lm, dict):
                continue
            out[name] = {
                "last_month_shifts": lm,
                "last_month_shifts_for": dv["last_month_shifts_for"]
                if isinstance(dv.get("last_month_shifts_for"), dict)
                else {},
                "last_month_shifts_note": dv.get("last_month_shifts_note", ""),
            }
    root_lm = data.get("last_month_shifts")
    if isinstance(root_lm, dict) and root_lm:
        rmeta = data.get("last_month_shifts_for")
        meta = rmeta if isinstance(rmeta, dict) else {}
        d0 = str(meta.get("department", "")).strip()
        if d0 and d0 not in out:
            out[d0] = {
                "last_month_shifts": dict(root_lm),
                "last_month_shifts_for": dict(meta),
                "last_month_shifts_note": data.get("last_month_shifts_note", ""),
            }
    return out


def _load_hospital_config_raw() -> dict | None:
    """hospital_config.json 전체(부서별 last_month_shifts 포함)."""
    if not _HOSPITAL_CONFIG_PATH.is_file():
        return None
    try:
        with open(_HOSPITAL_CONFIG_PATH, encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else None
    except (OSError, json.JSONDecodeError, TypeError, ValueError):
        return None


def _get_last_month_disk_blob(dept: str) -> tuple[dict | None, dict | None]:
    """
    디스크에서 부서 이월 근무 dict 와 last_month_shifts_for 메타.
    부서 행이 없으면 루트 last_month_shifts(department 메타 일치 시) 레거시.
    """
    raw = _load_hospital_config_raw()
    if not raw:
        return None, None
    drow = (raw.get("departments") or {}).get(dept)
    if isinstance(drow, dict) and "last_month_shifts" in drow:
        lm = drow["last_month_shifts"]
        if isinstance(lm, dict):
            meta = drow["last_month_shifts_for"]
            return lm, meta if isinstance(meta, dict) else {}
    rmeta = raw.get("last_month_shifts_for")
    meta = rmeta if isinstance(rmeta, dict) else {}
    if str(meta.get("department", "")).strip() != str(dept).strip():
        return None, None
    lm = raw.get("last_month_shifts")
    if isinstance(lm, dict):
        return lm, meta
    return None, None


def _validate_carry_json_for_persist(raw: str | None, nurse_names: list[str]) -> tuple[bool, str, dict | None]:
    """
    이월 JSON 저장 전 검증. (성공 여부, 메시지, 인덱스→carry 파싱 결과).
    빈 문자열은 {} 로 두는 것으로 간주.
    """
    s = (raw or "").strip()
    if not s:
        return True, "", {}
    try:
        data = json.loads(s)
    except json.JSONDecodeError as e:
        return False, f"JSON 파싱 오류 — 저장하지 않았습니다: {e}", None
    if not isinstance(data, dict):
        return False, "이월 데이터는 JSON 객체(이름 또는 인덱스 키)여야 합니다. 저장하지 않았습니다.", None
    if not data:
        return True, "", {}
    dumped = json.dumps(data, ensure_ascii=False)
    parsed = _parse_carry_in_text(dumped, nurse_names)
    if parsed is False:
        return (
            False,
            "이월 JSON의 키·값이 현재 부서 명단과 맞지 않습니다. 저장하지 않았습니다.",
            None,
        )
    return True, "", parsed


def _carry_parsed_to_name_dict(parsed: dict | None, nurse_names: list[str]) -> dict[str, list[str]]:
    if not parsed:
        return {}
    out: dict[str, list[str]] = {}
    for idx, seq in parsed.items():
        if isinstance(idx, int) and 0 <= idx < len(nurse_names) and isinstance(seq, (list, tuple)):
            out[nurse_names[idx]] = [str(x).strip() for x in seq if str(x).strip()]
    return out


def _persist_department_last_month_to_hospital_config(
    dept: str,
    year: int,
    month: int,
    carry_raw_text: str,
    nurse_names: list[str],
) -> tuple[bool, str]:
    """현재 부서의 last_month_shifts 를 hospital_config.json 부서 블록에 원자적으로 기록."""
    ok, msg, parsed = _validate_carry_json_for_persist(carry_raw_text, nurse_names)
    if not ok:
        return False, msg
    raw = _load_hospital_config_raw()
    if not raw:
        return False, "hospital_config.json을 읽을 수 없습니다."
    depts = raw.get("departments")
    if not isinstance(depts, dict) or dept not in depts:
        return False, f"hospital_config.json에 부서 «{dept}»가 없습니다."
    drow = depts[dept]
    if not isinstance(drow, dict):
        return False, "부서 설정 형식이 올바르지 않습니다."
    name_map = _carry_parsed_to_name_dict(parsed if parsed else None, nurse_names)
    drow["last_month_shifts"] = name_map
    drow["last_month_shifts_for"] = {
        "year": int(year),
        "month": int(month),
        "department": str(dept).strip(),
    }
    try:
        _atomic_write_json(_HOSPITAL_CONFIG_PATH, raw)
    except OSError as e:
        return False, f"파일 저장 실패(권한·경로): {e}"
    _carry_local_storage_put(dept, int(year), int(month), str(carry_raw_text or ""))
    return True, ""


def _hydrate_carry_textarea_from_disk(dept: str, year: int, month: int) -> None:
    """
    부서·연월이 바뀌면 파일 최신본으로 이월 입력칸을 채움(해당 연·월 메타가 일치할 때만).
    세션에 동일 (dept,y,m) 이면 건너뜀. 입력 상태 키는 부서×연월로 분리한다.
    """
    ctx = (str(dept).strip(), int(year), int(month))
    if st.session_state.get("_carry_prefill_ctx") == ctx:
        return
    key = _carry_widget_session_key(dept, year, month)
    leg = f"carry_txt_{dept}"
    lm, meta = _get_last_month_disk_blob(dept)
    ok = (
        isinstance(lm, dict)
        and bool(lm)
        and _carry_meta_applies_to_period(
            meta if isinstance(meta, dict) else None, int(year), int(month),
        )
    )
    if ok:
        st.session_state[key] = json.dumps(lm, ensure_ascii=False, indent=2)
    else:
        if key not in st.session_state and leg in st.session_state:
            st.session_state[key] = st.session_state[leg]
        elif key not in st.session_state:
            st.session_state[key] = ""
    st.session_state["_carry_prefill_ctx"] = ctx


def _bundle_from_hospital_json(
    path: Path,
    *,
    legacy_list_only: bool,
) -> dict | None:
    if not path.is_file():
        return None
    try:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
    except (OSError, json.JSONDecodeError, TypeError, ValueError):
        return None
    if not isinstance(data, dict):
        return None
    raw_dep = data.get("departments")
    if legacy_list_only and isinstance(raw_dep, dict) and raw_dep:
        if not all(isinstance(v, list) and len(v) >= 1 for v in raw_dep.values()):
            return None
    norm = _normalize_departments_blob(raw_dep)
    if norm is None:
        return None
    flat, dept_meta = norm
    fp_out = _forbidden_pairs_from_disk(data.get("forbidden_pairs"))
    pg_out = _pregnant_nurses_from_disk(data.get("pregnant_nurses"))
    n4_out = _pregnant_nurses_from_disk(data.get("n_max4_nurses"))
    dh_out: dict[str, str] = {}
    raw_h = data.get("dept_holidays")
    if isinstance(raw_h, dict):
        for _dk, dv in raw_h.items():
            selected_dept = str(_dk).strip()
            if selected_dept in flat:
                dh_out[selected_dept] = str(dv) if dv is not None else ""
    last_month_by_dept = _build_last_month_by_dept_from_raw(data)
    return {
        "departments": flat,
        "dept_meta": dept_meta,
        "active_dept": data.get("active_dept"),
        "forbidden_pairs": fp_out,
        "pregnant_nurses": pg_out,
        "n_max4_nurses": n4_out,
        "dept_holidays": dh_out,
        "last_month_by_dept": last_month_by_dept,
    }


def _load_hospital_config_bundle() -> dict | None:
    # hospital_config.json 우선 — 파싱 성공 시 user_departments.json으로 넘어가지 않음
    b = None
    if _HOSPITAL_CONFIG_PATH.is_file():
        b = _bundle_from_hospital_json(_HOSPITAL_CONFIG_PATH, legacy_list_only=False)
    if b is None and _DEPT_SAVE_PATH.is_file():
        b = _bundle_from_hospital_json(_DEPT_SAVE_PATH, legacy_list_only=True)
    _bundle_ensure_emergency_room(b)
    return b


def _session_schedule_requests_entries_for_dept(dept: str) -> dict[str, dict]:
    """세션 dept_requests[부서]의 DataFrame을 hospital_config.departments[].schedule_requests 엔트리 형식으로 변환."""
    out: dict[str, dict] = {}
    sub = (st.session_state.get("dept_requests") or {}).get(dept)
    if not isinstance(sub, dict):
        return out
    nurses = st.session_state.departments.get(dept)
    if not isinstance(nurses, list) or not nurses:
        return out
    hols = str((st.session_state.get("dept_holidays") or {}).get(dept, "") or "")
    days = get_april_days(_parse_holidays(hols))
    req_col_labels = [_day_label_compact(d) for d in days]
    for period_pk, df in sub.items():
        if not isinstance(period_pk, str) or df is None:
            continue
        if not hasattr(df, "shape"):
            continue
        ym = _period_pk_to_year_month(period_pk)
        if ym is None:
            continue
        yy, mm = ym
        try:
            df2 = _prepare_requests_df_for_current_table(df, nurses, req_col_labels)
            cleaned = _clean_req_df(df2)
        except (TypeError, ValueError, KeyError):
            continue
        out[period_pk] = {
            "year": int(yy),
            "month": int(mm),
            "nurse_names": [str(x) for x in nurses],
            "columns": [str(x) for x in req_col_labels],
            "data": cleaned.values.tolist(),
        }
    return out


def _hydrate_dept_requests_from_hospital_file_into_session() -> None:
    """재시작·F5 후 세션에 신청 근무가 비어 있으면 hospital_config.json 의 schedule_requests 로 채운다."""
    st.session_state.setdefault("dept_requests", {})
    raw = _load_hospital_config_raw()
    if not raw or not isinstance(raw.get("departments"), dict):
        return
    for dname, drow in raw["departments"].items():
        dname = str(dname).strip()
        if not dname or dname not in st.session_state.departments:
            continue
        if not isinstance(drow, dict):
            continue
        sr = drow.get("schedule_requests")
        if not isinstance(sr, dict) or not sr:
            continue
        nurses = list(st.session_state.departments[dname])
        if not nurses:
            continue
        hols = str((st.session_state.get("dept_holidays") or {}).get(dname, "") or "")
        days = get_april_days(_parse_holidays(hols))
        req_col_labels = [_day_label_compact(d) for d in days]
        sub = st.session_state.dept_requests.setdefault(dname, {})
        if not isinstance(sub, dict):
            sub = {}
            st.session_state.dept_requests[dname] = sub
        for period_pk in list(sr.keys()):
            if not isinstance(period_pk, str):
                continue
            existing = sub.get(period_pk)
            if existing is not None and hasattr(existing, "shape"):
                try:
                    if int(existing.shape[0]) > 0:
                        continue
                except (TypeError, ValueError, AttributeError):
                    pass
            df = _try_load_requests_from_hospital_config(dname, period_pk, nurses, req_col_labels)
            if df is None:
                continue
            try:
                sub[period_pk] = _prepare_requests_df_for_current_table(df, nurses, req_col_labels)
            except (TypeError, ValueError, KeyError):
                continue


def _save_hospital_config_to_disk() -> bool:
    if "departments" not in st.session_state:
        return False
    existing_sr: dict[str, dict] = {}
    _persist_carry_meta: dict = {}
    _raw_exist: dict = {}
    if _HOSPITAL_CONFIG_PATH.is_file():
        try:
            with open(_HOSPITAL_CONFIG_PATH, encoding="utf-8") as f:
                _raw_exist = json.load(f)
            if not isinstance(_raw_exist, dict):
                _raw_exist = {}
            if _raw_exist:
                for _k, _v in (_raw_exist.get("departments") or {}).items():
                    if isinstance(_v, dict) and isinstance(_v.get("schedule_requests"), dict):
                        existing_sr[str(_k)] = dict(_v["schedule_requests"])
                for _ck in (
                    "last_month_shifts",
                    "last_month_shifts_for",
                    "last_month_shifts_note",
                ):
                    if _ck in _raw_exist:
                        _persist_carry_meta[_ck] = _raw_exist[_ck]
        except (OSError, json.JSONDecodeError, TypeError, ValueError):
            _raw_exist = {}
    depts_out: dict = {}
    meta = st.session_state.get("dept_meta", {})
    for nm, nurses in st.session_state.departments.items():
        if not isinstance(nurses, list):
            continue
        m = meta.get(nm) or {}
        up = str(m.get("unit_profile") or "ward").strip().lower()
        if up not in ("icu", "er", "ward"):
            up = "ward"
        row = {
            "nurses": list(nurses),
            "general_code": str(m.get("general_code") or "").strip(),
            "admin_code": str(m.get("admin_code") or "").strip(),
            "unit_profile": up,
        }
        rn = m.get("rule_note")
        if rn:
            row["rule_note"] = str(rn).strip()
        # 디스크에 있던 schedule_requests + 세션 dept_requests(연도|월 키 동일) 병합 — 세션이 우선
        sr_disk = dict(existing_sr.get(nm, {}))
        sr_sess = _session_schedule_requests_entries_for_dept(nm)
        row["schedule_requests"] = {**sr_disk, **sr_sess}
        _exist_dept_row = (_raw_exist.get("departments") or {}).get(nm)
        if isinstance(_exist_dept_row, dict):
            for _lmk in (
                "last_month_shifts",
                "last_month_shifts_for",
                "last_month_shifts_note",
            ):
                if _lmk in _exist_dept_row:
                    row[_lmk] = _exist_dept_row[_lmk]
        depts_out[nm] = row
    dep_keys = set(st.session_state.departments.keys())
    payload = {
        "version": 1,
        "active_dept": st.session_state.get("active_dept", ""),
        "departments": depts_out,
        "forbidden_pairs": {
            selected_dept: v
            for selected_dept, v in st.session_state.get("dept_forbidden_pairs", {}).items()
            if selected_dept in dep_keys
        },
        "pregnant_nurses": {
            selected_dept: v
            for selected_dept, v in st.session_state.get("dept_pregnant", {}).items()
            if selected_dept in dep_keys
        },
        "n_max4_nurses": {
            selected_dept: v
            for selected_dept, v in st.session_state.get("dept_n_max4", {}).items()
            if selected_dept in dep_keys
        },
        "dept_holidays": {
            selected_dept: str(v) if v is not None else ""
            for selected_dept, v in st.session_state.get("dept_holidays", {}).items()
            if selected_dept in dep_keys
        },
    }
    payload.update(_persist_carry_meta)
    ok, err = _atomic_write_json_safe(_HOSPITAL_CONFIG_PATH, payload)
    if ok:
        try:
            st.session_state["_hospital_config_mtime_seen"] = float(
                _HOSPITAL_CONFIG_PATH.stat().st_mtime
            )
        except OSError:
            pass
        return True
    _enqueue_warning(f"hospital_config.json 저장 실패 — {err or '알 수 없음'}")
    return False


def _apply_nurse_data_editor_state(
    base_df: pd.DataFrame,
    editor_key: str,
    *,
    name_col: str = "이름",
) -> pd.DataFrame:
    """st.data_editor(session_state)의 edited / added / deleted_rows 를 단일 이름 열 DataFrame에 반영."""
    out = base_df.reset_index(drop=True).copy()
    if name_col not in out.columns and len(out.columns) >= 1:
        name_col = str(out.columns[0])
    raw = st.session_state.get(editor_key)
    if not isinstance(raw, dict):
        return out
    idx_id = "_index"
    er = raw.get("edited_rows")
    if isinstance(er, dict) and er:
        for row_id, changes in er.items():
            try:
                ri = int(row_id)
            except (TypeError, ValueError):
                continue
            if ri < 0 or ri >= len(out):
                continue
            if not isinstance(changes, dict):
                continue
            for col_name, val in changes.items():
                if col_name == idx_id:
                    continue
                if col_name not in out.columns:
                    continue
                j = out.columns.get_loc(col_name)
                out.iat[ri, j] = val
    dr = raw.get("deleted_rows") or []
    if dr:
        to_drop: list[int] = []
        for idx in dr:
            try:
                to_drop.append(int(idx))
            except (TypeError, ValueError):
                continue
        for di in sorted(set(to_drop), reverse=True):
            if 0 <= di < len(out):
                out = out.drop(index=di).reset_index(drop=True)
    for add in raw.get("added_rows") or []:
        if not isinstance(add, dict):
            continue
        val = add.get(name_col)
        if val is None and add:
            val = next((add[k] for k in add if k != idx_id), "")
        row_data = {name_col: "" if val is None else val}
        for c in out.columns:
            if c not in row_data:
                row_data[c] = None
        out = pd.concat([out, pd.DataFrame([row_data])], ignore_index=True)
    return out


def _nurse_roster_dataframe_has_changes(
    base_df: pd.DataFrame,
    edited_df: pd.DataFrame,
) -> bool:
    """명단 data_editor: 원본 DF와 반환 DF를 셀 단위로 비교(행 수·열·값 하나)."""
    if not isinstance(base_df, pd.DataFrame) or not isinstance(edited_df, pd.DataFrame):
        return True
    b = base_df.reset_index(drop=True)
    e = edited_df.reset_index(drop=True)
    if len(b) != len(e):
        return True
    if list(b.columns) != list(e.columns):
        return True
    if b.empty:
        return False
    col = "이름" if "이름" in b.columns else str(b.columns[0])
    jb = int(b.columns.get_loc(col))
    je = int(e.columns.get_loc(col))
    for i in range(len(b)):
        a = b.iat[i, jb]
        c = e.iat[i, je]
        sa = "" if pd.isna(a) else str(a)
        sb = "" if pd.isna(c) else str(c)
        if sa != sb:
            return True
    return False


def _filter_constraints_for_roster(dept: str, updated_nurses: list[str]) -> None:
    """명단 변경 후 함께 근무 불가·임산부·N≤4 명단을 현재 이름에 맞게 정리."""
    _fp = st.session_state.dept_forbidden_pairs.get(dept, [])

    def _fp_all_names_ok(p):
        ns = _fp_row_names_from_entry(p)
        return bool(ns) and all(n in updated_nurses for n in ns)

    st.session_state.dept_forbidden_pairs[dept] = [p for p in _fp if _fp_all_names_ok(p)]
    _pgn = st.session_state.setdefault("dept_pregnant", {}).get(dept, [])
    if isinstance(_pgn, list):
        st.session_state["dept_pregnant"][dept] = [n for n in _pgn if n in updated_nurses]
    _n4n = st.session_state.setdefault("dept_n_max4", {}).get(dept, [])
    if isinstance(_n4n, list):
        st.session_state["dept_n_max4"][dept] = [n for n in _n4n if n in updated_nurses]


_ROSTER_SAVE_TOAST_MSG = "✅ 명단 변경사항이 안전하게 저장되었습니다."


def _sync_roster_session_and_save_to_disk(
    dept: str,
    updated_nurses: list[str],
    *,
    toast_on_success: bool,
) -> bool:
    """디스크 직전 세션 명단을 edited 결과로 덮어쓴 뒤 저장. 성공 시 선택적으로 toast."""
    dept = str(dept).strip()
    if not dept or dept not in st.session_state.departments:
        return False
    names = _clean_nurse_names_list(updated_nurses)
    if not names:
        names = ["수간호사"]
    _filter_constraints_for_roster(dept, names)
    st.session_state.departments[dept] = list(names)
    ok = _save_hospital_config_to_disk()
    if ok and toast_on_success:
        st.toast(_ROSTER_SAVE_TOAST_MSG)
    return ok


def _persist_nurse_roster_from_editor_key(dept: str, editor_key: str) -> None:
    """Data Editor 변경 짉후: 세션 명단·부대 제약 동기화 후 hospital_config.json 저장( rerun 없음 )."""
    dept = str(dept).strip()
    if not dept or dept not in st.session_state.departments:
        return
    base = st.session_state.departments.get(dept)
    if not isinstance(base, list):
        return
    _df = pd.DataFrame({"이름": list(base)})
    merged = _apply_nurse_data_editor_state(_df, editor_key, name_col="이름")
    _cols = list(merged.columns)
    _col_n = "이름" if "이름" in _cols else (_cols[0] if _cols else "이름")
    _cells: list[object] = []
    for _, row in merged.iterrows():
        _cell = row[_col_n] if _col_n in row.index else None
        _cells.append(_cell)
    updated = _clean_nurse_names_list(_cells)
    if not updated:
        updated = ["수간호사"]
    if (
        tuple(updated) == tuple(base)
        and not _nurse_roster_dataframe_has_changes(
            _df.reset_index(drop=True),
            merged.reset_index(drop=True),
        )
    ):
        return
    _sync_roster_session_and_save_to_disk(dept, updated, toast_on_success=True)


def _on_nurse_roster_data_editor_change() -> None:
    if not st.session_state.get("dept_admin_verified"):
        return
    dept = str(st.session_state.get("active_dept") or "").strip()
    if not dept:
        return
    gen = int((st.session_state.nurse_gen or {}).get(dept, 0))
    editor_key = f"nurse_tbl_{dept}_g{gen}"
    _persist_nurse_roster_from_editor_key(dept, editor_key)


def _refresh_departments_from_disk_if_file_newer() -> None:
    """hospital_config.json이 세션 로드 이후 갱신되었으면 부서별 명단만 디스크 기준으로 맞춤."""
    if "departments" not in st.session_state:
        return
    if not _HOSPITAL_CONFIG_PATH.is_file():
        return
    try:
        mtime = float(_HOSPITAL_CONFIG_PATH.stat().st_mtime)
    except OSError:
        return
    last = float(st.session_state.get("_hospital_config_mtime_seen") or 0.0)
    if mtime <= last + 1e-6:
        return
    b = _load_hospital_config_bundle()
    if not b or not isinstance(b.get("departments"), dict):
        st.session_state["_hospital_config_mtime_seen"] = mtime
        return
    disk = b["departments"]
    for dn, names in disk.items():
        if dn not in st.session_state.departments or not isinstance(names, list):
            continue
        if list(st.session_state.departments[dn]) != list(names):
            st.session_state.departments[dn] = list(names)
    st.session_state["_hospital_config_mtime_seen"] = mtime


def _fp_row_names_from_entry(row) -> list[str] | None:
    """저장 행 → 정렬된 고유 이름 2~4명 또는 None."""
    if not row or not isinstance(row, (list, tuple)) or len(row) < 2:
        return None
    if isinstance(row[0], list):
        names = sorted({str(x).strip() for x in row[0] if str(x).strip()})
    else:
        a, b = str(row[0]).strip(), str(row[1]).strip()
        if not a or not b:
            return None
        names = sorted({a, b})
    if len(names) < 2 or len(names) > 4:
        return None
    return names


def _fp_pairs_to_indices(nurse_names: list[str], pairs: list) -> list[tuple[int, int, frozenset]]:
    """이름 그룹 2~4명(+적용 시프트) → 쌍 전개 (i, j, frozenset('D','E','N')). 수간호사 포함."""
    idx = {name: i for i, name in enumerate(nurse_names)}
    merged: dict[tuple[int, int], frozenset] = {}
    for row in pairs or []:
        names = _fp_row_names_from_entry(row)
        if not names:
            continue
        inds: list[int] = []
        bad = False
        for nm in names:
            if nm not in idx:
                bad = True
                break
            inds.append(idx[nm])
        if bad:
            continue
        inds = sorted(set(inds))
        if len(inds) < 2:
            continue
        if isinstance(row[0], list):
            sh_raw = row[1] if len(row) > 1 else None
        else:
            sh_raw = row[2] if len(row) > 2 else None
        if isinstance(sh_raw, (list, tuple, set, frozenset)):
            sh = frozenset(x for x in sh_raw if x in ("D", "E", "N"))
        else:
            sh = frozenset({"D", "E", "N"})
        if not sh:
            sh = frozenset({"D", "E", "N"})
        for ia in range(len(inds)):
            for ib in range(ia + 1, len(inds)):
                i, j = inds[ia], inds[ib]
                key = (min(i, j), max(i, j))
                merged[key] = merged.get(key, frozenset()) | sh
    return [(i, j, s) for (i, j), s in merged.items()]


def _period_storage_key(year: int, month: int) -> str:
    """신청·생성 근무를 연·월마다 따로 보관할 때 사용 (월 바꿔도 다른 달 데이터 유지)."""
    return f"{int(year)}|{int(month)}"


def _period_pk_to_year_month(period_pk: str) -> tuple[int, int] | None:
    """`_period_storage_key` 역변환 — shift_requests.json 키에 쓸 연·월."""
    parts = str(period_pk).strip().split("|", 1)
    if len(parts) != 2:
        return None
    try:
        return int(parts[0]), int(parts[1])
    except ValueError:
        return None


def _carry_widget_session_key(dept: str, year: int, month: int) -> str:
    """이월 JSON 텍스트를 부서·연월마다 별도 세션 슬롯에 보관 (부서 간·달 간 뒤섞임 방지)."""
    return f"carry_txt_{str(dept).strip()}__{_period_storage_key(int(year), int(month))}"


def _summarize_carry_lm_dict(lm: dict | None) -> str:
    if not lm:
        return "데이터 없음"
    lengths = [len(v) for v in lm.values() if isinstance(v, (list, tuple))]
    if not lengths:
        return "빈 객체"
    L0 = lengths[0]
    uneven = any(x != L0 for x in lengths)
    return (
        f"간호사 {len(lm)}명 · 이월 {L0}일"
        + (" (담당자마다 일수 다름)" if uneven else "")
    )


def _dept_carry_status_line(dept: str, year: int, month: int) -> str:
    """디스크 기준 해당 부서 이월이 «이번 생성 연·월»과 맞는지 한 줄 요약."""
    lm, meta = _get_last_month_disk_blob(dept)
    if lm is None or not isinstance(lm, dict):
        return "파일에 이월 없음"
    if not isinstance(meta, dict):
        return _summarize_carry_lm_dict(lm)
    try:
        y = int(meta.get("year", 0))
        m = int(meta.get("month", 0))
    except (TypeError, ValueError):
        return _summarize_carry_lm_dict(lm)
    if y != int(year) or m != int(month):
        return f"저장분 {y}년 {m}월 — 지금 표시 월({year}년 {month}월)과 다름"
    return "✓ 적용 · " + _summarize_carry_lm_dict(lm)


def _carry_meta_applies_to_period(meta: dict | None, year: int, month: int) -> bool:
    """last_month_shifts_for 가 없거나 비어 있으면 현재 표시 월에 맞는 것으로 본다."""
    if not isinstance(meta, dict) or not meta:
        return True
    try:
        yy = int(meta.get("year", 0))
        mm = int(meta.get("month", 0))
    except (TypeError, ValueError):
        return True
    if yy == 0 and mm == 0:
        return True
    return yy == int(year) and mm == int(month)


def _carry_ls_load_all(localS) -> dict[str, str]:
    if localS is None:
        return {}
    raw = localS.getItem(_LS_CARRY_ITEM_KEY_RY)
    if raw is None or raw == "":
        return {}
    try:
        data = json.loads(raw)
        return data if isinstance(data, dict) else {}
    except (json.JSONDecodeError, TypeError):
        return {}


def _carry_ls_save_all(localS, blob: dict[str, str]) -> None:
    try:
        ctr = int(st.session_state.get("_ls_carry_ctr", 0)) + 1
        st.session_state["_ls_carry_ctr"] = ctr
        localS.setItem(
            _LS_CARRY_ITEM_KEY_RY,
            json.dumps(blob, ensure_ascii=False),
            key=f"carry_ls_{ctr}",
        )
    except (TypeError, ValueError, OSError):
        pass


def _carry_local_storage_put(dept: str, year: int, month: int, text: str) -> None:
    """이월 JSON 텍스트를 브라우저에 백업(서버 재시작·세션 초기화 후에도 복구용)."""
    localS = _duty_local_storage()
    if localS is None:
        return
    blob = _carry_ls_load_all(localS)
    sk = _carry_widget_session_key(dept, int(year), int(month))
    blob[sk] = text if text is not None else ""
    _carry_ls_save_all(localS, blob)


def _seed_carry_session_from_persisted_sources(year: int, month: int) -> None:
    """
    세션에 없는 부서×연월 이월 칸만 채움. 순서: hospital_config.json → localStorage.
    앱/탭 리부팅 후에도 디스크·브라우저에 남은 내용이 입력 칸에 복구된다.
    """
    depts = st.session_state.get("departments")
    if not isinstance(depts, dict):
        return
    y_i, m_i = int(year), int(month)
    localS = _duty_local_storage()
    ls_blob = _carry_ls_load_all(localS) if localS is not None else {}
    for dname in depts:
        key = _carry_widget_session_key(dname, y_i, m_i)
        if key in st.session_state:
            continue
        lm, meta = _get_last_month_disk_blob(dname)
        filled = False
        if isinstance(lm, dict) and lm and _carry_meta_applies_to_period(
            meta if isinstance(meta, dict) else None, y_i, m_i,
        ):
            st.session_state[key] = json.dumps(lm, ensure_ascii=False, indent=2)
            filled = True
        if not filled:
            ls_text = ls_blob.get(key)
            if isinstance(ls_text, str) and ls_text.strip():
                st.session_state[key] = ls_text


def _set_schedule_edit_mode(dept: str, period_pk: str, enabled: bool) -> None:
    """근무표 ✏️ 수정 / 취소 — on_click에서 호출(세션 정규화 후 플래그 반영)."""
    if not dept or not period_pk:
        return
    st.session_state.edit_mode.setdefault(dept, {})
    sub = st.session_state.edit_mode[dept]
    if not isinstance(sub, dict):
        sub = {}
        st.session_state.edit_mode[dept] = sub
    if enabled:
        sub[period_pk] = True
    else:
        sub.pop(period_pk, None)


def _migrate_period_stores_if_needed() -> None:
    """기존 세션: 부서→표 단일 저장 → 부서→연월→표."""
    y = st.session_state.sel_year
    m = st.session_state.sel_month
    pk = _period_storage_key(y, m)

    def _first_nonempty(d: dict):
        for v in d.values():
            if v is not None:
                return v
        return None

    dr = st.session_state.get("dept_requests")
    if isinstance(dr, dict) and dr:
        fn = _first_nonempty(dr)
        if fn is not None and not isinstance(fn, dict):
            new_dr = {}
            for dept, val in dr.items():
                new_dr[dept] = {}
                if val is not None and hasattr(val, "shape"):
                    new_dr[dept][pk] = val
            st.session_state.dept_requests = new_dr
        elif fn is None:
            st.session_state.dept_requests = {str(d): {} for d in dr}

    ds = st.session_state.get("dept_schedules")
    if isinstance(ds, dict) and ds:
        fn = _first_nonempty(ds)
        inner_is_bundle = (
            isinstance(fn, dict)
            and fn
            and isinstance(next(iter(fn.values())), dict)
            and "schedule" in next(iter(fn.values()))
        )
        if inner_is_bundle:
            pass
        elif fn is not None and isinstance(fn, dict) and "schedule" in fn:
            new_ds = {}
            for dept, val in ds.items():
                new_ds[dept] = {}
                if val is not None and isinstance(val, dict) and "schedule" in val:
                    new_ds[dept][pk] = val
            st.session_state.dept_schedules = new_ds
        elif fn is None:
            st.session_state.dept_schedules = {str(d): {} for d in ds}

    em = st.session_state.get("edit_mode")
    if isinstance(em, dict) and em:
        v0 = next(iter(em.values()))
        if not isinstance(v0, dict):
            st.session_state.edit_mode = {
                d: ({pk: bool(v)} if v else {}) for d, v in em.items()
            }


def _sync_selected_dept() -> None:
    """부서별 격리: session의 데이터 키는 항상 active_dept와 같은 selected_dept를 쓴다."""
    st.session_state.selected_dept = st.session_state.active_dept


def _init_state():
    _repair_hospital_config_file_emergency_dept()
    loaded_holidays: dict[str, str] | None = None
    if "departments" in st.session_state:
        _refresh_departments_from_disk_if_file_newer()
    if "departments" not in st.session_state:
        loaded = _load_hospital_config_bundle()
        if loaded:
            st.session_state.departments = loaded["departments"]
            ad = loaded.get("active_dept") or ""
            st.session_state.active_dept = (
                ad if ad in st.session_state.departments else _primary_dept_key(st.session_state.departments)
            )
            st.session_state.dept_meta = dict(loaded.get("dept_meta") or {})
            for dn in st.session_state.departments:
                st.session_state.dept_meta.setdefault(dn, _default_dept_meta())
            for _dm in st.session_state.dept_meta.values():
                _dm.setdefault("general_code", "")
                _dm.setdefault("admin_code", "")
                _dm.setdefault("unit_profile", "ward")
                if not (_dm.get("admin_code") or "").strip() and (_dm.get("access_code") or "").strip():
                    _dm["admin_code"] = str(_dm.get("access_code") or "").strip()
            lfp = loaded.get("forbidden_pairs")
            if isinstance(lfp, dict):
                st.session_state.dept_forbidden_pairs = {
                    str(k): v for k, v in lfp.items() if isinstance(v, list)
                }
            else:
                st.session_state.dept_forbidden_pairs = {}
            lpg = loaded.get("pregnant_nurses")
            if isinstance(lpg, dict):
                st.session_state.dept_pregnant = {
                    str(k): list(v) if isinstance(v, list) else []
                    for k, v in lpg.items()
                }
            else:
                st.session_state.dept_pregnant = {}
            ln4 = loaded.get("n_max4_nurses")
            if isinstance(ln4, dict):
                st.session_state.dept_n_max4 = {
                    str(k): list(v) if isinstance(v, list) else []
                    for k, v in ln4.items()
                }
            else:
                st.session_state.dept_n_max4 = {}
            dh = loaded.get("dept_holidays")
            if isinstance(dh, dict) and dh:
                loaded_holidays = {
                    str(k): str(v) if v is not None else "" for k, v in dh.items()
                }
        else:
            _seed = _default_hospital_config_payload()
            _raw_seed = _seed.get("departments")
            _norm_seed = _normalize_departments_blob(_raw_seed)
            if _norm_seed:
                st.session_state.departments = _norm_seed[0]
                st.session_state.dept_meta = dict(_norm_seed[1])
                for dn in st.session_state.departments:
                    st.session_state.dept_meta.setdefault(dn, _default_dept_meta())
            else:
                _row0 = _er_department_hospital_row()
                st.session_state.departments = {_ER_DEPT_NAME: list(_row0["nurses"])}
                st.session_state.dept_meta = {
                    _ER_DEPT_NAME: _default_dept_meta(
                        "er",
                        _row0["general_code"],
                        _row0["admin_code"],
                        _row0["rule_note"],
                    ),
                }
            st.session_state.dept_forbidden_pairs = {}
            st.session_state.dept_pregnant = {}
            st.session_state.dept_n_max4 = {}
            ad0 = _seed.get("active_dept") or ""
            if ad0 in st.session_state.departments:
                st.session_state.active_dept = ad0
    if "dept_forbidden_pairs" not in st.session_state:
        _ld = _load_hospital_config_bundle()
        if _ld and isinstance(_ld.get("forbidden_pairs"), dict):
            st.session_state.dept_forbidden_pairs = {
                str(k): v for k, v in _ld["forbidden_pairs"].items() if isinstance(v, list)
            }
        else:
            st.session_state.dept_forbidden_pairs = {}
    if "active_dept" not in st.session_state:
        st.session_state.active_dept = _primary_dept_key(st.session_state.departments)
    # 연도·월
    if "sel_year" not in st.session_state:
        st.session_state.sel_year = 2026
    if "sel_month" not in st.session_state:
        st.session_state.sel_month = 5
    # 부서별 데이터 (dict of dict)
    for key in ("dept_schedules", "dept_requests", "dept_holidays", "nurse_gen", "edit_mode"):
        if key not in st.session_state:
            st.session_state[key] = {}
    if loaded_holidays:
        for selected_dept, dv in loaded_holidays.items():
            if selected_dept in st.session_state.departments:
                st.session_state.dept_holidays[selected_dept] = dv
    # 규칙 위반 팝업 제어
    if "show_violations" not in st.session_state:
        st.session_state.show_violations = False
    if "violations" not in st.session_state:
        st.session_state.violations = []
    st.session_state.setdefault("_warning_queue", [])
    _migrate_period_stores_if_needed()
    # 레거시 마이그레이션 이후: hospital_config.json 의 schedule_requests → 세션 dept_requests
    _hydrate_dept_requests_from_hospital_file_into_session()
    st.session_state.setdefault("dept_forbidden_pairs", {})
    if "dept_pregnant" not in st.session_state:
        _ldpg = _load_hospital_config_bundle()
        if _ldpg and isinstance(_ldpg.get("pregnant_nurses"), dict):
            st.session_state.dept_pregnant = {
                str(k): list(v) if isinstance(v, list) else []
                for k, v in _ldpg["pregnant_nurses"].items()
            }
        else:
            st.session_state.dept_pregnant = {}
    if "dept_n_max4" not in st.session_state:
        _ldn4 = _load_hospital_config_bundle()
        if _ldn4 and isinstance(_ldn4.get("n_max4_nurses"), dict):
            st.session_state.dept_n_max4 = {
                str(k): list(v) if isinstance(v, list) else []
                for k, v in _ldn4["n_max4_nurses"].items()
            }
        else:
            st.session_state.dept_n_max4 = {}
    st.session_state.setdefault("dept_pregnant", {})
    st.session_state.setdefault("dept_n_max4", {})
    st.session_state.setdefault("dept_meta", {})
    for _dn in st.session_state.departments:
        st.session_state.dept_meta.setdefault(_dn, _default_dept_meta())
    for _dm in st.session_state.dept_meta.values():
        _dm.setdefault("general_code", "")
        _dm.setdefault("admin_code", "")
        _dm.setdefault("unit_profile", "ward")
        if not (_dm.get("admin_code") or "").strip() and (_dm.get("access_code") or "").strip():
            _dm["admin_code"] = str(_dm.get("access_code") or "").strip()
    _ensure_emergency_department_session_state()
    st.session_state.setdefault("dept_auth_ok", {})
    _sync_selected_dept()
    _seed_carry_session_from_persisted_sources(
        int(st.session_state.sel_year), int(st.session_state.sel_month)
    )
    if _HOSPITAL_CONFIG_PATH.is_file():
        try:
            st.session_state["_hospital_config_mtime_seen"] = float(
                _HOSPITAL_CONFIG_PATH.stat().st_mtime
            )
        except OSError:
            pass

_init_state()
_carry_warn_pending = st.session_state.pop("_carry_persist_warning", None)
if _carry_warn_pending:
    st.warning(
        "신청 근무는 저장되었으나, 이월 근무 JSON 검증에 실패해 hospital_config.json의 last_month_shifts는 "
        f"바꾸지 않았습니다. ({_carry_warn_pending})"
    )


def _effective_unit_profile(dept: str) -> str:
    st.session_state.setdefault("dept_meta", {})
    m = st.session_state.dept_meta.get(dept) or {}
    up = str(m.get("unit_profile") or "").strip().lower()
    if up in ("icu", "er", "ward"):
        return up
    return _app.infer_unit_profile(dept)

# 네잎 클로버 SVG (연두 #81C784 계열) — 헤더용 단순 기하 도형
_DS_BRAND_SVG = """
<svg xmlns="http://www.w3.org/2000/svg" width="30" height="30" viewBox="0 0 32 32" aria-hidden="true">
  <g fill="#81C784">
    <circle cx="16" cy="9.5" r="6.2"/>
    <circle cx="23.5" cy="16" r="6.2"/>
    <circle cx="16" cy="22.5" r="6.2"/>
    <circle cx="8.5" cy="16" r="6.2"/>
  </g>
  <circle cx="16" cy="16" r="3.8" fill="#A5D6A7"/>
  <rect x="14.9" y="20" width="2.2" height="7.5" rx="1.1" fill="#66BB6A"/>
</svg>
"""


def _render_app_brand_header() -> None:
    """앱 최상단 브랜드 바: flex 헤더 + 클로버 SVG + Duty Solution (st.markdown HTML/CSS)."""
    st.markdown(
        f"""
<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/pretendard@1.3.9/dist/web/variable/pretendardvariable-dynamic-subset.min.css">
<div class="ds-brand-header" role="banner">
  <div class="ds-brand-header__inner">
    <span class="ds-brand-header__icon">{_DS_BRAND_SVG}</span>
    <span class="ds-brand-header__title">Duty Solution</span>
  </div>
</div>
""",
        unsafe_allow_html=True,
    )


def _parse_carry_in_text(raw: str, nurse_names: list[str]):
    """
    전월 말 근무 JSON → {간호사인덱스: [시프트,...]} 또는 None.
    파싱 실패 시 False.
    """
    if raw is None or not str(raw).strip():
        return None
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        return False
    if not isinstance(data, dict):
        return False
    idx = {name: i for i, name in enumerate(nurse_names)}
    out = {}
    for k, v in data.items():
        if isinstance(k, bool):
            continue
        if isinstance(k, int):
            i = k
        else:
            ks = str(k).strip()
            if ks.isdigit():
                i = int(ks)
            elif ks in idx:
                i = idx[ks]
            else:
                continue
        if not (isinstance(i, int) and 0 <= i < len(nurse_names)):
            continue
        if not isinstance(v, list):
            continue
        seq = [str(x).strip() for x in v if str(x).strip()]
        if seq:
            out[i] = seq
    return out if out else None


def _carry_from_hospital_last_month_shifts(
    bundle: dict | None,
    active_dept: str,
    year: int,
    month: int,
    nurse_names: list[str],
) -> dict | None:
    """부서별 또는 루트(레거시) last_month_shifts → carry_in. 연·월·부서 메타가 맞을 때만."""
    if not bundle or not isinstance(bundle, dict):
        return None
    blob: dict | None = None
    by_dept = bundle.get("last_month_by_dept")
    if isinstance(by_dept, dict):
        blob = by_dept.get(str(active_dept).strip())
    if blob:
        lm = blob.get("last_month_shifts")
        meta = blob.get("last_month_shifts_for") or {}
    else:
        lm = bundle.get("last_month_shifts")
        meta = bundle.get("last_month_shifts_for") or {}
    if not isinstance(lm, dict) or not lm:
        return None
    try:
        y = int(meta.get("year", 0))
        m = int(meta.get("month", 0))
    except (TypeError, ValueError):
        return None
    if y != int(year) or m != int(month):
        return None
    dept_need = str(meta.get("department", "")).strip()
    if dept_need and dept_need != str(active_dept).strip():
        return None
    return _parse_carry_in_text(json.dumps(lm, ensure_ascii=False), nurse_names)


def _merge_carry_with_hospital_last_month(
    parsed: dict | None | bool,
    bundle: dict | None,
    active_dept: str,
    year: int,
    month: int,
    nurse_names: list[str],
):
    """수기 JSON(parsed) 우선·동일 인덱스 덮어쓰기, 비어 있으면 last_month_shifts만 사용."""
    if parsed is False:
        return False
    cfg = _carry_from_hospital_last_month_shifts(
        bundle, active_dept, year, month, nurse_names,
    )
    if parsed is None:
        return cfg
    if not cfg:
        return parsed
    merged = dict(cfg)
    merged.update(parsed)
    return merged


def _carry_virtual_timeline_caption(year: int, month: int, carry_merged) -> str | None:
    """
    이월 리스트 길이 L → 당월 1일 직전일이 마지막 칸이 되도록 가상 Day -L … -1 로 고정.
    (4일이면 4일 전부터, 5일이면 5일 전부터 소급)
    """
    if carry_merged is False or not carry_merged or not isinstance(carry_merged, dict):
        return None
    from datetime import date, timedelta

    lens = [len(v) for v in carry_merged.values() if isinstance(v, (list, tuple)) and v]
    if not lens:
        return None
    lo, hi = min(lens), max(lens)
    if lo != hi:
        return (
            f"⚠️ 이월 일수가 간호사마다 다릅니다(최소 {lo}일·최대 {hi}일). "
            f"규칙 적용 전 **모든 행의 배열 길이를 동일하게** 맞춰 주세요."
        )
    L = hi
    mf = date(int(year), int(month), 1)
    first_d = mf - timedelta(days=L)
    last_d = mf - timedelta(days=1)
    return (
        f"**연속성(가상 타임라인):** 이월 **{L}일** → "
        f"`{first_d.year}-{first_d.month:02d}-{first_d.day:02d}` ~ "
        f"`{last_d.year}-{last_d.month:02d}-{last_d.day:02d}` "
        f"(말일 = 당월 1일 전날). 솔버는 이 구간을 Day -{L}…-1 로 두고 "
        f"당월과 합쳐 슬라이딩 윈도로 연속 N·연속근무·N-OF-D 를 적용합니다."
    )


def _month_archive_key(year: int, month: int) -> str:
    return f"{int(year)}|{int(month)}"


def _prev_year_month(year: int, month: int) -> tuple[int, int]:
    if month <= 1:
        return year - 1, 12
    return year, month - 1


def _schedule_to_jsonable(sched: dict) -> dict:
    out = {}
    for n, row in sched.items():
        nk = str(int(n) if not isinstance(n, str) else n)
        out[nk] = {str(int(d) if not isinstance(d, str) else d): v for d, v in row.items()}
    return out


def _schedule_from_jsonable(data: dict) -> dict:
    out = {}
    for nk, row in (data or {}).items():
        n = int(nk)
        out[n] = {}
        for dk, v in (row or {}).items():
            out[n][int(dk)] = v
    return out


def _load_schedule_archive() -> dict:
    if not _SCHEDULE_ARCHIVE_PATH.is_file():
        return {}
    try:
        with open(_SCHEDULE_ARCHIVE_PATH, encoding="utf-8") as f:
            return json.load(f)
    except (OSError, json.JSONDecodeError, TypeError):
        return {}


def _save_schedule_archive(archive: dict) -> None:
    try:
        with open(_SCHEDULE_ARCHIVE_PATH, "w", encoding="utf-8") as f:
            json.dump(archive, f, ensure_ascii=False, indent=2)
    except OSError:
        pass


def _load_schedule_requests_archive() -> dict:
    if not _SCHEDULE_REQUESTS_PATH.is_file():
        return {}
    try:
        with open(_SCHEDULE_REQUESTS_PATH, encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except (OSError, json.JSONDecodeError, TypeError):
        return {}


def _save_schedule_requests_archive(archive: dict) -> None:
    try:
        tmp = _SCHEDULE_REQUESTS_PATH.with_suffix(".json.tmp")
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(archive, f, ensure_ascii=False, indent=2)
        os.replace(tmp, _SCHEDULE_REQUESTS_PATH)
    except OSError:
        try:
            if tmp.is_file():
                tmp.unlink(missing_ok=True)
        except OSError:
            pass


def _schedule_requests_snapshot_rowshape_ok(
    snap: dict | None,
    nurses: list[str],
    req_col_labels: list[str],
) -> bool:
    """명단·일수가 현재 표와 맞으면 헤더 문자열이 달라도 데이터 행을 살린다(표시 라벨 변경 대비)."""
    if not snap or not isinstance(snap, dict):
        return False
    ns = snap.get("nurse_names")
    cs = snap.get("columns")
    data = snap.get("data")
    if not isinstance(ns, list) or not isinstance(cs, list) or not isinstance(data, list):
        return False
    if [str(x) for x in ns] != [str(x) for x in nurses]:
        return False
    if len(cs) != len(req_col_labels) or len(data) != len(ns):
        return False
    w = len(cs)
    for row in data:
        if not isinstance(row, list) or len(row) != w:
            return False
    return True


def _schedule_requests_snapshot_matches(
    snap: dict | None,
    nurses: list[str],
    req_col_labels: list[str],
) -> bool:
    if not snap or not isinstance(snap, dict):
        return False
    ns = snap.get("nurse_names")
    cs = snap.get("columns")
    data = snap.get("data")
    if not isinstance(ns, list) or not isinstance(cs, list) or not isinstance(data, list):
        return False
    if [str(x) for x in ns] != [str(x) for x in nurses]:
        return False
    if [str(x) for x in cs] != [str(x) for x in req_col_labels]:
        return False
    if len(data) != len(ns):
        return False
    w = len(cs)
    for row in data:
        if not isinstance(row, list) or len(row) != w:
            return False
    return True


def _req_cell_str(x: object) -> str:
    """신청 근무 칸: None / NaN / pd.NA → 빈 문자열(표시·JSON·Selectbox 일관)."""
    if x is None:
        return ""
    try:
        if pd.isna(x):
            return ""
    except TypeError:
        pass
    s = str(x).strip()
    if s in ("", "None", "nan", "<NA>"):
        return ""
    return s


def _load_shift_requests_json_root() -> dict:
    if not _SHIFT_REQUESTS_PATH.is_file():
        return {}
    try:
        with open(_SHIFT_REQUESTS_PATH, encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except (OSError, json.JSONDecodeError, TypeError, ValueError):
        return {}


def _shift_requests_period_key(dept: str, year: int, month: int) -> str:
    """화면 선택값과 동일 규칙: current_key = f'{부서}_{연도}_{월}'."""
    return f"{str(dept).strip()}_{int(year)}_{int(month)}"


def _try_load_requests_from_shift_requests_json(
    selected_dept: str,
    year: int,
    month: int,
    nurses: list[str],
    req_col_labels: list[str],
) -> pd.DataFrame | None:
    root = _load_shift_requests_json_root()
    if not root:
        return None
    ck = _shift_requests_period_key(selected_dept, year, month)
    snap = root.get(ck)
    if not isinstance(snap, dict):
        return None
    return _snapshot_to_requests_df(snap, nurses, req_col_labels)


def _snapshot_to_requests_df(
    snap: dict,
    nurses: list[str],
    req_col_labels: list[str],
) -> pd.DataFrame | None:
    if not _schedule_requests_snapshot_matches(snap, nurses, req_col_labels):
        if not _schedule_requests_snapshot_rowshape_ok(snap, nurses, req_col_labels):
            return None
    rows = []
    for row in snap["data"]:
        rows.append([_req_cell_str(c) for c in row])
    return pd.DataFrame(rows, index=list(nurses), columns=list(req_col_labels))


def _try_load_requests_from_archive(
    arch: dict,
    selected_dept: str,
    period_pk: str,
    nurses: list[str],
    req_col_labels: list[str],
) -> pd.DataFrame | None:
    if not arch or not selected_dept:
        return None
    sub = arch.get(str(selected_dept))
    if not isinstance(sub, dict):
        return None
    snap = sub.get(period_pk)
    return _snapshot_to_requests_df(snap, nurses, req_col_labels)


def _try_load_requests_from_disk(
    selected_dept: str,
    period_pk: str,
    nurses: list[str],
    req_col_labels: list[str],
) -> pd.DataFrame | None:
    return _try_load_requests_from_archive(
        _load_schedule_requests_archive(),
        selected_dept,
        period_pk,
        nurses,
        req_col_labels,
    )


def _try_load_requests_from_hospital_config(
    selected_dept: str,
    period_pk: str,
    nurses: list[str],
    req_col_labels: list[str],
) -> pd.DataFrame | None:
    if not selected_dept or not _HOSPITAL_CONFIG_PATH.is_file():
        return None
    try:
        with open(_HOSPITAL_CONFIG_PATH, encoding="utf-8") as f:
            data = json.load(f)
    except (OSError, json.JSONDecodeError, TypeError, ValueError):
        return None
    if not isinstance(data, dict):
        return None
    depts = data.get("departments")
    if not isinstance(depts, dict):
        return None
    d = depts.get(str(selected_dept).strip())
    if not isinstance(d, dict):
        return None
    sr = d.get("schedule_requests")
    if not isinstance(sr, dict):
        return None
    snap = sr.get(period_pk)
    return _snapshot_to_requests_df(snap, nurses, req_col_labels)


def _try_load_requests_from_saved_sources(
    selected_dept: str,
    period_pk: str,
    nurses: list[str],
    req_col_labels: list[str],
    req_arch: dict,
) -> pd.DataFrame | None:
    """신청 근무 디스크/백업 로드 — shift_requests.json(부서_연도_월) → hospital_config → 아카이브."""
    if not selected_dept or not period_pk:
        return None
    _ym = _period_pk_to_year_month(period_pk)
    if _ym is not None:
        _yy, _mm = _ym
        _sh = _try_load_requests_from_shift_requests_json(
            selected_dept, _yy, _mm, nurses, req_col_labels,
        )
        if _sh is not None:
            return _sh
    hc_df = _try_load_requests_from_hospital_config(
        selected_dept, period_pk, nurses, req_col_labels
    )
    if hc_df is not None:
        return hc_df
    return _try_load_requests_from_archive(
        req_arch, selected_dept, period_pk, nurses, req_col_labels
    )


def _load_requests_dataframe_for_selected_dept(
    selected_dept: str,
    period_pk: str,
    nurses: list[str],
    req_col_labels: list[str],
    req_arch: dict,
) -> pd.DataFrame | None:
    """부서별 필터: hospital_config.departments[부서].schedule_requests → 동일 부서 키의 schedule_requests.json(또는 req_arch)."""
    sd = str(selected_dept).strip()
    if not sd or not period_pk:
        return None
    hc_df = _try_load_requests_from_hospital_config(
        sd, period_pk, nurses, req_col_labels
    )
    if hc_df is not None:
        return hc_df
    ar_df = _try_load_requests_from_archive(
        req_arch, sd, period_pk, nurses, req_col_labels
    )
    if ar_df is not None:
        return ar_df
    return _try_load_requests_from_disk(sd, period_pk, nurses, req_col_labels)


def _prepare_requests_df_for_current_table(
    df: pd.DataFrame,
    nurses: list[str],
    req_col_labels: list[str],
) -> pd.DataFrame:
    """불러온 스냅샷을 표(인덱스=명단, 열=날짜 헤더)와 동일 형식으로 정규화."""
    out = df.fillna("").apply(lambda col: col.map(_req_cell_str))
    out = _normalize_req_shift_cells(
        _clean_req_df(out), frozenset(REQUEST_SHIFT_OPTIONS)
    )
    out = out.copy()
    idx = list(nurses)
    if len(out) != len(idx) or [str(x) for x in out.index] != [str(x) for x in idx]:
        try:
            out = out.reindex(index=idx, fill_value="")
        except Exception:
            out = pd.DataFrame(
                [[""] * len(req_col_labels) for _ in idx],
                index=idx,
                columns=list(req_col_labels),
            )
    out.index = idx
    if list(out.columns) != list(req_col_labels):
        out = out.reindex(columns=list(req_col_labels), fill_value="")
        out = out.fillna("").apply(lambda col: col.map(_req_cell_str))
    return out


def _persist_schedule_requests(
    selected_dept: str,
    period_pk: str,
    year: int,
    month: int,
    nurses: list[str],
    req_col_labels: list[str],
    df: pd.DataFrame,
) -> None:
    if not selected_dept or not period_pk:
        return
    cleaned = _clean_req_df(df)
    entry = {
        "year": int(year),
        "month": int(month),
        "nurse_names": [str(x) for x in nurses],
        "columns": [str(x) for x in req_col_labels],
        "data": cleaned.values.tolist(),
    }
    localS = _duty_local_storage()
    if localS is not None:
        arch = _parse_requests_archive_raw(localS.getItem(_LS_ARCHIVE_ITEM_KEY))
        if not arch:
            arch = _load_schedule_requests_archive()
        arch.setdefault(str(selected_dept), {})[period_pk] = entry
        ctr = int(st.session_state.get("_ls_write_ctr", 0)) + 1
        st.session_state["_ls_write_ctr"] = ctr
        localS.setItem(
            _LS_ARCHIVE_ITEM_KEY,
            json.dumps(arch, ensure_ascii=False),
            key=f"ls_persist_{ctr}",
        )
    else:
        arch = _load_schedule_requests_archive()
        arch.setdefault(str(selected_dept), {})[period_pk] = entry
        _save_schedule_requests_archive(arch)


def _dept_row_payload_from_session(dept_name: str) -> dict:
    nurses = st.session_state.departments.get(dept_name)
    if not isinstance(nurses, list):
        nurses = []
    nurses = _clean_nurse_names_list(nurses)
    if not nurses:
        nurses = ["수간호사"]
    raw_meta = (st.session_state.get("dept_meta") or {}).get(dept_name)
    meta = dict(raw_meta) if isinstance(raw_meta, dict) else _default_dept_meta()
    up = str(meta.get("unit_profile") or "ward").strip().lower()
    if up not in ("icu", "er", "ward"):
        up = "ward"
    row: dict = {
        "nurses": [str(x) for x in nurses],
        "general_code": str(meta.get("general_code") or "").strip(),
        "admin_code": str(meta.get("admin_code") or "").strip(),
        "unit_profile": up,
    }
    rn = meta.get("rule_note")
    if rn:
        row["rule_note"] = str(rn).strip()
    return row


def _save_dept_schedule_requests_to_hospital_config(
    selected_dept: str,
    period_pk: str,
    year: int,
    month: int,
    nurses: list[str],
    req_col_labels: list[str],
    df: pd.DataFrame,
) -> bool:
    if not selected_dept or not period_pk:
        return False
    cleaned = _clean_req_df(df)
    entry = {
        "year": int(year),
        "month": int(month),
        "nurse_names": [str(x) for x in nurses],
        "columns": [str(x) for x in req_col_labels],
        "data": cleaned.values.tolist(),
    }
    path = _HOSPITAL_CONFIG_PATH
    dk = str(selected_dept).strip()
    try:
        if path.is_file():
            with open(path, encoding="utf-8") as f:
                data = json.load(f)
            if not isinstance(data, dict):
                data = {}
        else:
            data = {}
        depts = data.setdefault("departments", {})
        if not isinstance(depts, dict):
            depts = {}
            data["departments"] = depts
        raw = depts.get(dk)
        if isinstance(raw, list):
            base = _dept_row_payload_from_session(dk)
            base["schedule_requests"] = {period_pk: entry}
            depts[dk] = base
        elif isinstance(raw, dict):
            sr = raw.get("schedule_requests")
            if not isinstance(sr, dict):
                sr = {}
            sr = dict(sr)
            sr[period_pk] = entry
            fresh = _dept_row_payload_from_session(dk)
            depts[dk] = {**raw, **fresh, "schedule_requests": sr}
        else:
            base = _dept_row_payload_from_session(dk)
            base["schedule_requests"] = {period_pk: entry}
            depts[dk] = base
        data.setdefault("version", 1)
        _atomic_write_json(path, data)
        with open(path, encoding="utf-8") as f:
            verify = json.load(f)
        vdepts = verify.get("departments") if isinstance(verify, dict) else None
        vd = vdepts.get(dk) if isinstance(vdepts, dict) else None
        vsr = vd.get("schedule_requests") if isinstance(vd, dict) else None
        if not isinstance(vsr, dict) or period_pk not in vsr:
            return False
        vent = vsr.get(period_pk)
        return (
            isinstance(vent, dict)
            and vent.get("data") == entry["data"]
            and vent.get("columns") == entry["columns"]
            and vent.get("nurse_names") == entry["nurse_names"]
        )
    except (OSError, TypeError, ValueError, json.JSONDecodeError):
        return False


def _delete_schedule_requests_period(selected_dept: str, period_pk: str) -> None:
    localS = _duty_local_storage()
    if localS is not None:
        arch = _parse_requests_archive_raw(localS.getItem(_LS_ARCHIVE_ITEM_KEY))
        if not arch:
            return
        sub = arch.get(str(selected_dept))
        if not isinstance(sub, dict) or period_pk not in sub:
            return
        sub.pop(period_pk, None)
        if not sub:
            arch.pop(str(selected_dept), None)
        ctr = int(st.session_state.get("_ls_write_ctr", 0)) + 1
        st.session_state["_ls_write_ctr"] = ctr
        localS.setItem(
            _LS_ARCHIVE_ITEM_KEY,
            json.dumps(arch, ensure_ascii=False),
            key=f"ls_delp_{ctr}",
        )
        return
    arch = _load_schedule_requests_archive()
    sub = arch.get(str(selected_dept))
    if not isinstance(sub, dict) or period_pk not in sub:
        return
    sub.pop(period_pk, None)
    if not sub:
        arch.pop(str(selected_dept), None)
    _save_schedule_requests_archive(arch)


def _delete_schedule_requests_dept(selected_dept: str) -> None:
    localS = _duty_local_storage()
    if localS is not None:
        arch = _parse_requests_archive_raw(localS.getItem(_LS_ARCHIVE_ITEM_KEY))
        if str(selected_dept) in arch:
            arch.pop(str(selected_dept), None)
            ctr = int(st.session_state.get("_ls_write_ctr", 0)) + 1
            st.session_state["_ls_write_ctr"] = ctr
            localS.setItem(
                _LS_ARCHIVE_ITEM_KEY,
                json.dumps(arch, ensure_ascii=False),
                key=f"ls_deld_{ctr}",
            )
        return
    arch = _load_schedule_requests_archive()
    if str(selected_dept) in arch:
        arch.pop(str(selected_dept), None)
        _save_schedule_requests_archive(arch)


def _archive_put_month(selected_dept: str, year: int, month: int, nurse_names: list[str], schedule: dict) -> None:
    """해당 연·월 근무표를 디스크 아카이브에 저장 (자동 이월용)."""
    if not selected_dept or not nurse_names or not schedule:
        return
    arch = _load_schedule_archive()
    arch.setdefault(str(selected_dept), {})[_month_archive_key(year, month)] = {
        "nurse_names": [str(x) for x in nurse_names],
        "schedule": _schedule_to_jsonable(schedule),
    }
    _save_schedule_archive(arch)


def _build_carry_from_prev_month(
    selected_dept: str,
    year: int,
    month: int,
    nurse_names: list[str],
    n_days: int = CARRY_AUTO_DAYS,
) -> tuple[dict[int, list[str]] | None, str | None]:
    """
    직전 달 아카이브에서 마지막 n_days일 근무를 추출 → carry_in 형식.
    성공 시 (dict, None), 실패 시 (None, 메시지).
    오직 selected_dept 아카이브·현재 부서 명단만 사용한다.
    """
    py, pm = _prev_year_month(year, month)
    arch = _load_schedule_archive()
    sub_prev = arch.get(str(selected_dept))
    if not isinstance(sub_prev, dict):
        return None, f"{py}년 {pm}월에 저장된 근무표가 없습니다. 먼저 그 달에 근무표를 생성·저장하세요."
    entry = sub_prev.get(_month_archive_key(py, pm))
    if not entry:
        return None, f"{py}년 {pm}월에 저장된 근무표가 없습니다. 먼저 그 달에 근무표를 생성·저장하세요."
    old_names = [str(x) for x in (entry.get("nurse_names") or [])]
    sched = _schedule_from_jsonable(entry.get("schedule") or {})
    last_day = _calendar.monthrange(py, pm)[1]
    start_d = max(1, last_day - int(n_days) + 1)
    day_list = list(range(start_d, last_day + 1))
    name_to_si = {n: i for i, n in enumerate(old_names)}
    carry: dict[int, list[str]] = {}
    for i, nm in enumerate(nurse_names):
        si = name_to_si.get(str(nm).strip())
        seq = []
        for d in day_list:
            if si is None:
                seq.append("OF")
            else:
                v = sched.get(si, {}).get(d, "")
                seq.append(str(v).strip() if v not in (None, "") else "OF")
        carry[i] = seq
    return carry, None


# ── 연도·월 전역 상수 동기화 (렌더링마다 app 모듈 갱신)
_app.set_period(st.session_state.sel_year, st.session_state.sel_month)

# 부서 로그인 여부(_auth_ok)는 상단 부서 선택 직후에 계산한다.


def _enqueue_warning(message: str) -> None:
    """다음 실행에서 ⚠️ 기타 알림 Expander에 표시할 문구를 큐에 넣는다."""
    m = str(message).strip()
    if not m:
        return
    st.session_state.setdefault("_warning_queue", []).append(m)


# ════════════════════════════════════════════════════════════════════════════════
#  규칙 위반 팝업 다이얼로그 (세션: st.session_state.show_violations)
# ════════════════════════════════════════════════════════════════════════════════
def _violations_dialog_dismissed() -> None:
    """X·바깥 클릭·ESC로 닫을 때 플래그 해제(전체 리런 시 모달이 다시 열리지 않도록)."""
    st.session_state.show_violations = False


def _violations_modal_close_click() -> None:
    """닫기 버튼: 대화상자는 프래그먼트 리런만 일으킬 수 있어 전체 앱 리런으로 확실히 닫는다."""
    st.session_state.show_violations = False
    st.rerun(scope="app")


@st.dialog(
    "📋 생성 근무표 — 검토 메모",
    width="small",
    on_dismiss=_violations_dialog_dismissed,
)
def _show_violations_dialog():
    issues = list(st.session_state.get("violations") or [])
    errors = [v for v in issues if v.get("level") == "error"]
    warns = [v for v in issues if v.get("level") == "warn"]

    if not issues:
        st.success("✅ 모든 규칙을 만족합니다!")
    else:
        st.success(
            "✅ 근무표는 이미 화면에 반영되었습니다. "
            "아래는 참고·수정용 검토 목록입니다. (주간 2일 휴무 등은 노란 경고로 표시됩니다.)"
        )
        st.caption(f"🔴 검토(오류 표기) {len(errors)}건 &nbsp;|&nbsp; 🟡 경고 {len(warns)}건")
        st.markdown("---")

        if errors:
            st.markdown("**🔴 검토(기존 오류 등급 메시지)**")
            for v in errors:
                st.error(v.get("msg", ""), icon="🚨")

        if warns:
            st.markdown("**🟡 경고**")
            st.markdown("\n".join(f"- {v.get('msg', '')}" for v in warns))

    st.divider()
    st.markdown(
        '<div style="min-height:0.75rem;"></div>',
        unsafe_allow_html=True,
    )
    st.button(
        "닫기",
        key="violations_review_modal_close",
        type="primary",
        use_container_width=True,
        help="검토 창을 닫고 근무표 화면으로 돌아갑니다.",
        on_click=_violations_modal_close_click,
    )


# 검토 팝업 — `_auth_ok` 는 부서 로그인 스트립에서 설정한 뒤 아래에서 연다.

# ── 안전하게 active_dept 보정 (부서 삭제 후 남은 부서로 자동 전환)
if st.session_state.active_dept not in st.session_state.departments:
    st.session_state.active_dept = _primary_dept_key(st.session_state.departments)
_sync_selected_dept()

# ════════════════════════════════════════════════════════════════════════════════
#  헬퍼 함수
# ════════════════════════════════════════════════════════════════════════════════
def _parse_holidays(text: str) -> list[int]:
    result = []
    for tok in text.replace("，", ",").split(","):
        tok = tok.strip()
        if tok.isdigit():
            d = int(tok)
            if 1 <= d <= _app.NUM_DAYS:
                result.append(d)
    return sorted(set(result))


def _day_label(day: dict) -> str:
    mark = "🔴" if day["is_holiday"] else ("🔵" if day["is_weekend"] else "")
    return f"{day['day']}({day['weekday_name']}){mark}"


def _day_label_compact(day: dict) -> str:
    """신청 근무 표 헤더용 — 가로 폭 최소 (일만 + 표시)."""
    d = day["day"]
    if day["is_holiday"]:
        return f"{d}♦"
    if day["is_weekend"]:
        return f"{d}·"
    return str(d)


def _monday_week_split_style(day: dict) -> str:
    """월요일(weekday==0) 칸 왼쪽 — 일요일·월요일 사이 주 구분 빨간 굵은 세로선."""
    if day.get("weekday") == 0:
        # border shorthand 이후에 left만 덮어씀 (일·월 사이 한 줄)
        return "border-left:6px solid #B71C1C;"
    return ""


def _inject_week_split_css(days: list) -> None:
    """st.data_editor: 월요일 열 왼쪽 굵은 빨간 선 (인덱스 열 다음 = nth-child 2가 1일)."""
    indices = [j for j, d in enumerate(days) if d.get("weekday") == 0]
    if not indices:
        return
    parts: list[str] = []
    for j in indices:
        n = j + 2  # 1=행 이름, 2=1일, …
        parts.extend(
            [
                f'section[data-testid="stMain"] [data-testid="stDataFrame"] thead th:nth-child({n})',
                f'section[data-testid="stMain"] [data-testid="stDataFrame"] tbody td:nth-child({n})',
                f'section[data-testid="stMain"] [data-testid="stDataFrame"] tr > th:nth-child({n})',
                f'section[data-testid="stMain"] [data-testid="stDataFrame"] tr > td:nth-child({n})',
                f'[data-testid="stDataFrame"] thead th:nth-child({n})',
                f'[data-testid="stDataFrame"] tbody td:nth-child({n})',
            ]
        )
    st.markdown(
        "<style>"
        + ",\n".join(parts)
        + " {\n  border-left: 6px solid #B71C1C !important;\n"
        "  box-shadow: none !important;\n}\n</style>",
        unsafe_allow_html=True,
    )


def _make_requests_df(nurses: list[str], days: list) -> pd.DataFrame:
    num_days = _app.NUM_DAYS
    return pd.DataFrame(
        [[""] * num_days for _ in range(len(nurses))],
        index=nurses,
        columns=[_day_label_compact(d) for d in days],
    )


def _df_to_requests(
    df: pd.DataFrame,
    days: list,
    nurses: list[str] | None = None,
) -> dict:
    """행 순서가 아닌 **명단(nurses) 순 인덱스**와 맞추어 dict[간호사인덱스][일]=시프트 생성."""
    if nurses is not None:
        try:
            _aligned = df.reindex(index=list(nurses))
        except (TypeError, ValueError, KeyError):
            _aligned = df
    else:
        _aligned = df
    result: dict[int, dict[int, str]] = {}
    for i in range(len(_aligned)):
        for j, day in enumerate(days):
            raw = _aligned.iloc[i, j]
            try:
                val = "" if pd.isna(raw) else str(raw).strip()
            except (TypeError, ValueError):
                val = str(raw).strip() if raw is not None else ""
            if val and val not in ("", "None", "nan"):
                result.setdefault(i, {})[day["day"]] = val
    return result


# 근무표 HTML 미리보기·엑셀 다운로드 공통 셀 색
_PREVIEW_FG_BLACK = "#000000"
_PREVIEW_BG_DE = "#FFFFFF"
_PREVIEW_BG_OF_PINK = "#F8BBD0"
_PREVIEW_BG_LEAVE_YELLOW = "#FFF59D"


def _preview_shift_bg_fg(shift: str) -> tuple[str, str]:
    """미리보기 셀 (배경, 글자). N만 기존 앱 색 유지, 나머지 글자는 검정."""
    if not shift:
        return "#FFFFFF", "#BDBDBD"
    if shift == "N":
        return (
            SHIFT_COLORS.get("N", "#E8EAF6"),
            SHIFT_TEXT_COLORS.get("N", "#283593"),
        )
    if shift in ("D", "E"):
        return _PREVIEW_BG_DE, _PREVIEW_FG_BLACK
    if shift in ("OF", "NO"):
        return _PREVIEW_BG_OF_PINK, _PREVIEW_FG_BLACK
    if shift in ("연", "공", "EDU", "경"):
        return _PREVIEW_BG_LEAVE_YELLOW, _PREVIEW_FG_BLACK
    bg = SHIFT_COLORS.get(shift, "#FFFFFF")
    return bg, _PREVIEW_FG_BLACK


def _preview_shift_matches_filter(shift: str, preview_mode: str) -> bool:
    """미리보기 필터: 해당 시프트를 이 모드에서 표시할지."""
    if preview_mode == "all":
        return True
    if not shift:
        return False
    if preview_mode == "D":
        return shift == "D"
    if preview_mode == "E":
        return shift == "E"
    if preview_mode == "N":
        return shift == "N"
    if preview_mode == "off":
        return shift in ("OF", "OH", "NO", "연")
    return True


def _render_schedule_html(
    schedule: dict,
    nurse_names: list,
    days: list,
    requests: dict | None = None,
    preview_mode: str = "all",
) -> str:
    num = len(nurse_names)
    requests = requests or {}
    _pm = preview_mode if preview_mode in ("all", "D", "E", "N", "off") else "all"
    th = lambda txt, bg, extra="", fg="#37474F": (
        f'<th style="background:{bg};color:{fg};padding:4px 2px;'
        f'text-align:center;white-space:nowrap;{extra}">{txt}</th>'
    )
    _hdr: list[str] = ["<tr>"]
    _hdr.append(th("간호사", "#ECEFF1",
                   "min-width:80px;padding:5px 8px;font-size:11px;"
                   "position:sticky;left:0;z-index:5;", "#263238"))
    for day in days:
        if day["is_holiday"]:
            dbg, dfg = "#FFEBEE", "#C62828"
        elif day["is_weekend"]:
            dbg, dfg = "#E3F2FD", "#1565C0"
        else:
            dbg, dfg = "#F5F5F5", "#455A64"
        _wsp = _monday_week_split_style(day)
        _hdr.append(th(
            f"{day['day']}<br><span style='font-size:9px'>{day['weekday_name']}</span>",
            dbg, f"min-width:36px;{_wsp}", dfg,
        ))
    if _pm == "all":
        _sum_keys = ["N", "D", "E", "OF", "OH", "연"]
    elif _pm == "D":
        _sum_keys = ["D"]
    elif _pm == "E":
        _sum_keys = ["E"]
    elif _pm == "N":
        _sum_keys = ["N"]
    else:
        _sum_keys = ["OF", "OH", "연"]

    for lbl in _sum_keys:
        bg, fg = _preview_shift_bg_fg(lbl)
        _hdr.append(th(
            f"{lbl}<br><span style='font-size:9px'>합계</span>",
            bg, "min-width:36px;", fg,
        ))
    _hdr.append("</tr>")
    _header_html = "".join(_hdr)
    _body: list[str] = []
    _n_sum_cols = len(_sum_keys)

    for n_idx, name in enumerate(nurse_names):
        ns = schedule.get(n_idx, {})
        counts = {"N": 0, "D": 0, "E": 0, "OF": 0, "OH": 0, "연": 0}
        row_bg = "#FAFAFA" if n_idx % 2 == 0 else "#F5F5F5"
        cells = [
            f'<td style="background:#ECEFF1;color:#263238;font-weight:700;'
            f'padding:4px 8px;white-space:nowrap;position:sticky;left:0;'
            f'border-right:2px solid #CFD8DC;">{name}</td>'
        ]
        nurse_req = requests.get(n_idx, {})
        for day in days:
            d_num = day["day"]
            shift = ns.get(d_num, "")
            _wsp = _monday_week_split_style(day)
            vis = _pm == "all" or _preview_shift_matches_filter(shift, _pm)
            if vis:
                bg, fg = _preview_shift_bg_fg(shift)
                disp = shift
                is_requested = nurse_req.get(d_num) == shift and shift != ""
            else:
                bg, fg = "#EEEEEE", "#BDBDBD"
                disp = ""
                is_requested = False
            underline = "text-decoration:underline;text-underline-offset:3px;" if is_requested else ""
            cells.append(
                f'<td style="background:{bg};color:{fg};font-weight:700;{underline}'
                f'padding:3px 1px;text-align:center;border:1px solid #E0E0E0;{_wsp}">{disp}</td>'
            )
            if shift == "N":
                counts["N"] += 1
            elif shift == "D":
                counts["D"] += 1
            elif shift == "E":
                counts["E"] += 1
            elif shift in ("OF", "NO"):
                counts["OF"] += 1
            elif shift == "OH":
                counts["OH"] += 1
            elif shift == "연":
                counts["연"] += 1

        for key in _sum_keys:
            bg, fg = _preview_shift_bg_fg(key)
            cells.append(
                f'<td style="background:{bg};color:{fg};font-weight:700;'
                f'text-align:center;padding:3px;">{counts[key]}</td>'
            )
        _body.append(f'<tr style="background:{row_bg};">' + "".join(cells) + "</tr>")

    def _append_summary_row(lbl: str, sk_or_fn, *, color_key: str):
        if isinstance(sk_or_fn, str):
            sk = sk_or_fn

            def _cnt(dn: int) -> int:
                return sum(
                    1 for n in range(num) if schedule.get(n, {}).get(dn) == sk
                )
        else:
            _cnt = sk_or_fn
        hbg, hfg = _preview_shift_bg_fg(color_key)
        bg, data_fg = hbg, hfg
        if color_key in ("D", "E"):
            hbg = _PREVIEW_BG_DE
            hfg = _PREVIEW_FG_BLACK
            bg = _PREVIEW_BG_DE
            data_fg = _PREVIEW_FG_BLACK
        cells = [
            f'<td style="background:{hbg};color:{hfg};font-weight:700;'
            f'padding:4px 8px;white-space:nowrap;position:sticky;left:0;'
            f'border-right:2px solid #CFD8DC;">{lbl}</td>'
        ]
        for day in days:
            dn = day["day"]
            cnt = _cnt(dn)
            _wsp = _monday_week_split_style(day)
            cells.append(
                f'<td style="background:{bg};color:{data_fg};font-weight:700;text-align:center;'
                f'padding:3px;border:1px solid #E0E0E0;{_wsp}">{cnt}</td>'
            )
        cells += ["<td></td>"] * _n_sum_cols
        _body.append("<tr>" + "".join(cells) + "</tr>")

    if _pm == "all":
        for lbl, sk in [("D인원", "D"), ("E인원", "E"), ("N인원", "N")]:
            hbg, hfg = _preview_shift_bg_fg(sk)
            bg, data_fg = hbg, hfg
            if sk in ("D", "E"):
                hbg = _PREVIEW_BG_DE
                hfg = _PREVIEW_FG_BLACK
                bg = _PREVIEW_BG_DE
                data_fg = _PREVIEW_FG_BLACK
            cells = [
                f'<td style="background:{hbg};color:{hfg};font-weight:700;'
                f'padding:4px 8px;white-space:nowrap;position:sticky;left:0;'
                f'border-right:2px solid #CFD8DC;">{lbl}</td>'
            ]
            for day in days:
                cnt = sum(1 for n in range(num) if schedule.get(n, {}).get(day["day"]) == sk)
                _wsp = _monday_week_split_style(day)
                cells.append(
                    f'<td style="background:{bg};color:{data_fg};font-weight:700;text-align:center;'
                    f'padding:3px;border:1px solid #E0E0E0;{_wsp}">{cnt}</td>'
                )
            cells += ["<td></td>"] * _n_sum_cols
            _body.append("<tr>" + "".join(cells) + "</tr>")
    elif _pm == "D":
        _append_summary_row("D인원", "D", color_key="D")
    elif _pm == "E":
        _append_summary_row("E인원", "E", color_key="E")
    elif _pm == "N":
        _append_summary_row("N인원", "N", color_key="N")
    elif _pm == "off":
        _append_summary_row(
            "OF·NO인원",
            lambda dn: sum(
                1 for n in range(num)
                if schedule.get(n, {}).get(dn) in ("OF", "NO")
            ),
            color_key="OF",
        )
        _append_summary_row("OH인원", "OH", color_key="OH")
        _append_summary_row("연인원", "연", color_key="연")

    return (
        '<div class="duty-generated-schedule-wrap" style="overflow-x:scroll;width:100%;max-width:100%;'
        'min-width:0;box-sizing:border-box;border-radius:10px;'
        'box-shadow:0 2px 12px rgba(0,0,0,0.09);-webkit-overflow-scrolling:touch;">'
        '<table style="border-collapse:collapse;font-size:12px;width:max-content;">'
        "<thead>" + _header_html + "</thead>"
        "<tbody>" + "".join(_body) + "</tbody>"
        "</table></div>"
    )


def _render_requests_preview_html(df: pd.DataFrame, nurse_names: list, days: list) -> str:
    """신청 근무 data_editor 내용 — 생성 근무표와 동일한 헤더·색상(합계 열 없음)."""
    col_labels = [_day_label_compact(d) for d in days]
    dfn = df.reindex(index=list(nurse_names), columns=col_labels, fill_value="")
    th = lambda txt, bg, extra="", fg="#37474F": (
        f'<th style="background:{bg};color:{fg};padding:4px 2px;'
        f'text-align:center;white-space:nowrap;{extra}">{txt}</th>'
    )
    _hdr: list[str] = ["<tr>"]
    _hdr.append(th("간호사", "#ECEFF1",
                   "min-width:80px;padding:5px 8px;font-size:11px;"
                   "position:sticky;left:0;z-index:5;", "#263238"))
    for day in days:
        if day["is_holiday"]:
            dbg, dfg = "#FFEBEE", "#C62828"
        elif day["is_weekend"]:
            dbg, dfg = "#E3F2FD", "#1565C0"
        else:
            dbg, dfg = "#F5F5F5", "#455A64"
        _wsp = _monday_week_split_style(day)
        _hdr.append(th(
            f"{day['day']}<br><span style='font-size:9px'>{day['weekday_name']}</span>",
            dbg, f"min-width:36px;{_wsp}", dfg,
        ))
    _hdr.append("</tr>")
    _header_html = "".join(_hdr)
    _body: list[str] = []
    for n_idx, name in enumerate(nurse_names):
        row_bg = "#FAFAFA" if n_idx % 2 == 0 else "#F5F5F5"
        cells = [
            f'<td style="background:#ECEFF1;color:#263238;font-weight:700;'
            f'padding:4px 8px;white-space:nowrap;position:sticky;left:0;'
            f'border-right:2px solid #CFD8DC;">{name}</td>'
        ]
        for j, day in enumerate(days):
            raw = dfn.iat[n_idx, j]
            shift = "" if (raw is None or str(raw).strip() in ("", "None", "nan")) else str(raw).strip()
            if shift:
                bg, fg = _preview_shift_bg_fg(shift)
            else:
                bg, fg = "#FFFFFF", "#BDBDBD"
            _wsp = _monday_week_split_style(day)
            cells.append(
                f'<td style="background:{bg};color:{fg};font-weight:700;'
                f'padding:3px 1px;text-align:center;border:1px solid #E0E0E0;{_wsp}">{shift}</td>'
            )
        _body.append(f'<tr style="background:{row_bg};">' + "".join(cells) + "</tr>")
    return (
        '<div class="duty-generated-schedule-wrap" style="overflow-x:scroll;width:100%;max-width:100%;'
        'min-width:0;box-sizing:border-box;border-radius:10px;'
        'box-shadow:0 2px 12px rgba(0,0,0,0.09);-webkit-overflow-scrolling:touch;">'
        '<table style="border-collapse:collapse;font-size:12px;width:max-content;">'
        "<thead>" + _header_html + "</thead>"
        "<tbody>" + "".join(_body) + "</tbody>"
        "</table></div>"
    )


def _show_schedule_preview_iframe(
    html_fragment: str, num_nurses: int, *, extra_rows: int = 5,
) -> None:
    """Streamlit 본문이 표 너비까지 늘어나 `st.markdown` 가로 스크롤이 안 생기는 경우 — iframe에서 스크롤."""
    # srcdoc 안에서 스크립트 태그로 잘못 해석되는 경우 방지
    safe = html_fragment.replace("</script>", "<\\/script>")
    # 간호사 행 + (생성표: 합계·요약 행 / 신청표: 헤더만) — extra_rows로 여유 조절
    _h = min(72 + max(num_nurses + extra_rows, 8) * 28, 1400)
    doc = (
        "<!DOCTYPE html><html><head><meta charset=\"utf-8\"/>"
        "<style>"
        "html,body{margin:0;padding:6px;background:#fafafa;}"
        "body{overflow:auto;overflow-x:auto;overflow-y:auto;-webkit-overflow-scrolling:touch;}"
        ".duty-generated-schedule-wrap{overflow:visible!important;width:max-content!important;"
        "max-width:none!important;min-width:0;}"
        "</style></head><body>"
        f"{safe}</body></html>"
    )
    components.html(doc, width=None, height=_h, scrolling=True)


def _schedule_to_edit_df(schedule: dict, nurse_names: list, days: list) -> pd.DataFrame:
    """schedule dict → data_editor용 DataFrame (행=간호사, 열=날짜)"""
    rows = []
    for n_idx, name in enumerate(nurse_names):
        row = {}
        for day in days:
            row[_day_label(day)] = schedule.get(n_idx, {}).get(day["day"], "")
        rows.append(row)
    return pd.DataFrame(rows, index=nurse_names)


def _edit_df_to_schedule(df: pd.DataFrame, days: list) -> dict:
    """data_editor DataFrame → schedule dict"""
    schedule = {}
    for n_idx in range(len(df)):
        schedule[n_idx] = {}
        for j, day in enumerate(days):
            val = str(df.iloc[n_idx, j]).strip()
            if val and val not in ("", "None", "nan"):
                schedule[n_idx][day["day"]] = val
    return schedule


def _clean_req_df(df: pd.DataFrame) -> pd.DataFrame:
    return df.apply(lambda col: col.map(_req_cell_str))


def _normalize_req_shift_cells(df: pd.DataFrame, allowed: frozenset[str]) -> pd.DataFrame:
    """저장·표시: Selectbox 옵션에 없는 값은 빈 칸으로 맞춤(연동 오류 방지)."""

    def cell(x: object) -> str:
        s = _req_cell_str(x)
        return s if s in allowed else ""

    return df.apply(lambda col: col.map(cell))


_REQ_SCHEDULE_BATCH_SAVE_TOAST = "✅ 신청 근무가 일괄 저장되었습니다."


def _snapshot_request_editor_for_save(
    base_df: pd.DataFrame,
    editor_key: str,
    return_df: pd.DataFrame | None,
) -> pd.DataFrame:
    """저장/생성 직전: st.data_editor의 session_state[editor_key](EditingState)와 반환 DF를 base에 병합.

    Streamlit 1.35+ 에서 session_state[key]는 edited_rows / added_rows / deleted_rows 를 담는 dict이다.
    반환 DataFrame은 위젯 상태가 이미 반영된 값이므로, 형태가 맞으면 최종본으로 우선한다.
    """
    idx_id = "_index"  # streamlit internal index column id
    out = base_df.copy()
    raw = st.session_state.get(editor_key)
    merged_any = False
    if isinstance(raw, dict):
        er = raw.get("edited_rows")
        if isinstance(er, dict) and er:
            merged_any = True
            for row_id, changes in er.items():
                try:
                    ri = int(row_id)
                except (TypeError, ValueError):
                    continue
                if ri < 0 or ri >= len(out):
                    continue
                for col_name, val in (changes or {}).items():
                    if col_name == idx_id:
                        old_lbl = out.index[ri]
                        new_lbl = _req_cell_str(val)
                        if new_lbl and str(new_lbl) != str(old_lbl):
                            out = out.rename(index={old_lbl: new_lbl})
                        continue
                    if col_name not in out.columns:
                        continue
                    j = out.columns.get_loc(col_name)
                    out.iat[ri, j] = _req_cell_str(val)
    if isinstance(return_df, pd.DataFrame):
        same_shape = len(return_df) == len(base_df) and list(return_df.columns) == list(
            base_df.columns
        )
        if same_shape:
            if merged_any:
                rt = _clean_req_df(return_df.copy())
                for c in out.columns:
                    if c in rt.columns:
                        out[c] = rt[c].tolist()
            else:
                out = _clean_req_df(return_df.copy())
        elif not merged_any:
            out = base_df.copy()
    return out


def _on_request_schedule_editor_change() -> None:
    """신청 근무 data_editor: 반환값 대신 session_state 위젯 dict를 파싱해 dept_requests를 갱신한다."""
    ctx = st.session_state.get("_req_editor_on_change_ctx")
    if not isinstance(ctx, dict):
        return
    editor_key = ctx.get("editor_key")
    dept = str(ctx.get("dept") or "").strip()
    period_pk = str(ctx.get("period_pk") or "").strip()
    crdf_key = ctx.get("crdf_key")
    if not editor_key or not dept or not period_pk:
        return
    raw = st.session_state.get(editor_key)
    if not isinstance(raw, dict):
        return
    er = raw.get("edited_rows")
    ar = raw.get("added_rows") or []
    dr = raw.get("deleted_rows") or []
    if (not isinstance(er, dict) or not er) and not ar and not dr:
        return
    _rq_sub = st.session_state.dept_requests.setdefault(dept, {})
    base_df = _rq_sub.get(period_pk)
    if base_df is None or not isinstance(base_df, pd.DataFrame):
        return
    _live_rq = _snapshot_request_editor_for_save(base_df, editor_key, None)
    _allowed = frozenset(REQUEST_SHIFT_OPTIONS)
    _live_rq = _normalize_req_shift_cells(_clean_req_df(_live_rq), _allowed)
    _live_rq = _live_rq.fillna("").apply(lambda c: c.map(_req_cell_str))
    hols = str(st.session_state.dept_holidays.get(dept, "") or "")
    days_cb = get_april_days(_parse_holidays(hols))
    req_col_labels_cb = [_day_label_compact(d) for d in days_cb]
    nurses = st.session_state.departments.get(dept, [])
    if not isinstance(nurses, list):
        nurses = []
    _raw_main = list(nurses)
    _cl_main = _clean_nurse_names_list(_raw_main)
    if not _cl_main:
        _cl_main = ["수간호사"]
    if _cl_main != _raw_main:
        st.session_state.departments[dept] = _cl_main
        nurses = _cl_main
        _save_hospital_config_to_disk()
    else:
        nurses = _cl_main
    _n_ext_main = _extend_nurses_to_dept_headcount(dept, list(nurses))
    if _n_ext_main != nurses:
        st.session_state.departments[dept] = _n_ext_main
        nurses = _n_ext_main
        _save_hospital_config_to_disk()
    else:
        nurses = _n_ext_main
    num_nurses = len(nurses)
    if (
        _live_rq.shape[0] == num_nurses
        and _live_rq.shape[1] == len(req_col_labels_cb)
        and list(_live_rq.columns) == list(req_col_labels_cb)
    ):
        _sd_rq = dept
        _nurses_req_row0 = list(nurses)
        _rq_sub[period_pk] = _live_rq
        st.session_state.dept_requests[_sd_rq] = _rq_sub
        _new_idx = [str(x).strip() if x is not None else "" for x in _live_rq.index.tolist()]
        _nurses_before = list(nurses)
        if len(_new_idx) == num_nurses and _new_idx != [str(x) for x in _nurses_before]:
            _fallback = [str(_nurses_before[i]) for i in range(num_nurses)]
            _pres = [
                _new_idx[i] if _new_idx[i] else _fallback[i]
                for i in range(num_nurses)
            ]
            _upd_n = _clean_nurse_names_list(_pres)
            if not _upd_n or len(_upd_n) != num_nurses:
                _upd_n = _pres
            if _upd_n != [str(x) for x in _nurses_before]:
                _fp = st.session_state.dept_forbidden_pairs.get(dept, [])

                def _fp_rq_names_ok(p):
                    ns = _fp_row_names_from_entry(p)
                    return bool(ns) and all(n in _upd_n for n in ns)

                st.session_state.dept_forbidden_pairs[dept] = [
                    p for p in _fp if _fp_rq_names_ok(p)
                ]
                _pg_rq = st.session_state.setdefault("dept_pregnant", {}).get(dept, [])
                if isinstance(_pg_rq, list):
                    st.session_state["dept_pregnant"][dept] = [
                        n for n in _pg_rq if n in _upd_n
                    ]
                _n4_rq = st.session_state.setdefault("dept_n_max4", {}).get(dept, [])
                if isinstance(_n4_rq, list):
                    st.session_state["dept_n_max4"][dept] = [
                        n for n in _n4_rq if n in _upd_n
                    ]
                st.session_state.departments[dept] = _upd_n
                _live_rq.index = list(_upd_n)
                _rq_sub[period_pk] = _live_rq
                st.session_state.dept_requests[_sd_rq] = _rq_sub
                _save_hospital_config_to_disk()
        _n_final_rq = _clean_nurse_names_list(
            list(st.session_state.departments.get(dept, _nurses_req_row0))
        )
        if len(_n_final_rq) != num_nurses:
            _n_final_rq = list(_nurses_req_row0)
        _live_rq_final = _prepare_requests_df_for_current_table(
            _live_rq, _n_final_rq, req_col_labels_cb
        )
        _rq_sub[period_pk] = _live_rq_final
        st.session_state.dept_requests[_sd_rq] = _rq_sub
        if crdf_key:
            st.session_state[str(crdf_key)] = _rq_sub[period_pk]


def _generate_excel(
    schedule,
    num_nurses,
    nurse_names,
    days,
    requests: dict | None = None,
) -> bytes:
    """미리보기(_render_schedule_html)와 동일: 일자 + 오른쪽 N/D/E/OH/OF/연 합계 6열 + D·E·N 일별 인원 행.

    requests: 간호사 인덱스 → {일(day 숫자): 신청 시프트}. 신청과 셀 값이 같으면 밑줄(신청 구분).
    """
    requests = requests or {}
    wb = openpyxl.Workbook(); ws = wb.active
    ws.title = f"{_app.YEAR}년 {_app.MONTH}월 근무표"
    ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _s_thin = Side(style="thin")
    thin = Border(left=_s_thin, right=_s_thin, top=_s_thin, bottom=_s_thin)
    # 엑셀: 월요일 칸 왼쪽 굵은 검정 세로선(일↔월 주 구분)
    _s_week = Side(style="medium", color="000000")

    def _excel_day_border(day: dict) -> Border:
        if day.get("weekday") == 0:
            return Border(left=_s_week, right=_s_thin, top=_s_thin, bottom=_s_thin)
        return thin

    def _xrgb(h: str) -> str:
        return h.replace("#", "").upper()

    def _px(sk: str) -> tuple[str, str]:
        """미리보기와 동일 (배경/글자 HEX, 알파벳 대문자 6자리)."""
        bg, fg = _preview_shift_bg_fg(sk)
        return _xrgb(bg), _xrgb(fg)

    _hdr_name_bg, _hdr_name_fg = _xrgb("#ECEFF1"), _xrgb("#263238")
    num_days = len(days)
    # 합계 열 순서: 미리보기와 동일 (N, D, E, OF, OH, 연)
    _sum_keys = ("N", "D", "E", "OF", "OH", "연")
    _sum_start = 2 + num_days   # 첫 합계 열 (1=이름, 2..num_days+1=일자)
    _last_col = _sum_start + len(_sum_keys) - 1

    year_label = _app.YEAR
    month_label = _app.MONTH
    ws.merge_cells(f"A1:{get_column_letter(_last_col)}1")
    c = ws["A1"]; c.value = f"{year_label}년 {month_label}월 근무표"
    c.fill = PatternFill("solid", fgColor=_hdr_name_bg); c.alignment = ctr
    c.font = Font(bold=True, size=14, color=_hdr_name_fg)
    ws.row_dimensions[1].height = 28

    h = ws.cell(2, 1, "간호사")
    h.fill = PatternFill("solid", fgColor=_hdr_name_bg)
    h.font = Font(bold=True, color=_hdr_name_fg, size=10)
    h.alignment = ctr; h.border = thin; ws.column_dimensions["A"].width = 11

    for d, day in enumerate(days):
        col = d + 2
        cell = ws.cell(2, col, f"{day['day']}\n{day['weekday_name']}")
        cell.alignment = ctr; cell.border = _excel_day_border(day)
        if day["is_holiday"]:
            bg, tfg = "FFEBEE", "C62828"
        elif day["is_weekend"]:
            bg, tfg = "E3F2FD", "1565C0"
        else:
            bg, tfg = "F5F5F5", "455A64"
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(bold=True, color=tfg, size=9)
        ws.column_dimensions[get_column_letter(col)].width = 4.5

    for i, sk in enumerate(_sum_keys):
        col = _sum_start + i
        c = ws.cell(2, col, f"{sk}\n합계"); c.alignment = ctr; c.border = thin
        _bg, _fg = _px(sk)
        c.fill = PatternFill("solid", fgColor=_bg)
        c.font = Font(bold=True, color=_fg, size=9)
        ws.column_dimensions[get_column_letter(col)].width = 5.5
    ws.row_dimensions[2].height = 28

    _day_lo = get_column_letter(2)
    _day_hi = get_column_letter(num_days + 1)

    def _sum_formula_for_row(row: int, sk: str) -> str:
        """해당 행의 일자 열 범위에 대한 합계(미리보기 집계와 동일). OF = OF+NO."""
        rng = f"${_day_lo}{row}:${_day_hi}{row}"
        if sk == "OF":
            return f'=COUNTIF({rng},"OF")+COUNTIF({rng},"NO")'
        return f'=COUNTIF({rng},"{sk}")'

    _first_body = 3
    _last_body = 2 + len(nurse_names)

    def _requested_shift(nr: dict, day_num: int):
        """신청 dict는 일자 키가 int 또는 str일 수 있음."""
        if not nr:
            return None
        return nr.get(day_num, nr.get(str(day_num)))

    for n_idx, name in enumerate(nurse_names):
        row = n_idx + 3
        nc = ws.cell(row, 1, name)
        nc.fill = PatternFill("solid", fgColor=_hdr_name_bg)
        nc.font = Font(bold=True, color=_hdr_name_fg, size=9)
        nc.alignment = ctr; nc.border = thin; ws.row_dimensions[row].height = 18
        ns = schedule.get(n_idx, {})
        nurse_req = requests.get(n_idx) or requests.get(str(n_idx)) or {}
        if not isinstance(nurse_req, dict):
            nurse_req = {}
        for d, day in enumerate(days):
            shift = ns.get(day["day"], ""); col = d + 2
            cell = ws.cell(row, col, shift); cell.alignment = ctr; cell.border = _excel_day_border(day)
            bg, fg = _px(shift)
            cell.fill = PatternFill("solid", fgColor=bg)
            dn = day["day"]
            is_requested = bool(shift) and (_requested_shift(nurse_req, dn) == shift)
            cell.font = Font(
                color=fg,
                size=9,
                bold=True,
                underline="single" if is_requested else None,
            )
        for i, sk in enumerate(_sum_keys):
            col = _sum_start + i
            bg, fg = _px(sk)
            c = ws.cell(row, col, _sum_formula_for_row(row, sk))
            c.alignment = ctr; c.border = thin
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(color=fg, bold=True, size=10)

    sr = len(nurse_names) + 3
    for idx, (lbl, sk) in enumerate([("D 인원", "D"), ("E 인원", "E"), ("N 인원", "N")]):
        row = sr + idx; lc = ws.cell(row, 1, lbl)
        lb, lf = _px(sk)
        if sk in ("D", "E"):
            lb, lf = _xrgb(_PREVIEW_BG_DE), _xrgb(_PREVIEW_FG_BLACK)
        lc.fill = PatternFill("solid", fgColor=lb); lc.font = Font(bold=True, color=lf, size=9)
        lc.alignment = ctr; lc.border = thin; ws.row_dimensions[row].height = 16
        for d in range(num_days):
            col = d + 2
            day = days[d]
            letter = get_column_letter(col)
            fml = f'=COUNTIF(${letter}${_first_body}:${letter}${_last_body},"{sk}")'
            cell = ws.cell(row, col, fml); cell.alignment = ctr; cell.border = _excel_day_border(day)
            cell.fill = PatternFill("solid", fgColor=lb)
            cell.font = Font(bold=True, color=lf, size=9)
        for j in range(len(_sum_keys)):
            ec = ws.cell(row, _sum_start + j, "")
            ec.alignment = ctr; ec.border = thin
            ec.fill = PatternFill("solid", fgColor=lb)
            ec.font = Font(bold=True, color=lf, size=9)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


# ════════════════════════════════════════════════════════════════════════════════
#  Duty Solution 브랜드 헤더 → 부서 로그인 (단일 암호)
# ════════════════════════════════════════════════════════════════════════════════
_render_app_brand_header()

_MONTH_NAMES = [
    "1월", "2월", "3월", "4월", "5월", "6월",
    "7월", "8월", "9월", "10월", "11월", "12월",
]

dept_list = _ordered_dept_keys(st.session_state.departments)
try:
    _dept_sb_idx = dept_list.index(st.session_state.active_dept)
except ValueError:
    _dept_sb_idx = 0

with st.container(border=True):
    st.markdown(
        '<p style="margin:0;padding:2px 0 4px 0;font-size:12px;font-weight:700;color:#37474F;">'
        "🔐 부서 로그인 — 암호는 <code>hospital_config.json</code> 해당 부서 설정과 동일해야 합니다.</p>",
        unsafe_allow_html=True,
    )
    _lc1, _lc2, _lc3, _lc4 = st.columns([2.1, 1.55, 0.5, 0.5], gap="small")
    with _lc1:
        _sel_dept_ui = st.selectbox(
            "부서",
            dept_list,
            index=_dept_sb_idx,
            key="dept_selectbox",
            label_visibility="collapsed",
        )
    st.session_state.active_dept = _sel_dept_ui
    _sync_selected_dept()
    with _lc2:
        st.text_input(
            "dept_pw_login",
            type="password",
            key="dept_password_input",
            placeholder="부서 암호",
            label_visibility="collapsed",
            autocomplete="current-password",
        )
    with _lc3:
        st.markdown("<div style='height:2px'></div>", unsafe_allow_html=True)
        if st.button("로그인", key="btn_dept_login", type="primary", use_container_width=True):
            _adr_l = str(st.session_state.active_dept).strip()
            st.session_state.setdefault("dept_meta", {})
            st.session_state.dept_meta.setdefault(_adr_l, _default_dept_meta())
            _secrets = _dept_login_secrets(st.session_state.dept_meta.get(_adr_l))
            _try_pw = (st.session_state.get("dept_password_input") or "").strip()
            if not _secrets:
                st.warning(
                    "이 부서에 저장된 암호가 없습니다. hospital_config.json에서 부서 코드를 설정한 뒤 다시 시도해 주세요."
                )
            elif _try_pw in _secrets:
                st.session_state.setdefault("dept_auth_ok", {})[_adr_l] = True
                st.session_state["_dept_login_flash"] = True
                st.session_state["_dept_login_flash_name"] = _adr_l
                st.rerun()
            else:
                st.error("부서 암호가 일치하지 않습니다.")
    with _lc4:
        st.markdown("<div style='height:2px'></div>", unsafe_allow_html=True)
        if st.button("로그아웃", key="btn_dept_logout_top", use_container_width=True):
            _adr_out = str(st.session_state.active_dept).strip()
            st.session_state.setdefault("dept_auth_ok", {}).pop(_adr_out, None)
            st.rerun()

active_dept = str(st.session_state.get("active_dept") or "").strip()
_auth_ok = bool(st.session_state.get("dept_auth_ok", {}).get(active_dept))
if not _auth_ok:
    st.session_state.pop("_pending_schedule_generate", None)
    st.session_state.show_violations = False

if st.session_state.pop("_dept_login_flash", None):
    _fn = str(st.session_state.pop("_dept_login_flash_name", "") or "").strip()
    if _fn:
        st.success(f"✅ [{_fn}] 인증되었습니다. 해당 부서 기능을 사용할 수 있습니다.")

if st.session_state.show_violations and _auth_ok:
    _show_violations_dialog()

st.markdown(
    '<div style="height:10px;min-height:10px;margin:0;padding:0;" aria-hidden="true"></div>',
    unsafe_allow_html=True,
)

# ════════════════════════════════════════════════════════════════════════════════
#  상단 설정 패널 (연·월 — 근무표·신청 표; 부서는 위 로그인 줄과 동일)
# ════════════════════════════════════════════════════════════════════════════════

_sync_selected_dept()

with st.container(border=True):
    _f1, _f2, _f3 = st.columns([5, 1, 1], gap="small")
    with _f2:
        sel_year = st.selectbox(
            "연도",
            list(range(2024, 2032)),
            index=list(range(2024, 2032)).index(st.session_state.sel_year),
            key="year_selectbox",
            label_visibility="collapsed",
        )
    with _f3:
        sel_month = st.selectbox(
            "월",
            list(range(1, 13)),
            index=st.session_state.sel_month - 1,
            format_func=lambda m: _MONTH_NAMES[m - 1],
            key="month_selectbox",
            label_visibility="collapsed",
        )
    with _f1:
        with st.container():
            st.markdown(
                '<div style="margin:0;padding:12px 8px 0 8px;box-sizing:border-box;width:100%;max-width:100%;">'
                '<p class="ds-main-schedule-title" '
                'style="margin:0 0 20px 0;padding:0;font-size:clamp(26px,5.5vw,42px);font-weight:800;'
                "color:#1E3A8A;line-height:1.15;word-wrap:break-word;overflow-wrap:break-word;max-width:100%;\">"
                '<span aria-hidden="true" style="font-size:1em;line-height:1;display:inline-block;">🗓️</span>'
                "&nbsp;교대근무간호사 근무표 생성</p></div>",
                unsafe_allow_html=True,
            )

    st.session_state.setdefault("dept_meta", {})
    st.session_state.dept_meta.setdefault(active_dept, _default_dept_meta())

    # 연·월 변경 시 앱 기간만 갱신 (신청·생성 근무는 부서×연월별로 유지)
    if sel_year != st.session_state.sel_year or sel_month != st.session_state.sel_month:
        st.session_state.sel_year = sel_year
        st.session_state.sel_month = sel_month
        _app.set_period(sel_year, sel_month)
        st.rerun()

    st.markdown(
        '<hr style="margin:0.06rem 0;border:none;border-top:1px solid #e0e0e0;">',
        unsafe_allow_html=True,
    )

    warning_list: list[str] = []
    _wq = st.session_state.get("_warning_queue")
    if isinstance(_wq, list) and _wq:
        warning_list.extend(x.strip() for x in _wq if isinstance(x, str) and x.strip())
        st.session_state["_warning_queue"] = []
    if _auth_ok and warning_list:
        with st.expander(
            f"⚠️ 기타 알림 {len(warning_list)}건",
            expanded=False,
        ):
            st.markdown("\n".join(f"- {line}" for line in warning_list))

    _can_manage_dept = bool(_auth_ok)
    st.session_state["dept_admin_verified"] = bool(_auth_ok)

    if not _auth_ok:
        st.info(
            "위에서 **부서를 선택**하고 **부서 암호**로 로그인하면 명단·근무표·신청 근무 기능을 사용할 수 있습니다.",
            icon="🔐",
        )

    nurses = st.session_state.departments[active_dept]
    if not isinstance(nurses, list):
        nurses = []
    _raw_roster = list(nurses)
    _cl_roster = _clean_nurse_names_list(_raw_roster)
    if not _cl_roster:
        _cl_roster = ["수간호사"]
    if _cl_roster != _raw_roster:
        st.session_state.departments[active_dept] = _cl_roster
        nurses = _cl_roster
        _save_hospital_config_to_disk()
    else:
        nurses = _cl_roster
    _nurse_ext = _extend_nurses_to_dept_headcount(active_dept, list(nurses))
    if _nurse_ext != nurses:
        st.session_state.departments[active_dept] = _nurse_ext
        nurses = _nurse_ext
        _save_hospital_config_to_disk()
    gen = st.session_state.nurse_gen.get(active_dept, 0)
    # 부서·연월 전환 시 파일 우선으로 이월 칸 동기화(앱 기동·리부팅 후에도 빈 {} 로 덮어쓰지 않음)
    _hydrate_carry_textarea_from_disk(active_dept, sel_year, sel_month)

    if _can_manage_dept:
        _r0b, _r0c, _r0d = st.columns([0.72, 0.75, 0.82], gap="small")
        with _r0b:
            with st.expander("➕ 부서", expanded=False):
                new_dept_input = st.text_input(
                    "부서 이름", key="new_dept_input",
                    placeholder="예: 본관 5병동",
                    label_visibility="collapsed",
                )
                if st.button("부서 추가", key="btn_add_dept", use_container_width=True):
                    name = new_dept_input.strip()
                    if not name:
                        _enqueue_warning("부서 이름을 입력하세요.")
                        st.rerun()
                    elif name in st.session_state.departments:
                        st.error("이미 존재하는 부서입니다.")
                    else:
                        st.session_state.departments[name] = _default_nurses(9)
                        st.session_state.dept_meta[name] = _default_dept_meta()
                        st.session_state.active_dept = name
                        _sync_selected_dept()
                        if "new_dept_input" in st.session_state:
                            st.session_state.new_dept_input = ""
                        st.rerun()

        with _r0c:
            with st.expander(f"👩 명단({len(nurses)})", expanded=False):
                st.markdown(
                    '<p class="roster-editor-hint" style="margin:0.2rem 0 0.95rem 0;padding:0 2px;'
                    'font-size:0.8rem;line-height:1.5;color:rgba(49,51,63,0.88);">'
                    "부서 로그인 후에만 수정 가능합니다. 표 <strong>+</strong> 로 행을 늘리거나 줄입니다. "
                    "첫 행은 수간호사로 쓰는 것을 권장합니다.</p>",
                    unsafe_allow_html=True,
                )
                _nurses_before_editor = list(nurses)
                _ndf = pd.DataFrame({"이름": list(nurses)})
                _ned = st.data_editor(
                    _ndf,
                    column_config={
                        "이름": st.column_config.TextColumn(
                            "이름",
                            help="수간호사·일반 간호사 이름 (변경 즉시 hospital_config.json 동기화)",
                            width=260,
                        )
                    },
                    num_rows="dynamic",
                    key=f"nurse_tbl_{active_dept}_g{gen}",
                    on_change=_on_nurse_roster_data_editor_change,
                    use_container_width=True,
                    hide_index=True,
                    disabled=False,
                )
                _cols = list(_ned.columns)
                _col_n = "이름" if "이름" in _cols else (_cols[0] if _cols else "이름")
                _raw_name_cells: list[object] = []
                for _, row in _ned.iterrows():
                    _cell = row[_col_n] if _col_n in row.index else None
                    _raw_name_cells.append(_cell)
                updated_nurses = _clean_nurse_names_list(_raw_name_cells)
                if not updated_nurses:
                    updated_nurses = ["수간호사"]
                _prev_len = len(nurses)
                _baseline_clean = _clean_nurse_names_list(list(_nurses_before_editor))
                _roster_cells_changed = _nurse_roster_dataframe_has_changes(
                    _ndf.reset_index(drop=True),
                    _ned.reset_index(drop=True),
                )
                _roster_names_changed = tuple(updated_nurses) != tuple(_baseline_clean)
                if _roster_cells_changed or _roster_names_changed:
                    _sync_roster_session_and_save_to_disk(
                        active_dept,
                        updated_nurses,
                        toast_on_success=True,
                    )
                else:
                    _filter_constraints_for_roster(active_dept, updated_nurses)
                    st.session_state.departments[active_dept] = list(updated_nurses)
                if len(updated_nurses) != _prev_len:
                    st.session_state.dept_requests[active_dept] = {}
                    st.session_state.dept_schedules[active_dept] = {}
                    st.session_state.nurse_gen[active_dept] = gen + 1
                    _delete_schedule_requests_dept(active_dept)
                    st.rerun()
                _rq_pk = _period_storage_key(sel_year, sel_month)
                _rq_sub = st.session_state.dept_requests.setdefault(active_dept, {})
                if not isinstance(_rq_sub, dict):
                    _rq_sub = {}
                    st.session_state.dept_requests[active_dept] = _rq_sub
                df_existing = _rq_sub.get(_rq_pk)
                if df_existing is not None and len(df_existing) == len(updated_nurses):
                    df_existing.index = updated_nurses

        with _r0d:
            with st.expander("📅 휴일", expanded=False):
                default_hols = st.session_state.dept_holidays.get(active_dept, "")
                holidays_raw = st.text_input(
                    "공휴일",
                    value=default_hols if default_hols is not None else "",
                    key=f"holidays_{active_dept}",
                    placeholder="",
                    label_visibility="collapsed",
                )
                st.session_state.dept_holidays[active_dept] = holidays_raw
                _hol_parsed = _parse_holidays(holidays_raw)
                if _hol_parsed:
                    badge = " · ".join(f"{h}일" for h in _hol_parsed)
                    st.markdown(
                        f'<div style="background:#E3F2FD;border:1px solid #90CAF9;border-radius:4px;'
                        f'padding:4px 8px;font-size:10px;color:#1565C0;">📌 {badge}</div>',
                        unsafe_allow_html=True,
                    )
    
        # 가로 2행: 함께 근무 불가·임산부 | 전월 이월 | 부서 삭제 | 근무표 생성
        _r1a, _r1b, _r1c, _r1d = st.columns([2.5, 1.72, 0.38, 1.05], gap="small")
        with _r1a:
            with st.expander("🙅 불가", expanded=False):
                st.markdown(
                    '<p style="font-size:11px;font-weight:600;margin:0 0 2px 0;color:#212121;">'
                    "🙅 함께 근무 불가</p>",
                    unsafe_allow_html=True,
                )
                st.markdown(
                    '<p class="fp-forbidden-help" style="font-size:10px;line-height:1.45;color:#616161;'
                    'margin:0 0 14px 0;padding-bottom:2px;">'
                    "<strong>수간호사 포함</strong> <strong>2~4명</strong>을 고릅니다. 선택한 사람들은 같은 날·같은 근무에 "
                    "동시에 배치되지 않습니다. 아래에서 <strong>D / E / N</strong> 중 적용할 근무를 고릅니다.</p>",
                    unsafe_allow_html=True,
                )
                _fp_list = st.session_state.dept_forbidden_pairs.setdefault(active_dept, [])
                st.markdown('<div class="fp-multiselect-anchor"></div>', unsafe_allow_html=True)
                _fp_pick = st.multiselect(
                    "👤 간호사 선택",
                    nurses,
                    key=f"fp_multi_{active_dept}",
                    max_selections=4,
                    label_visibility="visible",
                )
                st.markdown(
                    '<p style="font-size:11px;font-weight:600;color:#111111;margin:8px 0 4px 0;">적용 근무</p>',
                    unsafe_allow_html=True,
                )
                _fc1, _fc2, _fc3 = st.columns(3)
                with _fc1:
                    _chk_d = st.checkbox("D 근무 불가", value=True, key=f"fp_shift_d_{active_dept}")
                with _fc2:
                    _chk_e = st.checkbox("E 근무 불가", value=True, key=f"fp_shift_e_{active_dept}")
                with _fc3:
                    _chk_n = st.checkbox("N 근무 불가", value=True, key=f"fp_shift_n_{active_dept}")
                _fp_shift_sel = [s for s, ok in (("D", _chk_d), ("E", _chk_e), ("N", _chk_n)) if ok]
                if st.button("➕ 추가", key=f"fp_add_{active_dept}", use_container_width=True):
                    _nuniq = sorted(set(_fp_pick))
                    _fp_msg: str | None = None
                    if len(_nuniq) < 2:
                        _fp_msg = "2명 이상(최대 4명) 선택해 주세요."
                    elif len(_nuniq) > 4:
                        _fp_msg = "최대 4명까지 선택할 수 있습니다."
                    elif not _fp_shift_sel:
                        _fp_msg = "적용할 근무(D/E/N)를 하나 이상 선택해 주세요."
                    if _fp_msg:
                        _enqueue_warning(_fp_msg)
                        st.rerun()
                    else:
                        _gkey = tuple(_nuniq)
                        _shifts = sorted(_fp_shift_sel, key=lambda x: "DEN".index(x))
                        _found_i = None
                        for _ix, _row in enumerate(_fp_list):
                            _ex = _fp_row_names_from_entry(_row)
                            if _ex and tuple(_ex) == _gkey:
                                _found_i = _ix
                                break
                        if _found_i is not None:
                            _old = _fp_list[_found_i]
                            if isinstance(_old[0], list):
                                _prev = set(_old[1]) if len(_old) > 1 and isinstance(_old[1], list) else {"D", "E", "N"}
                            else:
                                _prev = (
                                    set(_old[2]) if len(_old) >= 3 and isinstance(_old[2], list) else {"D", "E", "N"}
                                )
                            _merged = sorted(_prev | set(_shifts), key=lambda x: "DEN".index(x))
                            _fp_list[_found_i] = [list(_nuniq), _merged]
                        else:
                            _fp_list.append([list(_nuniq), _shifts])
                        st.rerun()
                if _fp_list:
                    for _i, _pr in enumerate(list(_fp_list)):
                        _r1, _r2 = st.columns([5, 1])
                        with _r1:
                            _nm_disp = _fp_row_names_from_entry(_pr)
                            _lbl = " · ".join(_nm_disp) if _nm_disp else "?"
                            if isinstance(_pr[0], list):
                                _sh_disp = _pr[1] if len(_pr) > 1 and isinstance(_pr[1], list) else ["D", "E", "N"]
                            else:
                                _sh_disp = (
                                    _pr[2]
                                    if len(_pr) >= 3 and isinstance(_pr[2], list)
                                    else ["D", "E", "N"]
                                )
                            _tags = "".join(
                                f'<span style="display:inline-block;background:#ECEFF1;padding:1px 6px;'
                                f'border-radius:4px;margin:2px 4px 0 0;font-size:9px;">{s} 불가</span>'
                                for s in _sh_disp
                            )
                            st.markdown(
                                f'<div style="font-size:10px;color:#37474F;padding:1px 0;line-height:1.35;">'
                                f"🔗 {_lbl}<br/>{_tags}</div>",
                                unsafe_allow_html=True,
                            )
                        with _r2:
                            if st.button("삭제", key=f"fp_rm_{active_dept}_{_i}", use_container_width=True):
                                _fp_list.pop(_i)
                                st.rerun()
                else:
                    st.markdown(
                        '<p style="font-size:10px;color:#9E9E9E;margin:0;">등록된 쌍이 없습니다.</p>',
                        unsafe_allow_html=True,
                    )
    
                st.markdown(
                    '<hr style="margin:14px 0 10px 0;border:none;border-top:1px solid #E0E0E0;"/>'
                    '<p style="font-size:11px;font-weight:600;margin:0 0 4px 0;color:#212121;">'
                    "🤰 임산부 (법적·절대 규칙)</p>"
                    '<p style="font-size:10px;line-height:1.45;color:#616161;margin:0 0 10px 0;">'
                    "선택한 일반간호사는 <strong>나이트(N)에 절대 배정되지 않습니다</strong>. "
                    "N을 신청한 경우 생성 전에 신청을 수정해 주세요.</p>",
                    unsafe_allow_html=True,
                )
                _pg_map = st.session_state.setdefault("dept_pregnant", {})
                if active_dept not in _pg_map or not isinstance(_pg_map[active_dept], list):
                    _pg_map[active_dept] = []
                _pg_opts = nurses[1:] if len(nurses) > 1 else []
                _pg_prev = tuple(_pg_map[active_dept])
                _pg_sel = st.multiselect(
                    "👤 간호사 선택",
                    options=_pg_opts,
                    default=[n for n in _pg_map[active_dept] if n in _pg_opts],
                    key=f"preg_mu_{active_dept}_g{gen}",
                    label_visibility="visible",
                )
                if tuple(_pg_sel) != _pg_prev:
                    _pg_map[active_dept] = list(_pg_sel)

                st.markdown(
                    '<hr style="margin:14px 0 10px 0;border:none;border-top:1px solid #E0E0E0;"/>'
                    '<p style="font-size:11px;font-weight:600;margin:0 0 4px 0;color:#212121;">'
                    "🌙 N 최대 4개 제한</p>"
                    '<p style="font-size:10px;line-height:1.45;color:#616161;margin:0 0 10px 0;">'
                    "선택한 간호사는 해당 스케줄 기간 동안 <strong>나이트(N) 근무가 최대 4개까지만</strong> "
                    "배정됩니다 (5개 이상 배정 불가).</p>",
                    unsafe_allow_html=True,
                )
                _n4_map = st.session_state.setdefault("dept_n_max4", {})
                if active_dept not in _n4_map or not isinstance(_n4_map[active_dept], list):
                    _n4_map[active_dept] = []
                _n4_opts = nurses[1:] if len(nurses) > 1 else []
                _n4_prev = tuple(_n4_map[active_dept])
                _n4_sel = st.multiselect(
                    "👤 간호사 선택",
                    options=_n4_opts,
                    default=[n for n in _n4_map[active_dept] if n in _n4_opts],
                    key=f"n4_mu_{active_dept}_g{gen}",
                    label_visibility="visible",
                )
                if tuple(_n4_sel) != _n4_prev:
                    _n4_map[active_dept] = list(_n4_sel)

        with _r1b:
            with st.expander("📎 이월", expanded=False):
                _carry_ui_key = _carry_widget_session_key(active_dept, sel_year, sel_month)
                st.markdown(
                    '<p style="font-size:10px;line-height:1.4;color:#616161;margin:0 0 6px 0;">'
                    "이월 배열의 <strong>마지막 칸</strong>은 항상 <strong>당월 1일 전날</strong>로 간주됩니다. "
                    "4일이면 그 전 4일, 5일이면 5일이 가상 타임라인 앞부분(Day -L…-1)으로 합쳐집니다. "
                    "<code>hospital_config.json</code>의 <code>departments[부서].last_month_shifts</code>에 "
                    "부서·당월(생성 월) 기준으로 저장되며, JSON 오류 시 파일은 건드리지 않습니다.</p>",
                    unsafe_allow_html=True,
                )
                if CARRY_AUTO_FROM_ARCHIVE_ENABLED:
                    if st.button(
                        f"📥 직전 달 마지막 {CARRY_AUTO_DAYS}일 자동",
                        key=f"btn_carry_auto_{active_dept}",
                        use_container_width=True,
                    ):
                        _co, _em = _build_carry_from_prev_month(
                            active_dept, sel_year, sel_month, nurses, CARRY_AUTO_DAYS,
                        )
                        if _em:
                            _enqueue_warning(_em)
                            st.rerun()
                        else:
                            st.session_state[_carry_ui_key] = json.dumps(
                                {str(k): v for k, v in _co.items()},
                                ensure_ascii=False,
                            )
                            st.toast(
                                f"✅ 직전 달 마지막 {CARRY_AUTO_DAYS}일을 반영했습니다.",
                                icon="📎",
                            )
                            st.rerun()
                else:
                    st.caption(
                        "직전 달 마지막 일수 **자동 반영**은 현재 사용하지 않습니다. "
                        "전월 말 근무는 아래 JSON에만 입력해 주세요."
                    )
                _cpy, _cpm = _prev_year_month(sel_year, sel_month)
                st.caption(f"저장분: **{_cpy}년 {_cpm}월** (아래 JSON·미리보기는 **{sel_year}년 {sel_month}월** 생성 기준)")
                with st.expander("📋 부서별 이월 저장·적용 현황 (이번 표시 월)", expanded=False):
                    for _dn in _ordered_dept_keys(st.session_state.departments):
                        st.markdown(
                            "<p style=\"font-size:10px;margin:3px 0;line-height:1.35;color:#37474F;\">"
                            f"<strong>{_dn}</strong> — {_dept_carry_status_line(_dn, sel_year, sel_month)}</p>",
                            unsafe_allow_html=True,
                        )
                st.text_area(
                    "전월 말 JSON",
                    height=90,
                    key=_carry_ui_key,
                    placeholder=(
                        '{"0": ["OF"], "1": ["N", "N", "OF"], "2": ["D", "E"]}  ← 수간=0, 간호사=1…'
                    ),
                    label_visibility="collapsed",
                )
                if st.button(
                    "💾 이월 근무 hospital_config 저장",
                    key=f"btn_save_carry_{active_dept}_{sel_year}_{sel_month}",
                    use_container_width=True,
                    help="현재 부서 블록에 last_month_shifts 를 저장합니다. JSON 검증 실패 시 디스크에 쓰지 않습니다.",
                ):
                    _raw_c = (st.session_state.get(_carry_ui_key) or "").strip()
                    _c_ok, _c_msg = _persist_department_last_month_to_hospital_config(
                        active_dept,
                        int(sel_year),
                        int(sel_month),
                        _raw_c,
                        nurses,
                    )
                    if _c_ok:
                        st.success("✅ 이월 근무를 hospital_config.json(해당 부서)에 저장했습니다.")
                    else:
                        st.error(_c_msg)
                _carry_pv_raw = _parse_carry_in_text(
                    (st.session_state.get(_carry_ui_key) or "").strip(),
                    nurses,
                )
                _carry_pv = _merge_carry_with_hospital_last_month(
                    _carry_pv_raw,
                    _load_hospital_config_bundle(),
                    active_dept,
                    int(sel_year),
                    int(sel_month),
                    nurses,
                )
                _vt_cap = _carry_virtual_timeline_caption(
                    int(sel_year), int(sel_month), _carry_pv,
                )
                if _vt_cap:
                    st.caption(_vt_cap)
                if _carry_pv and _carry_pv is not False:
                    _pv_lines: list[str] = []
                    for _ci in range(len(nurses)):
                        _sq = _carry_pv.get(_ci, [])
                        if _sq:
                            _pv_lines.append(
                                f"{nurses[_ci]} — " + " · ".join(str(s) for s in _sq)
                            )
                    if _pv_lines:
                        st.markdown(
                            '<p style="font-size:10px;font-weight:600;color:#1565C0;margin:8px 0 4px 0;">'
                            f"전월 이월 근무 요약 "
                            f"(선택 부서 <strong>{active_dept}</strong> 명단만)</p>",
                            unsafe_allow_html=True,
                        )
                        st.markdown("\n".join(f"- {__ln}" for __ln in _pv_lines))
    
        with _r1c:
            st.markdown("<div style='min-height:1rem'></div>", unsafe_allow_html=True)
            if len(dept_list) > 1:
                if st.button(
                    "🗑️",
                    key="btn_del_dept",
                    help=f"'{active_dept}' 부서 삭제 (데이터 포함)",
                ):
                    for store in ("dept_schedules", "dept_requests", "dept_holidays", "nurse_gen"):
                        st.session_state[store].pop(active_dept, None)
                    _delete_schedule_requests_dept(active_dept)
                    _dfb = st.session_state.get("dept_forbidden_pairs")
                    if isinstance(_dfb, dict):
                        _dfb.pop(active_dept, None)
                    _dpg = st.session_state.get("dept_pregnant")
                    if isinstance(_dpg, dict):
                        _dpg.pop(active_dept, None)
                    _dn4 = st.session_state.get("dept_n_max4")
                    if isinstance(_dn4, dict):
                        _dn4.pop(active_dept, None)
                    _dm = st.session_state.get("dept_meta")
                    if isinstance(_dm, dict):
                        _dm.pop(active_dept, None)
                    _dauth = st.session_state.get("dept_auth_ok")
                    if isinstance(_dauth, dict):
                        _dauth.pop(active_dept, None)
                    del st.session_state.departments[active_dept]
                    st.session_state.active_dept = list(st.session_state.departments.keys())[0]
                    _sync_selected_dept()
                    st.rerun()
            else:
                st.caption("—")
    
        with _r1d:
            _hint_pk = _period_storage_key(st.session_state.sel_year, st.session_state.sel_month)
            _hint_sub = st.session_state.dept_schedules.get(active_dept, {})
            _has_sched = isinstance(_hint_sub, dict) and bool(_hint_sub.get(_hint_pk))
            _gen_lbl = "🗓️ 재생성" if _has_sched else "🗓️ 생성"
            if st.button(
                _gen_lbl,
                type="primary",
                use_container_width=True,
                key="btn_generate",
                help="신청으로 적은 칸은 유지하고, 자동 배정 칸만 규칙에 맞게 다시 짭니다. 재생성마다 패턴이 달라질 수 있습니다.",
            ):
                st.session_state["_pending_schedule_generate"] = True
            if _has_sched:
                st.caption("✅ 생성됨 · 재생성 가능")

    _roster_readonly = not _auth_ok
    if _roster_readonly:
        with st.expander(f"👩 명단 ({len(nurses)}명 · 열람 전용)", expanded=False):
            st.markdown(
                '<p class="roster-editor-hint" style="margin:0.2rem 0 0.95rem 0;padding:0 2px;'
                'font-size:0.8rem;line-height:1.5;color:rgba(49,51,63,0.88);">'
                "이름·인원 수정은 위에서 해당 부서 <strong>로그인</strong> 후에만 가능합니다.</p>",
                unsafe_allow_html=True,
            )
            _ndf_ro = pd.DataFrame({"이름": list(nurses)})
            st.data_editor(
                _ndf_ro,
                column_config={
                    "이름": st.column_config.TextColumn("이름", width=260),
                },
                num_rows="fixed",
                key=f"nurse_tbl_ro_{active_dept}_g{gen}",
                use_container_width=True,
                hide_index=True,
                disabled=True,
            )

    if _auth_ok:
        st.markdown(
            '<p style="font-size:11px;color:#B71C1C;margin:10px 0 6px 0;line-height:1.5;">'
            "⚠️ 본 명단에는 개인정보가 포함되어 있으므로, 부서 암호가 유출되지 않게 주의하십시오.</p>",
            unsafe_allow_html=True,
        )

    holidays = _parse_holidays(st.session_state.dept_holidays.get(active_dept, ""))


_show_req_ui = bool(_auth_ok)
if not _show_req_ui:
    st.info(
        "📋 **신청 근무** 영역은 위에서 해당 부서 **로그인**을 완료한 뒤에만 아래에 열립니다."
    )
    st.stop()

# ════════════════════════════════════════════════════════════════════════════════
#  MAIN – 변수 준비
# ════════════════════════════════════════════════════════════════════════════════
st.session_state.selected_dept = active_dept
nurses      = st.session_state.departments[active_dept]   # 최신 명단 (수간호사 포함 총원)
if not isinstance(nurses, list):
    nurses = []
_raw_main = list(nurses)
_cl_main = _clean_nurse_names_list(_raw_main)
if not _cl_main:
    _cl_main = ["수간호사"]
if _cl_main != _raw_main:
    st.session_state.departments[active_dept] = _cl_main
    nurses = _cl_main
    _save_hospital_config_to_disk()
else:
    nurses = _cl_main
_n_ext_main = _extend_nurses_to_dept_headcount(active_dept, list(nurses))
if _n_ext_main != nurses:
    st.session_state.departments[active_dept] = _n_ext_main
    nurses = _n_ext_main
    _save_hospital_config_to_disk()
num_nurses  = len(nurses)  # 예: 11이면 수간 1 + 일반간호사 10
days        = get_april_days(holidays)
# 신청 근무 표는 짧은 열 제목(한 화면에 한 달)
req_col_labels = [_day_label_compact(d) for d in days]
gen         = st.session_state.nurse_gen.get(active_dept, 0)
_period_pk  = _period_storage_key(st.session_state.sel_year, st.session_state.sel_month)
# 위젯 키: 부서·연월·명단·세대별로 분리 (전환 시 편집 누적 방지). 저장 시 session_state[key] edited_rows + 반환 DF 병합.
_req_editor_widget_key = f"request_editor__{active_dept}__{_period_pk}__n{num_nurses}__g{gen}"

# requests_df 준비 — 부서·연월·gen 변경 시 파일에서 자동 시드 → 세션 유효 시 유지 → 아카이브 → 빈 표
# 부서 키는 항상 selected_dept(= active_dept)와 동일하게 사용
_rq_sub = st.session_state.dept_requests.setdefault(st.session_state.selected_dept, {})
if not isinstance(_rq_sub, dict):
    _rq_sub = {}
    st.session_state.dept_requests[st.session_state.selected_dept] = _rq_sub

_ls_obj = _duty_local_storage()
if _ls_obj is None:
    _req_arch = _load_schedule_requests_archive()
else:
    _req_arch = _requests_archive_from_local_storage(_ls_obj)

# 부서·연월·명단 세대(gen)가 바뀐 경우에만 디스크에서 시드(매 rerun마다 시드하면 편집 직후 값이 이전 JSON로 덮임)
_req_disk_seed_tuple = (
    str(st.session_state.selected_dept).strip(),
    str(_period_pk),
    int(gen),
)
if st.session_state.get("_req_disk_seed_ctx_done") != _req_disk_seed_tuple:
    st.session_state["_req_disk_seed_ctx_done"] = _req_disk_seed_tuple
    _seed_df = _try_load_requests_from_saved_sources(
        st.session_state.selected_dept, _period_pk, nurses, req_col_labels, _req_arch
    )
    if _seed_df is not None:
        _seed_df = _prepare_requests_df_for_current_table(_seed_df, nurses, req_col_labels)
        _rq_sub[_period_pk] = _seed_df
        st.session_state.dept_requests[str(st.session_state.selected_dept).strip()] = _rq_sub
        st.session_state["request_editor_key"] = _req_editor_widget_key
        st.session_state.pop(_req_editor_widget_key, None)
        st.session_state.pop("request_editor", None)

if st.session_state.pop("_force_ls_reload", False):
    # 🔄 이전 기록: shift_requests.json 키 `{부서}_{연도}_{월}` → 없으면 hospital_config → schedule_requests.json
    _sd_reload = str(st.session_state.selected_dept).strip()
    _yr = int(st.session_state.sel_year)
    _mo = int(st.session_state.sel_month)
    _current_key = _shift_requests_period_key(_sd_reload, _yr, _mo)
    _df_reload_raw = _try_load_requests_from_shift_requests_json(
        _sd_reload, _yr, _mo, nurses, req_col_labels,
    )
    if _df_reload_raw is None:
        _df_reload_raw = _load_requests_dataframe_for_selected_dept(
            _sd_reload, _period_pk, nurses, req_col_labels, _req_arch,
        )
    if _df_reload_raw is not None:
        _df_reload = _prepare_requests_df_for_current_table(
            _df_reload_raw, nurses, req_col_labels
        )
        _rq_sub[_period_pk] = _df_reload
        st.session_state.dept_requests[_sd_reload] = _rq_sub
        st.session_state["request_editor_key"] = _req_editor_widget_key
        for _ek in (_req_editor_widget_key, "request_editor"):
            st.session_state.pop(_ek, None)
        st.session_state["_req_ls_load_ok_msg"] = True
        st.rerun()
    st.session_state["_req_ls_load_warn_msg"] = True
    st.rerun()

df_req = _rq_sub.get(_period_pk)
_col_ok = (
    df_req is not None
    and df_req.shape[0] == num_nurses
    and df_req.shape[1] == len(req_col_labels)
    and list(df_req.columns) == list(req_col_labels)
    and [str(x) for x in df_req.index] == [str(x) for x in nurses]
)
if not _col_ok:
    if df_req is None:
        # 세션에 해당 월 데이터가 없을 때만 디스크/아카이브에서 시드 (편집 중인 작업본을 매 rerun 덮어쓰지 않음)
        _saved_src_df = _try_load_requests_from_saved_sources(
            st.session_state.selected_dept, _period_pk, nurses, req_col_labels, _req_arch
        )
        if _saved_src_df is not None:
            df_req = _prepare_requests_df_for_current_table(
                _saved_src_df, nurses, req_col_labels
            )
            _rq_sub[_period_pk] = df_req
            _auto_tk = f"_ls_auto_loaded_toast_{active_dept}_{_period_pk}"
            if not st.session_state.get(_auto_tk):
                st.session_state[_auto_tk] = True
                st.session_state["_req_ls_load_ok_msg"] = True
        else:
            df_req = _make_requests_df(nurses, days)
            _rq_sub[_period_pk] = df_req
    else:
        # 이미 세션 DF가 있는데 행·열만 어긋난 경우: JSON 재주입 없이 보간
        try:
            _same_n = len(df_req) == num_nurses
            _cols_match = list(df_req.columns) == list(req_col_labels)
            _set_idx = {str(x) for x in df_req.index}
            _set_nu = {str(x) for x in nurses}
            if _same_n and _cols_match and _set_idx == _set_nu:
                df_req = (
                    df_req.reindex(index=list(nurses), columns=list(req_col_labels), fill_value="")
                    .fillna("")
                    .apply(lambda c: c.map(_req_cell_str))
                )
            elif _same_n and _cols_match and _set_idx != _set_nu:
                df_req = _prepare_requests_df_for_current_table(
                    df_req, nurses, req_col_labels
                )
            else:
                df_req = (
                    df_req.reindex(index=list(nurses), columns=list(req_col_labels), fill_value="")
                    .fillna("")
                    .apply(lambda c: c.map(_req_cell_str))
                )
            _rq_sub[_period_pk] = df_req
        except (TypeError, ValueError, KeyError):
            df_req = _make_requests_df(nurses, days)
            _rq_sub[_period_pk] = df_req
else:
    # 이미 행·열·인덱스가 명단과 일치: 매 rerun마다 _prepare 로 새 DataFrame을 만들면
    # st.data_editor 가 동일 key 라도 내부 스냅샷과 어긋나 깜빡임·포커스 상실이 난다.
    pass

# 신청 표 원본은 _rq_sub[_period_pk] 단일 객체로 유지. data_editor 직전 fillna/copy 금지(포커스 튐 방지).
_crdf_key = f"current_req_df__{active_dept}__{_period_pk}__g{gen}"
st.session_state[_crdf_key] = _rq_sub[_period_pk]

if st.session_state.pop("_req_ls_load_ok_msg", None):
    st.success("신청 근무를 성공적으로 불러왔습니다.")
if st.session_state.pop("_req_ls_load_warn_msg", None):
    st.warning("해당 부서/월에 저장된 기록이 없습니다.")

_inject_week_split_css(days)

# ════════════════════════════════════════════════════════════════════════════════
#  MAIN – 생성된 근무표
# ════════════════════════════════════════════════════════════════════════════════
_sched_sub = st.session_state.dept_schedules.get(active_dept, {})
sched_data = _sched_sub.get(_period_pk) if isinstance(_sched_sub, dict) else None

if _can_manage_dept and sched_data:
    schedule    = sched_data["schedule"]
    sched_names = sched_data["nurse_names"]
    sched_hols  = sched_data["holidays"]
    sched_days  = get_april_days(sched_hols)
    # 검증·표시용 총원 = 저장된 명단 길이(솔버·validate_schedule 동기화의 단일 기준)
    sched_n     = len(sched_names)
    sched_reqs  = sched_data.get("requests", {})

    # 세션 위반 목록 동기화: 이월은 검증에만 반영(솔버와 동일하게 당월 하드는 이미 반영된 표)
    _fp_sched_v = _fp_pairs_to_indices(
        sched_names,
        st.session_state.get("dept_forbidden_pairs", {}).get(active_dept, []),
    )
    _cr_raw_sv = (
        st.session_state.get(
            _carry_widget_session_key(
                active_dept,
                int(st.session_state.sel_year),
                int(st.session_state.sel_month),
            ),
            "",
        )
        or ""
    )
    _cr_p_sv = _parse_carry_in_text(_cr_raw_sv, sched_names)
    _carry_sched_v = _merge_carry_with_hospital_last_month(
        None if _cr_p_sv is False else _cr_p_sv,
        _load_hospital_config_bundle(),
        active_dept,
        int(st.session_state.sel_year),
        int(st.session_state.sel_month),
        sched_names,
    )
    if _carry_sched_v is False:
        _carry_sched_v = None
    _n4_v_raw = st.session_state.get("dept_n_max4", {}).get(active_dept, [])
    _n4_for_validate = _n4_v_raw if isinstance(_n4_v_raw, list) and _n4_v_raw else None
    st.session_state.violations = validate_schedule(
        schedule,
        len(sched_names),
        sched_hols,
        forbidden_pairs=_fp_sched_v or None,
        nurse_names=sched_names,
        carry_in=_carry_sched_v,
        requests=sched_reqs or None,
        unit_profile=_effective_unit_profile(active_dept),
        n_max4_nurses=_n4_for_validate,
    )

    st.markdown("<hr>", unsafe_allow_html=True)

    # 검토·경고: ⚠️ 버튼·검토 모달·기타 알림 Expander (중복 st.warning 제거)

    # ── 수정 모드 (✏️ 눌렀을 때만 편집 표 — 평소는 컬러 미리보기만)
    _em_sub = st.session_state.edit_mode.setdefault(active_dept, {})
    if not isinstance(_em_sub, dict):
        _em_sub = {}
        st.session_state.edit_mode[active_dept] = _em_sub
    is_edit = bool(_em_sub.get(_period_pk, False))

    # ── 헤더 버튼 행 ───────────────────────────────────────────────────────────
    col_t, col_edit, col_vld, col_dl = st.columns([3, 1, 1, 1])
    with col_t:
        edit_badge = ' <span style="color:#E65100;font-size:12px;">✏️ 수정 중</span>' if is_edit else ""
        st.markdown(f"""
        <div class="card" style="padding:12px 20px;margin-bottom:8px;">
          <div class="card-title">📋 생성된 근무표 &nbsp;
            <span class="dept-badge">{active_dept}</span>{edit_badge}
          </div>
          <div class="card-sub">{_app.YEAR}년 {_app.MONTH}월 · {sched_n}명</div>
        </div>
        """, unsafe_allow_html=True)

    with col_edit:
        st.markdown("<div style='margin-top:18px'></div>", unsafe_allow_html=True)
        if not is_edit:
            st.button(
                "✏️ 수정",
                use_container_width=True,
                key=f"btn_sched_edit_on_{active_dept}_{_period_pk}",
                on_click=_set_schedule_edit_mode,
                args=(active_dept, _period_pk, True),
            )
        else:
            st.button(
                "❌ 취소",
                use_container_width=True,
                key=f"btn_sched_edit_off_{active_dept}_{_period_pk}",
                on_click=_set_schedule_edit_mode,
                args=(active_dept, _period_pk, False),
            )

    with col_vld:
        st.markdown("<div style='margin-top:18px'></div>", unsafe_allow_html=True)
        vld_issues = st.session_state.get("violations", [])
        err_cnt  = sum(1 for v in vld_issues if v["level"] == "error")
        warn_cnt = sum(1 for v in vld_issues if v["level"] == "warn")
        btn_label = (
            "✅ 규칙 통과" if not vld_issues
            else f"⚠️ {err_cnt}오류/{warn_cnt}경고"
        )
        if st.button(
            btn_label,
            use_container_width=True,
            key=f"btn_violations_{active_dept}_{_period_pk}",
        ):
            st.session_state.show_violations = True
            st.rerun()

    with col_dl:
        st.markdown("<div style='margin-top:18px'></div>", unsafe_allow_html=True)
        excel_bytes = _generate_excel(
            schedule, sched_n, sched_names, sched_days, requests=sched_reqs or {}
        )
        st.download_button(
            "📥 엑셀 다운로드", data=excel_bytes,
            file_name=f"{_app.YEAR}년_{_app.MONTH}월_근무표_{active_dept}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    _preview_mode_labels = {
        "all": "전체 보기",
        "D": "D(데이)만",
        "E": "E(이브닝)만",
        "N": "N(나이트)만",
        "off": "OF·OH·NO·연(휴게)만",
    }
    _pm_sel = st.radio(
        "미리보기 표시",
        ("all", "D", "E", "N", "off"),
        format_func=lambda k: _preview_mode_labels[k],
        key=f"sched_preview_filter_{active_dept}_{_period_pk}",
        horizontal=True,
        help="선택한 근무 유형만 강조하고, 나머지 날짜 칸은 회색으로 숨깁니다.",
    )
    _show_schedule_preview_iframe(
        _render_schedule_html(
            schedule,
            sched_names,
            sched_days,
            sched_reqs,
            preview_mode=_pm_sel,
        ),
        sched_n,
    )

    if is_edit:
        st.caption(
            "아래 편집 표는 **항상 전체 근무**가 보입니다. "
            "위 색상 표에서 필터로 D·E·N 등만 골라 확인할 수 있습니다. "
            "편집 내용은 **💾 저장** 후 반영됩니다."
        )
        st.info("셀을 클릭하면 근무를 변경할 수 있습니다. 수정 후 **💾 저장**을 눌러 확정하세요.", icon="✏️")
        st.markdown(
            '<div class="duty-schedule-editor-hscroll" aria-hidden="true"></div>',
            unsafe_allow_html=True,
        )
        _sched_shift_options = [""] + SHIFT_NAMES
        edit_df = _schedule_to_edit_df(schedule, sched_names, sched_days)
        col_cfg = {
            lbl: st.column_config.SelectboxColumn(lbl, options=_sched_shift_options, width="small")
            for lbl in edit_df.columns
        }
        edited = st.data_editor(
            edit_df,
            column_config=col_cfg,
            use_container_width=True,
            height=min(42 * sched_n + 90, 700),
            key=f"sched_editor_{active_dept}_{_period_pk}",
            num_rows="fixed",
        )
        save_col, _ = st.columns([1, 3])
        with save_col:
            if st.button(
                "💾 저장",
                type="primary",
                use_container_width=True,
                key=f"btn_save_sched_edit_{active_dept}_{_period_pk}",
            ):
                new_schedule = _edit_df_to_schedule(edited, sched_days)
                st.session_state.dept_schedules.setdefault(active_dept, {})[_period_pk]["schedule"] = new_schedule
                _archive_put_month(
                    active_dept,
                    st.session_state.sel_year,
                    st.session_state.sel_month,
                    sched_names,
                    new_schedule,
                )
                _fp_ed = _fp_pairs_to_indices(
                    sched_names,
                    st.session_state.get("dept_forbidden_pairs", {}).get(active_dept, []),
                )
                _cr_p_ed = _parse_carry_in_text(
                    st.session_state.get(
                        _carry_widget_session_key(
                            active_dept,
                            int(st.session_state.sel_year),
                            int(st.session_state.sel_month),
                        ),
                        "",
                    )
                    or "",
                    sched_names,
                )
                _carry_ed = _merge_carry_with_hospital_last_month(
                    None if _cr_p_ed is False else _cr_p_ed,
                    _load_hospital_config_bundle(),
                    active_dept,
                    int(st.session_state.sel_year),
                    int(st.session_state.sel_month),
                    sched_names,
                )
                if _carry_ed is False:
                    _carry_ed = None
                _n4_ed_raw = st.session_state.get("dept_n_max4", {}).get(active_dept, [])
                _n4_for_ed = _n4_ed_raw if isinstance(_n4_ed_raw, list) and _n4_ed_raw else None
                issues = validate_schedule(
                    new_schedule, len(sched_names), sched_hols,
                    forbidden_pairs=_fp_ed or None,
                    nurse_names=sched_names,
                    carry_in=_carry_ed,
                    requests=sched_reqs or None,
                    unit_profile=_effective_unit_profile(active_dept),
                    n_max4_nurses=_n4_for_ed,
                )
                st.session_state.violations     = issues
                st.session_state.show_violations = False
                _em_sub[_period_pk] = False
                if issues:
                    err_c = sum(1 for v in issues if v["level"] == "error")
                    war_c = sum(1 for v in issues if v["level"] == "warn")
                    st.toast(f"💾 저장 완료 — 위반 {err_c}오류/{war_c}경고 발견", icon="⚠️")
                else:
                    st.toast("💾 저장 완료! 모든 규칙 통과", icon="✅")
                st.rerun()

# ════════════════════════════════════════════════════════════════════════════════
#  MAIN – 신청 근무 입력 달력 (부서 로그인 후 이 섹션부터 표시)
# ════════════════════════════════════════════════════════════════════════════════
_req_fb = st.session_state.pop("_req_save_feedback", None)
if isinstance(_req_fb, dict) and not _req_fb.get("disk_verify", True):
    st.warning(
        "hospital_config.json에서 방금 저장한 신청 근무를 다시 읽지 못했습니다. "
        "파일 경로·권한을 확인해 주세요."
    )

st.markdown(
    '<hr style="margin:6px 0 4px 0;border:none;border-top:1.5px solid #90A4AE;">',
    unsafe_allow_html=True,
)
st.markdown(
    f"""
<div class="card" style="padding:6px 10px;margin-bottom:4px;">
  <div class="card-title" style="font-size:14px;margin-bottom:2px;line-height:1.15;">📝 신청 근무 입력 &nbsp;
    <span class="dept-badge" style="font-size:10px;padding:2px 8px;">{active_dept}</span>
  </div>
</div>
""",
    unsafe_allow_html=True,
)

# 범례 (작은 칩 형태)
legend_items = [
    ("A1","수간호사"), ("D","데이"), ("E","이브닝"), ("N","나이트"),
    ("OF","휴무"), ("OH","휴일"),
    ("NO","N 누적 20회마다 생기는 휴무. 발생일은 사람마다 다름(약 3개월에 1회 수준). 자동배정 없음·수기 입력."),
    ("연","연차"), ("병","병가"), ("공","공가"), ("경","경조"), ("EDU","교육"),
]
_leg_chips = []
for shift, tip in legend_items:
    bg, fg = _preview_shift_bg_fg(shift)
    _leg_chips.append(
        f'<span title="{tip}" style="display:inline-block;background:{bg};color:{fg};'
        f'text-align:center;padding:0 3px;margin:0 1px 1px 0;border-radius:2px;'
        f'font-size:8px;font-weight:700;line-height:1.2;">{shift}</span>'
    )
st.markdown(
    f'<div style="display:flex;flex-wrap:wrap;align-items:center;gap:0;margin:0 0 2px 0;">'
    f'{"".join(_leg_chips)}</div>',
    unsafe_allow_html=True,
)

# data_editor (행高·헤더 최소화로 한 달 컬럼 한 화면에 가깝게)
_req_shift_allowed = frozenset(REQUEST_SHIFT_OPTIONS)
col_config = {
    lbl: st.column_config.SelectboxColumn(
        lbl,
        options=list(REQUEST_SHIFT_OPTIONS),
        width="small",
        required=False,
    )
    for lbl in req_col_labels
}
# 행高约 16px 목표 → 세로로 간호사 전원 한 화면에 가깝게
_req_table_h = min(16 * num_nurses + 44, 580)

# 미로그인 시 신청 표 편집 비활성화
_dept_adm_ok = bool(st.session_state.get("dept_admin_verified"))
_req_editor_disabled = not _auth_ok
st.session_state["_req_editor_on_change_ctx"] = {
    "editor_key": _req_editor_widget_key,
    "dept": active_dept,
    "period_pk": _period_pk,
    "crdf_key": _crdf_key,
}
st.session_state[_crdf_key] = _rq_sub[_period_pk]
st.data_editor(
    st.session_state[_crdf_key],
    column_config=col_config,
    use_container_width=True,
    height=_req_table_h,
    key=_req_editor_widget_key,
    num_rows="fixed",
    disabled=_req_editor_disabled,
    on_change=_on_request_schedule_editor_change,
)

if _req_editor_disabled:
    st.info(
        "🔐 위에서 해당 부서 **로그인**을 완료한 뒤에만 신청 표를 수정할 수 있습니다.",
        icon="🔒",
    )

st.caption(
    "편집 내용은 브라우저 세션에 바로 반영됩니다. **파일(JSON)에는** 아래 「💾 신청 근무 전체 저장」을 눌렀을 때만 기록됩니다."
)
_save_allowed = bool(_auth_ok)
_req_save_pad_l, _req_save_mid, _req_save_pad_r = st.columns([2, 2, 2])
with _req_save_mid:
    if st.button(
        "💾 신청 근무 전체 저장",
        type="primary",
        use_container_width=True,
        key=f"btn_save_all_requests_{active_dept}_{_period_pk}_g{gen}",
        help="현재 부서·연월의 신청 표를 hospital_config.json(해당 부서만)·브라우저·서버 백업에 반영합니다.",
        disabled=not _save_allowed,
    ):
        if not _save_allowed:
            st.warning("저장하려면 해당 부서 로그인이 필요합니다.")
        else:
            _sd_req = str(st.session_state.selected_dept).strip()
            _merged_save = _snapshot_request_editor_for_save(
                _rq_sub[_period_pk], _req_editor_widget_key, None
            )
            _ec_save = _normalize_req_shift_cells(_clean_req_df(_merged_save), _req_shift_allowed)
            _nurses_save_btn = _clean_nurse_names_list(
                list(st.session_state.departments.get(active_dept, nurses))
            )
            if len(_nurses_save_btn) != num_nurses or not _nurses_save_btn:
                _nurses_save_btn = list(nurses)
            _ec_save = _prepare_requests_df_for_current_table(
                _ec_save, _nurses_save_btn, req_col_labels
            )
            _persist_schedule_requests(
                _sd_req,
                _period_pk,
                st.session_state.sel_year,
                st.session_state.sel_month,
                _nurses_save_btn,
                req_col_labels,
                _ec_save,
            )
            _file_ok = _save_dept_schedule_requests_to_hospital_config(
                _sd_req,
                _period_pk,
                st.session_state.sel_year,
                st.session_state.sel_month,
                _nurses_save_btn,
                req_col_labels,
                _ec_save,
            )
            if _file_ok:
                _disk_df = _try_load_requests_from_hospital_config(
                    _sd_req, _period_pk, _nurses_save_btn, req_col_labels
                )
                _verify_ok = _disk_df is not None
                if _verify_ok:
                    _ec_disk = _prepare_requests_df_for_current_table(
                        _disk_df, _nurses_save_btn, req_col_labels
                    )
                    _rq_sub[_period_pk] = _ec_disk
                else:
                    _rq_sub[_period_pk] = _ec_save.copy()
                st.session_state.dept_requests[_sd_req] = _rq_sub
                st.session_state.pop(_req_editor_widget_key, None)
                st.session_state.pop("request_editor", None)
                st.toast(_REQ_SCHEDULE_BATCH_SAVE_TOAST)
                if not _verify_ok:
                    st.session_state["_req_save_feedback"] = {
                        "dept": _sd_req,
                        "disk_verify": False,
                    }
                _carry_all_raw = (
                    st.session_state.get(
                        _carry_widget_session_key(
                            _sd_req,
                            int(st.session_state.sel_year),
                            int(st.session_state.sel_month),
                        ),
                        "",
                    )
                    or ""
                ).strip()
                _carry_all_ok, _carry_all_msg = _persist_department_last_month_to_hospital_config(
                    _sd_req,
                    int(st.session_state.sel_year),
                    int(st.session_state.sel_month),
                    _carry_all_raw,
                    _nurses_save_btn,
                )
                if not _carry_all_ok:
                    st.session_state["_carry_persist_warning"] = _carry_all_msg
                st.rerun()
            else:
                st.warning("hospital_config.json에 반영되지 않았습니다. 파일 권한·경로를 확인해 주세요.")

# 근무표 생성: data_editor 직후 처리 (파일 하단까지 가지 않아 미적용·예외 누락 방지) — 관리자만
if _can_manage_dept and st.session_state.pop("_pending_schedule_generate", False):
    try:
        holidays = _parse_holidays(st.session_state.dept_holidays.get(active_dept, ""))
        days = get_april_days(holidays)
        req_df_gen = _normalize_req_shift_cells(
            _clean_req_df(
                _snapshot_request_editor_for_save(
                    _rq_sub[_period_pk], _req_editor_widget_key, None
                )
            ),
            _req_shift_allowed,
        )
        requests_gen = _df_to_requests(req_df_gen, days, nurses)
        _fp_idx = _fp_pairs_to_indices(
            nurses,
            st.session_state.get("dept_forbidden_pairs", {}).get(active_dept, []),
        )
        _pg_raw = st.session_state.get("dept_pregnant", {}).get(active_dept, [])
        _pg_for_solver = _pg_raw if isinstance(_pg_raw, list) and _pg_raw else None
        _n4_raw = st.session_state.get("dept_n_max4", {}).get(active_dept, [])
        _n4_for_solver = _n4_raw if isinstance(_n4_raw, list) and _n4_raw else None
        _carry_raw = (
            st.session_state.get(
                _carry_widget_session_key(
                    active_dept,
                    int(st.session_state.sel_year),
                    int(st.session_state.sel_month),
                ),
                "",
            )
            or ""
        )
        _carry_parse_gen = _parse_carry_in_text(_carry_raw, nurses)
        if _carry_parse_gen is False and _carry_raw.strip():
            st.warning(
                "전월 이월 JSON 형식이 올바르지 않습니다. 근무표 생성은 계속하며, "
                "검증의 이월 반영은 hospital_config의 저장본만 사용합니다. "
                "「💾 이월 근무 데이터 저장」 전에 JSON을 고쳐 주세요."
            )
        _carry_for_validate = _merge_carry_with_hospital_last_month(
            None if _carry_parse_gen is False else _carry_parse_gen,
            _load_hospital_config_bundle(),
            active_dept,
            int(st.session_state.sel_year),
            int(st.session_state.sel_month),
            nurses,
        )
        if _carry_for_validate is False:
            _carry_for_validate = None
        _sched_ex = st.session_state.dept_schedules.get(active_dept, {})
        _regen = isinstance(_sched_ex, dict) and bool(_sched_ex.get(_period_pk))
        _prev_sched_for_regen = None
        _regen_fix_cells = None
        if _regen:
            _e = _sched_ex.get(_period_pk)
            if isinstance(_e, dict):
                _prev_sched_for_regen = _e.get("schedule")
            # 인력(D/E/N) 오류도 해당 일 전체 셀이 포함되어, 재생성 시 이전 근무를 풀고 하드 인력을 맞추기 쉬움.
            _regen_fix_cells = error_cells_from_validation_issues(
                st.session_state.get("violations")
            )
            st.session_state["_schedule_regen_ctr"] = int(
                st.session_state.get("_schedule_regen_ctr", 0)
            ) + 1
        # 재생성마다 솔버 시드 분기(동일 신청이라도 다른 근무 패턴 탐색). 최초 생성도 시드 고정으로 재현성 완화.
        _ctr = int(st.session_state.get("_schedule_regen_ctr", 0))
        _seed = (
            (_ctr * 1_000_003)
            ^ hash(_period_pk)
            ^ hash(active_dept)
            ^ (int(st.session_state.sel_year) * 13)
            ^ (int(st.session_state.sel_month) * 97)
        ) & 0x7FFFFFFF
        with st.spinner(
            "⏳ 근무표를 다시 짜는 중입니다… (신청·인원 우선·탐색·신청 유지)"
            if _regen
            else "⏳ 근무표를 계산하는 중입니다… (인원·신청 우선·최대 약 1~1.5분 탐색, 시간 내 최선 해 표시)"
        ):
            # 솔버·검증 모두 carry(위젯·파일 병합) 사용; INFEASIBLE 시 schedule=None.
            _sol = solve_schedule(
                num_nurses,
                requests_gen,
                holidays,
                forbidden_pairs=_fp_idx or None,
                carry_in=_carry_for_validate,
                regenerate=_regen,
                rng_seed=_seed,
                nurse_names=nurses,
                carry_next_month=None,
                pregnant_nurses=_pg_for_solver,
                n_max4_nurses=_n4_for_solver,
                unit_profile=_effective_unit_profile(active_dept),
                previous_schedule=_prev_sched_for_regen if _regen else None,
                regeneration_fix_cells=_regen_fix_cells if _regen else None,
            )
            schedule = _sol[0]
            success = _sol[1]
            status = _sol[2]
            issues = list(_sol[3]) if len(_sol) > 3 and _sol[3] else []
        # schedule 이 있으면 저장·표시. success=False 는 솔버 폴백(UNKNOWN 등)일 때.
        if schedule is not None:
            _rq_sub[_period_pk] = req_df_gen
            _req_cols_gen = [_day_label_compact(d) for d in days]
            _persist_schedule_requests(
                active_dept,
                _period_pk,
                st.session_state.sel_year,
                st.session_state.sel_month,
                nurses,
                _req_cols_gen,
                req_df_gen,
            )
            st.toast("✅ 저장이 완료되었습니다", icon="✅")
            st.session_state.dept_schedules.setdefault(active_dept, {})[_period_pk] = {
                "schedule": schedule,
                "nurse_names": nurses.copy(),
                "holidays": holidays,
                "requests": requests_gen,
            }
            _archive_put_month(
                active_dept,
                st.session_state.sel_year,
                st.session_state.sel_month,
                nurses,
                schedule,
            )
            st.session_state.violations = issues
            st.session_state.show_violations = False
            if not success:
                st.warning(
                    f"⚠️ CP-SAT가 제한 시간 안에 완전한 해를 못 찾아 임시 초안을 채웠습니다. "
                    f"{status}"
                )
            if not issues:
                st.toast("✅ 근무표 생성 완료! 모든 규칙 통과", icon="🎉")
            else:
                errors = sum(1 for v in issues if v["level"] == "error")
                warns = sum(1 for v in issues if v["level"] == "warn")
                st.toast(
                    f"✅ 근무표를 표시했습니다 (검토 {errors}오류·{warns}경고). "
                    "상단 접힌 목록 또는 ⚠️ 버튼·수정 모드에서 수기로 다듬을 수 있습니다.",
                    icon="📋",
                )
            st.rerun()
        elif schedule is None:
            st.error(
                f"❌ 근무표 생성 불가: {status}\n\n"
                "N 절대 규칙·이월·일일 N 인원을 동시에 만족할 수 없을 수 있습니다. "
                "이월·신청·고정 휴가를 완화한 뒤 다시 시도해 주세요."
            )
            if issues:
                for _iss in issues:
                    if str(_iss.get("level")) == "error" and _iss.get("msg"):
                        st.caption(str(_iss["msg"]))
    except Exception as e:
        st.error("근무표 생성 중 오류가 발생했습니다. 아래 내용을 확인해 주세요.")
        st.exception(e)

st.markdown(
    '<p style="margin:4px 0 14px 0;font-size:14px;font-weight:700;color:#1A237E;'
    'line-height:1.35;position:relative;z-index:2;">👁️ 신청 근무 미리보기</p>',
    unsafe_allow_html=True,
)
_show_schedule_preview_iframe(
    _render_requests_preview_html(
        _clean_req_df(st.session_state[_crdf_key]), nurses, days
    ),
    num_nurses,
    extra_rows=2,
)

# 저장 영역 (전체 너비 — 좁은 열에 넣으면 버튼이 안 보이는 경우가 있음)
with st.container(border=True):
    _rs_head_l, _rs_head_r = st.columns([4, 1])
    with _rs_head_l:
        st.markdown(
            '<div class="req-save-panel">'
            '<h4 style="margin:0 0 8px 0;font-size:1.1rem;color:#111111;font-weight:700;">💾 신청 근무 저장</h4>'
            '<p style="margin:0 0 12px 0;font-size:13px;color:#222222;line-height:1.5;">'
            "신청 표는 위 <strong>💾 신청 근무 전체 저장</strong>을 눌렀을 때 "
            "<code>hospital_config.json</code>(해당 부서)·브라우저·<code>schedule_requests.json</code>에 기록됩니다. "
            "편집만 하고 저장하지 않으면 새로고침·세션 만료 시 내용이 사라질 수 있습니다. "
            "기존 <code>schedule_requests.json</code> 데이터는 최초 1회 localStorage로 옮겨질 수 있습니다. "
            "Google Sheets 연동은 추후 설정 시 추가 가능합니다.<br/>"
            + (
                '<strong>🗓️ 생성</strong>은 항상 위 표의 <strong>현재 내용</strong>을 사용합니다. '
                "(근무표 자동 생성은 위에서 <strong>부서 로그인</strong> 후 가능합니다.)</p></div>"
                if not _can_manage_dept
                else '<strong>🗓️ 생성</strong>은 항상 위 표의 <strong>현재 내용</strong>을 사용합니다.</p></div>'
            ),
            unsafe_allow_html=True,
        )
    with _rs_head_r:
        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
        if st.button(
            "🔄 이전 기록 불러오기",
            use_container_width=True,
            key=f"btn_ls_reload_{active_dept}_{_period_pk}_g{gen}",
            help=(
                "1) shift_requests.json 의 키 «부서_연도_월» (예: 본관5병동_2026_5) 스냅샷. "
                "2) 없으면 hospital_config → schedule_requests.json(또는 브라우저 백업). "
                "명단·열이 스냅샷과 일치할 때만 표에 반영됩니다."
            ),
        ):
            st.session_state["_force_ls_reload"] = True
            st.rerun()

    st.markdown(
        '<div class="req-save-status req-save-ok" style="background:#E8F5E9;border:1px solid #A5D6A7;'
        'border-radius:8px;padding:10px 14px;color:#111111;font-size:14px;margin:8px 0;line-height:1.45;">'
        "✅ 저장 버튼으로 확정한 뒤에는 파일·백업에서 불러올 수 있습니다.</div>",
        unsafe_allow_html=True,
    )

    if _LocalStorageCls is None:
        st.caption(
            "ℹ️ 패키지 `streamlit-local-storage`가 없어 서버 JSON만 사용 중입니다. "
            "Cloud에서도 유지하려면 `pip install streamlit-local-storage` 후 재실행하세요."
        )

    c_clear, _ = st.columns([1, 3])
    with c_clear:
        if st.button(
            "🗑️ 신청 전체 지우기",
            use_container_width=True,
            key=f"btn_clear_requests_{active_dept}_g{gen}",
        ):
            _delete_schedule_requests_period(active_dept, _period_pk)
            _rq_sub[_period_pk] = _make_requests_df(nurses, days)
            st.rerun()

# 테마·위젯 CSS보다 나중에 적용 — text_input 글자색 최종 고정(검정)
st.markdown(
    """
    <style>
    .stApp, section[data-testid="stSidebar"] { color-scheme: light !important; }
    section[data-testid="stSidebar"] > div:first-child { background: #ffffff !important; }
    section[data-testid="stMain"] [data-testid="stSelectbox"] [data-baseweb="select"],
    section[data-testid="stMain"] [data-testid="stMultiSelect"] [data-baseweb="select"],
    section[data-testid="stSidebar"] [data-testid="stSelectbox"] [data-baseweb="select"],
    section[data-testid="stSidebar"] [data-testid="stMultiSelect"] [data-baseweb="select"] {
        background-color: #ffffff !important;
        border: 1px solid #d1d5db !important;
        border-radius: 8px !important;
        box-sizing: border-box !important;
    }
    section[data-testid="stMain"] [data-testid="stSelectbox"] [data-baseweb="select"]:focus-within,
    section[data-testid="stMain"] [data-testid="stMultiSelect"] [data-baseweb="select"]:focus-within,
    section[data-testid="stSidebar"] [data-testid="stSelectbox"] [data-baseweb="select"]:focus-within,
    section[data-testid="stSidebar"] [data-testid="stMultiSelect"] [data-baseweb="select"]:focus-within {
        border-color: #4f46e5 !important;
        box-shadow: 0 0 0 1px #4f46e5 !important;
    }
    section[data-testid="stMain"] [data-testid="stSelectbox"] [data-baseweb="select"] > div,
    section[data-testid="stMain"] [data-testid="stMultiSelect"] [data-baseweb="select"] > div,
    section[data-testid="stSidebar"] [data-testid="stSelectbox"] [data-baseweb="select"] > div,
    section[data-testid="stSidebar"] [data-testid="stMultiSelect"] [data-baseweb="select"] > div {
        border: none !important;
        box-shadow: none !important;
    }
    .stTextInput input,
    section[data-testid="stSidebar"] [data-baseweb="input"] input {
        color: #000000 !important;
        -webkit-text-fill-color: #000000 !important;
        caret-color: #000000 !important;
        background-color: #ffffff !important;
    }
    section[data-testid="stSidebar"] [data-testid="stSelectbox"] [data-baseweb="select"] > div {
        background: #ffffff !important;
        box-shadow: none !important;
        border: none !important;
        display: flex !important;
        align-items: center !important;
        min-height: 54px !important;
        height: auto !important;
        max-height: none !important;
        padding: 0.55rem 2.5rem 0.55rem 0.75rem !important;
        box-sizing: border-box !important;
        overflow: visible !important;
    }
    section[data-testid="stSidebar"] [data-testid="stSelectbox"] [data-baseweb="select"] [role="combobox"],
    section[data-testid="stSidebar"] [data-testid="stSelectbox"] [data-baseweb="select"] p {
        color: #111111 !important;
        -webkit-text-fill-color: #111111 !important;
        font-weight: 600 !important;
        font-size: 16px !important;
        line-height: 1.4 !important;
        overflow: visible !important;
    }
    /* 메인 — 연도·월·부서 select 표시 글자 검정 (테마 덮어쓰기) */
    section[data-testid="stMain"] [data-testid="stSelectbox"] [data-baseweb="select"] > div {
        display: flex !important;
        align-items: center !important;
        min-height: 54px !important;
        padding: 0.55rem 2.5rem 0.55rem 0.75rem !important;
        overflow: visible !important;
    }
    section[data-testid="stMain"] [data-testid="stSelectbox"] [data-baseweb="select"] [role="combobox"],
    section[data-testid="stMain"] [data-testid="stSelectbox"] [data-baseweb="select"] [role="combobox"] p,
    section[data-testid="stMain"] [data-testid="stSelectbox"] [data-baseweb="select"] [role="combobox"] span {
        color: #111111 !important;
        -webkit-text-fill-color: #111111 !important;
        opacity: 1 !important;
        overflow: visible !important;
        font-size: 16px !important;
        line-height: 1.4 !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)
